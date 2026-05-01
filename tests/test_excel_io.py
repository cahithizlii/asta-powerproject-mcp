"""Test excel_io pure-Python helpers (CLAUDE.md RULE 14 brand)."""
import pytest
from openpyxl import Workbook
from excel_io import (
    BRAND_LACIVERT, BRAND_NAVY, BRAND_TURQUOISE, BRAND_LABEL_GRAY,
    ZEBRA, RAG_GREEN, RAG_AMBER, RAG_RED,
    apply_header_style, apply_rag_fill, apply_zebra_fill,
    build_tasks_sheet, read_tasks_sheet, build_summary_sheet,
    build_evm_sheet, build_dcma_sheet,
)


def test_brand_constants_exist():
    assert BRAND_LACIVERT == "FF0B1F4D"
    assert BRAND_NAVY == "FF3D4663"
    assert BRAND_TURQUOISE == "FF39B4CC"
    assert ZEBRA == "FFF0F3F8"


def test_rag_colors_distinct():
    assert RAG_GREEN.startswith("FF")
    assert RAG_AMBER.startswith("FF")
    assert RAG_RED.startswith("FF")
    assert len({RAG_GREEN, RAG_AMBER, RAG_RED}) == 3


def test_apply_header_style():
    wb = Workbook()
    ws = wb.active
    ws.append(["A", "B", "C"])
    apply_header_style(ws, row=1)
    cell = ws.cell(row=1, column=1)
    assert cell.font.bold is True
    assert cell.fill.start_color.value == BRAND_LACIVERT
    assert cell.font.color.value == "FFFFFFFF"
    assert cell.font.name == "Calibri"
    assert cell.font.size == 11


def test_apply_header_style_all_columns():
    wb = Workbook()
    ws = wb.active
    ws.append(["A", "B", "C", "D"])
    apply_header_style(ws, row=1)
    for col in range(1, 5):
        assert ws.cell(row=1, column=col).font.bold is True


def test_apply_rag_fill_pass():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="OK")
    apply_rag_fill(ws.cell(row=1, column=1), status="pass")
    assert ws.cell(row=1, column=1).fill.start_color.value == RAG_GREEN


def test_apply_rag_fill_fail():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="FAIL")
    apply_rag_fill(ws.cell(row=1, column=1), status="fail")
    assert ws.cell(row=1, column=1).fill.start_color.value == RAG_RED


def test_apply_rag_fill_amber():
    wb = Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="WARN")
    apply_rag_fill(cell, status="amber")
    assert cell.fill.start_color.value == RAG_AMBER


def test_apply_rag_fill_unknown_status_no_fill():
    """Unknown status string -> no fill applied (defensive)."""
    wb = Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="x")
    apply_rag_fill(cell, status="unknown_status")
    assert cell.fill.fill_type in (None, "none")


def test_apply_zebra_fill_even_row():
    wb = Workbook()
    ws = wb.active
    cell = ws.cell(row=2, column=1, value="x")
    apply_zebra_fill(cell, row_idx=2)
    assert cell.fill.start_color.value == ZEBRA


def test_apply_zebra_fill_odd_row_no_fill():
    wb = Workbook()
    ws = wb.active
    cell = ws.cell(row=3, column=1, value="x")
    apply_zebra_fill(cell, row_idx=3)
    assert cell.fill.fill_type in (None, "none")


# ---------- T95: Tasks + Summary sheet builders + readers ----------

def _sample_tasks():
    return [
        {"id": 1, "name": "Foundation", "duration_h": 80, "start": "2026-01-01",
         "finish": "2026-01-15", "percent_complete": 100, "critical": True,
         "summary": False},
        {"id": 2, "name": "Frame", "duration_h": 160, "start": "2026-01-16",
         "finish": "2026-02-15", "percent_complete": 50, "critical": False,
         "summary": False},
    ]


def test_build_tasks_sheet_creates_sheet():
    wb = Workbook()
    build_tasks_sheet(wb, _sample_tasks(), sheet_name="Tasks")
    assert "Tasks" in wb.sheetnames


def test_build_tasks_sheet_header_row():
    wb = Workbook()
    build_tasks_sheet(wb, _sample_tasks())
    ws = wb["Tasks"]
    headers = [c.value for c in ws[1]]
    assert "ID" in headers
    assert "Name" in headers
    assert "Duration (d)" in headers
    assert "%Complete" in headers


def test_build_tasks_sheet_data_rows():
    wb = Workbook()
    build_tasks_sheet(wb, _sample_tasks())
    ws = wb["Tasks"]
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == "Foundation"
    assert ws.cell(row=2, column=3).value == 10.0  # 80h / 8 = 10d


def test_build_tasks_sheet_excludes_summary():
    tasks = _sample_tasks() + [{"id": 0, "name": "Sum", "summary": True,
                                "duration_h": 0, "start": None, "finish": None,
                                "percent_complete": 0, "critical": False}]
    wb = Workbook()
    build_tasks_sheet(wb, tasks)
    ws = wb["Tasks"]
    assert ws.max_row == 3  # header + 2 real rows


def test_build_tasks_sheet_header_styled_lacivert():
    wb = Workbook()
    build_tasks_sheet(wb, _sample_tasks())
    ws = wb["Tasks"]
    assert ws.cell(row=1, column=1).fill.start_color.value == BRAND_LACIVERT


def test_build_tasks_sheet_pct_format():
    wb = Workbook()
    build_tasks_sheet(wb, _sample_tasks())
    ws = wb["Tasks"]
    assert ws.cell(row=2, column=6).number_format == "0%"


def test_read_tasks_sheet_round_trip(tmp_path):
    wb = Workbook()
    build_tasks_sheet(wb, _sample_tasks())
    xlsx = tmp_path / "round.xlsx"
    wb.save(str(xlsx))
    rows = read_tasks_sheet(str(xlsx), sheet_name="Tasks")
    assert len(rows) == 2
    assert rows[0]["name"] == "Foundation"
    assert rows[0]["duration_h"] == 80


def test_read_tasks_sheet_missing_sheet_returns_empty(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Other"
    xlsx = tmp_path / "x.xlsx"
    wb.save(str(xlsx))
    rows = read_tasks_sheet(str(xlsx), sheet_name="Tasks")
    assert rows == []


def test_build_summary_sheet():
    wb = Workbook()
    summary = {"BAC": 1000.0, "EAC": 1100.0, "SPI": 0.95, "CPI": 0.91,
               "rag": "amber", "executive_text": "Project on track."}
    build_summary_sheet(wb, summary)
    assert "Summary" in wb.sheetnames
    ws = wb["Summary"]
    assert "Summary" in str(ws.cell(row=1, column=1).value)


def test_build_summary_sheet_rag_colored():
    wb = Workbook()
    summary = {"rag": "red"}
    build_summary_sheet(wb, summary)
    ws = wb["Summary"]
    # Find RAG row (label = "Overall RAG")
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Overall RAG":
            assert ws.cell(row=r, column=2).fill.start_color.value == RAG_RED
            return
    pytest.fail("Overall RAG row not found")


def test_build_summary_sheet_handles_missing_metrics():
    wb = Workbook()
    build_summary_sheet(wb, {})
    ws = wb["Summary"]
    assert ws.cell(row=3, column=2).value == "N/A"


# ---------- T96: EVM sheet builder ----------

def _sample_evm():
    return {
        "metrics": {"BAC": 1000.0, "EV": 600.0, "AC": 650.0, "PV": 700.0,
                    "SV": -100.0, "CV": -50.0, "SPI": 0.857, "CPI": 0.923},
        "forecast": {"EAC1": 1050.0, "EAC2": 1083.5, "EAC3": 1100.0,
                     "ETC": 400.0, "VAC": -50.0,
                     "TCPI_BAC": 1.14, "TCPI_EAC": 1.05},
        "earned_schedule": {"AT": 5.0, "ES": 4.5, "SVt": -0.5, "SPIt": 0.9},
        "rag": "amber",
        "time_phased": [
            {"period": "2026-W01", "PV": 100, "EV": 80, "AC": 90,
             "cum_PV": 100, "cum_EV": 80, "cum_AC": 90},
            {"period": "2026-W02", "PV": 200, "EV": 180, "AC": 200,
             "cum_PV": 300, "cum_EV": 260, "cum_AC": 290},
        ],
    }


def test_build_evm_sheet_creates_two_sheets():
    wb = Workbook()
    build_evm_sheet(wb, _sample_evm())
    assert "EVM_Compute" in wb.sheetnames
    assert "EVM_TimePhased" in wb.sheetnames


def test_evm_compute_sheet_has_BAC_row():
    wb = Workbook()
    build_evm_sheet(wb, _sample_evm())
    ws = wb["EVM_Compute"]
    bac_found = any(row[0] == "BAC" and row[1] == 1000.0
                    for row in ws.iter_rows(values_only=True))
    assert bac_found


def test_evm_compute_sheet_rag_colored():
    wb = Workbook()
    build_evm_sheet(wb, _sample_evm())
    ws = wb["EVM_Compute"]
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Overall RAG":
            assert ws.cell(row=r, column=2).fill.start_color.value == RAG_AMBER
            return
    pytest.fail("Overall RAG row not found")


def test_evm_time_phased_sheet_has_data():
    wb = Workbook()
    build_evm_sheet(wb, _sample_evm())
    ws = wb["EVM_TimePhased"]
    # header + 2 data rows
    assert ws.max_row == 3
    assert ws.cell(row=2, column=1).value == "2026-W01"


def test_evm_handles_missing_time_phased():
    wb = Workbook()
    evm = {**_sample_evm(), "time_phased": []}
    build_evm_sheet(wb, evm)
    ws = wb["EVM_TimePhased"]
    # header only
    assert ws.max_row == 1


def test_evm_handles_missing_forecast():
    wb = Workbook()
    evm = {"metrics": {"BAC": 100, "EV": 50, "AC": 60, "PV": 80,
                       "SPI": 0.625, "CPI": 0.83}}
    build_evm_sheet(wb, evm)
    assert "EVM_Compute" in wb.sheetnames
    ws = wb["EVM_Compute"]
    # EAC1 row should have N/A
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "EAC1":
            assert ws.cell(row=r, column=2).value == "N/A"
            return
    pytest.fail("EAC1 row not found")


# ---------- T97: DCMA sheet builder ----------

def _sample_dcma():
    return {
        "rules": [
            {"id": 1, "name": "No Predecessor", "threshold": "<5%", "actual": 3.5,
             "actual_unit": "%", "status": "pass", "failed_count": 7,
             "total_count": 200, "failed_task_ids": []},
            {"id": 9, "name": "High Duration (>44d)", "threshold": "<5%",
             "actual": 7.5, "actual_unit": "%", "status": "fail",
             "failed_count": 15, "total_count": 200, "failed_task_ids": [186, 187]},
        ],
        "summary": {"pass_count": 13, "fail_count": 1, "overall_rag": "green",
                    "executive_text": "13/14 pass"},
        "drilldowns": {
            9: [{"id": 186, "name": "H00"}, {"id": 187, "name": "H01"}],
        },
    }


def test_build_dcma_sheet_creates_two_sheets():
    wb = Workbook()
    build_dcma_sheet(wb, _sample_dcma())
    assert "DCMA_Rules" in wb.sheetnames
    assert "DCMA_Failed" in wb.sheetnames


def test_dcma_rules_sheet_lists_all_rules():
    wb = Workbook()
    build_dcma_sheet(wb, _sample_dcma())
    ws = wb["DCMA_Rules"]
    assert ws.max_row == 3  # header + 2 rules


def test_dcma_rules_status_color():
    wb = Workbook()
    build_dcma_sheet(wb, _sample_dcma())
    ws = wb["DCMA_Rules"]
    headers = [c.value for c in ws[1]]
    status_col = headers.index("Status") + 1
    # Row 2 is rule 1 (pass) -> green
    assert ws.cell(row=2, column=status_col).fill.start_color.value == RAG_GREEN
    # Row 3 is rule 9 (fail) -> red
    assert ws.cell(row=3, column=status_col).fill.start_color.value == RAG_RED


def test_dcma_failed_sheet_lists_drilldowns():
    wb = Workbook()
    build_dcma_sheet(wb, _sample_dcma())
    ws = wb["DCMA_Failed"]
    assert ws.max_row >= 2
    headers = [c.value for c in ws[1]]
    tid_col = headers.index("Task ID") + 1
    task_ids = [ws.cell(row=r, column=tid_col).value
                for r in range(2, ws.max_row + 1)]
    assert 186 in task_ids
    assert 187 in task_ids


def test_dcma_failed_sheet_empty_when_no_drilldowns():
    wb = Workbook()
    build_dcma_sheet(wb, {"rules": [], "summary": {}, "drilldowns": {}})
    ws = wb["DCMA_Failed"]
    assert ws.max_row == 1  # header only


def test_dcma_failed_sheet_caps_at_10_per_rule():
    """Drilldown limited to 10 tasks per rule (avoid huge sheets)."""
    drilldowns = {
        9: [{"id": i, "name": f"T{i}"} for i in range(1, 21)],  # 20 tasks
    }
    dcma = {"rules": [{"id": 9, "name": "High Duration", "threshold": "<5%",
                       "actual": 100, "actual_unit": "%", "status": "fail",
                       "failed_count": 20, "total_count": 20}],
            "summary": {}, "drilldowns": drilldowns}
    wb = Workbook()
    build_dcma_sheet(wb, dcma)
    ws = wb["DCMA_Failed"]
    # header + 10 rows max
    assert ws.max_row == 11
