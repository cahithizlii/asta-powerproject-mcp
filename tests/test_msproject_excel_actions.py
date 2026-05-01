"""Test Phase 5c T99 export action helpers."""
import os, sys
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from msproject_mcp_core import (
    _msp_excel_export_hakedis, _msp_excel_export_tasks,
    _msp_excel_export_evm, _msp_excel_export_dcma,
)
from openpyxl import load_workbook

MSP_XML = os.path.join(os.path.dirname(__file__), "fixtures", "sample_msp.xml")


def test_export_hakedis_creates_workbook(tmp_path):
    xlsx = tmp_path / "hak.xlsx"
    r = _msp_excel_export_hakedis(file_path=MSP_XML, xlsx_path=str(xlsx))
    assert r["status"] == "ok"
    assert xlsx.exists()
    wb = load_workbook(str(xlsx), read_only=True)
    for s in ("Summary", "Tasks", "EVM_Compute", "EVM_TimePhased",
              "DCMA_Rules", "DCMA_Failed"):
        assert s in wb.sheetnames


def test_export_hakedis_missing_xlsx_path():
    r = _msp_excel_export_hakedis(file_path=MSP_XML)
    assert r["status"] == "error"
    assert "xlsx_path" in r["error"]


def test_export_tasks_creates_single_sheet(tmp_path):
    xlsx = tmp_path / "tasks.xlsx"
    r = _msp_excel_export_tasks(file_path=MSP_XML, xlsx_path=str(xlsx))
    assert r["status"] == "ok"
    wb = load_workbook(str(xlsx), read_only=True)
    assert "Tasks" in wb.sheetnames


def test_export_evm_creates_two_sheets(tmp_path):
    xlsx = tmp_path / "evm.xlsx"
    r = _msp_excel_export_evm(file_path=MSP_XML, xlsx_path=str(xlsx))
    assert r["status"] == "ok"
    wb = load_workbook(str(xlsx), read_only=True)
    assert "EVM_Compute" in wb.sheetnames
    assert "EVM_TimePhased" in wb.sheetnames


def test_export_dcma_creates_rules_sheet(tmp_path):
    xlsx = tmp_path / "dcma.xlsx"
    r = _msp_excel_export_dcma(file_path=MSP_XML, xlsx_path=str(xlsx))
    assert r["status"] == "ok"
    wb = load_workbook(str(xlsx), read_only=True)
    assert "DCMA_Rules" in wb.sheetnames


def test_export_evm_invalid_bucket(tmp_path):
    """bucket must be day/week/month."""
    xlsx = tmp_path / "evm.xlsx"
    r = _msp_excel_export_evm(file_path=MSP_XML, xlsx_path=str(xlsx),
                              bucket="quarterly")
    assert r["status"] == "error"


from openpyxl import Workbook
from msproject_mcp_core import (
    _msp_excel_import_tasks, _msp_excel_import_progress,
)


def test_import_tasks_round_trip(tmp_path):
    """Build a tasks xlsx, import it. Status must be ok or error (not crash)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.append(["ID", "Name", "Duration (d)", "Start", "Finish",
               "%Complete", "Critical", "Summary"])
    ws.append([1, "TestImport1", 5, None, None, 0, False, False])
    ws.append([2, "TestImport2", 3, None, None, 0, False, False])
    xlsx = tmp_path / "imp.xlsx"
    wb.save(str(xlsx))
    r = _msp_excel_import_tasks(xlsx_path=str(xlsx))
    # MSP COM round-trip - in unit tests MSP may or may not be running.
    # Assert call shape (not real task ids since MSP state isn't deterministic).
    assert r["status"] in ("ok", "error")
    if r["status"] == "ok":
        assert r["rows_imported"] == 2


def test_import_tasks_missing_file():
    r = _msp_excel_import_tasks(xlsx_path="/definitely/nonexistent.xlsx")
    assert r["status"] == "error"
    assert "not found" in r["error"].lower()


def test_import_tasks_no_xlsx_path():
    r = _msp_excel_import_tasks()
    assert r["status"] == "error"
    assert "xlsx_path" in r["error"]


def test_import_tasks_empty_sheet(tmp_path):
    """Empty Tasks sheet (header only) -> rows_imported=0, status ok."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.append(["ID", "Name", "Duration (d)"])
    xlsx = tmp_path / "empty.xlsx"
    wb.save(str(xlsx))
    r = _msp_excel_import_tasks(xlsx_path=str(xlsx))
    assert r["status"] == "ok"
    assert r["rows_imported"] == 0
    assert r.get("task_ids", []) == []


def test_import_progress_round_trip(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Progress"
    ws.append(["Task ID", "%Complete"])
    ws.append([1, 50])
    xlsx = tmp_path / "prog.xlsx"
    wb.save(str(xlsx))
    r = _msp_excel_import_progress(xlsx_path=str(xlsx))
    assert r["status"] in ("ok", "error")


def test_import_progress_missing_file():
    r = _msp_excel_import_progress(xlsx_path="/no/such.xlsx")
    assert r["status"] == "error"


def test_import_progress_no_xlsx_path():
    r = _msp_excel_import_progress()
    assert r["status"] == "error"


def test_import_progress_empty_sheet(tmp_path):
    """Empty Progress sheet -> rows_imported=0, status ok."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Progress"
    ws.append(["Task ID", "%Complete"])
    xlsx = tmp_path / "empty_prog.xlsx"
    wb.save(str(xlsx))
    r = _msp_excel_import_progress(xlsx_path=str(xlsx))
    assert r["status"] == "ok"
    assert r["rows_imported"] == 0
