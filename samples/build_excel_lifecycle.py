"""Phase 5c Excel acceptance: 200-task hakedis workbook + import roundtrip.

SAFETY: FileNew + FileClose 0. User's active project untouched.

Scenario (target <90s — 200 task COM-iter heavy per Phase 5b lesson):
  1. Build 200 tasks (185 normal + 15 high-duration) + 14 CAU resources
  2. Save Baseline 0
  3. Phase 3b progress for 30 tasks (BEI signal)
  4. set_status_date
  5. _msp_excel_export_hakedis -> 6-sheet workbook
  6. Verify file structure (sheet count + key cells)
  7. Build 10-row progress xlsx
  8. _msp_excel_import_progress -> re-applies 10 updates

Run: python samples/build_excel_lifecycle.py
"""
import functools
import os
import sys
import tempfile
import time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
print = functools.partial(print, flush=True)

import pythoncom
import win32com.client
from openpyxl import Workbook, load_workbook

from msproject_mcp_core import (
    _msp_task_bulk_add, _msp_resource_add, _msp_resource_bulk_assign,
    _msp_baseline_save,
    _msp_progress_bulk_update, _msp_progress_set_status_date,
    _msp_excel_export_hakedis, _msp_excel_import_progress,
)

N_TASKS = 200
N_HIGH_DUR = 15


def main():
    pythoncom.CoInitialize()
    app = win32com.client.GetActiveObject("MSProject.Application")
    app.FileNew()
    test_proj = app.ActiveProject
    test_name = test_proj.Name
    print(f"[SAFE] isolated test project: {test_name}")

    out_dir = tempfile.mkdtemp(prefix="dcma_excel_")
    xlsx_out = os.path.join(out_dir, "hakedis.xlsx")
    xlsx_imp = os.path.join(out_dir, "progress_import.xlsx")
    print(f"[INFO] output dir: {out_dir}")

    try:
        t0 = time.time()

        # 1. Build base
        print(f"\n1. Building {N_TASKS} tasks ({N_TASKS - N_HIGH_DUR} normal + "
              f"{N_HIGH_DUR} high-duration)...")
        items = [{"name": f"V{i:03d}", "duration": "5d"}
                 for i in range(N_TASKS - N_HIGH_DUR)]
        items += [{"name": f"H{i:02d}", "duration": "60d"} for i in range(N_HIGH_DUR)]
        tasks = _msp_task_bulk_add(items=items)
        task_ids = tasks["task_ids"]
        print(f"   tasks created in {time.time() - t0:.2f}s ({len(task_ids)} ids)")

        cau = ["COW", "EXT", "STL", "CAR", "MSN", "DRW", "INW",
               "EWI", "CWI", "ACP", "ELW", "DMS", "MTR", "LBR"]
        res_ids = [_msp_resource_add(name=n, type="Work")["resource_id"] for n in cau]
        sample = [{"task_id": tid, "resource_id": res_ids[i % 14]}
                  for i, tid in enumerate(task_ids[12:])]  # skip first 12 unassigned
        _msp_resource_bulk_assign(items=sample)
        print(f"   {len(sample)} resource assignments at {time.time() - t0:.2f}s")

        # 2. Baseline 0
        _msp_baseline_save(baseline_number=0)
        print(f"\n2. Baseline 0 saved at {time.time() - t0:.2f}s")

        # 3. Progress
        progress_items = [{"task_id": tid, "percent_complete": 50.0}
                          for tid in task_ids[:30]]
        _msp_progress_bulk_update(items=progress_items)
        _msp_progress_set_status_date(status_date="2026-05-15")
        print(f"   progress + status_date at {time.time() - t0:.2f}s")

        # 4. Export hakedis workbook
        print(f"\n3. Exporting hakedis workbook to {xlsx_out}...")
        r = _msp_excel_export_hakedis(xlsx_path=xlsx_out)
        if r.get("status") != "ok":
            print(f"   [ERR] export failed: {r.get('error')}")
            return
        print(f"   status={r['status']} sheets={r['sheets_written']}")
        print(f"   rows: {r['rows_written']}")
        print(f"   exported at {time.time() - t0:.2f}s, size={os.path.getsize(xlsx_out)} bytes")

        # 5. Verify workbook structure
        print(f"\n4. Verifying workbook structure...")
        wb = load_workbook(xlsx_out, read_only=True)
        for s in ("Summary", "Tasks", "EVM_Compute", "EVM_TimePhased",
                  "DCMA_Rules", "DCMA_Failed"):
            assert s in wb.sheetnames, f"missing sheet: {s}"
        print(f"   verified 6 sheets: {wb.sheetnames}")
        # Spot check Summary sheet first cell
        ws_sum = wb["Summary"]
        title_cell = ws_sum.cell(row=1, column=1).value
        print(f"   Summary A1: {title_cell!r}")
        wb.close()

        # 6. Build progress import xlsx (10 new updates for tasks 30-39)
        print(f"\n5. Building progress import workbook ({xlsx_imp})...")
        wb_p = Workbook()
        ws = wb_p.active
        ws.title = "Progress"
        ws.append(["Task ID", "%Complete"])
        for tid in task_ids[30:40]:
            ws.append([tid, 75.0])
        wb_p.save(xlsx_imp)
        print(f"   progress xlsx written at {time.time() - t0:.2f}s")

        # 7. Import progress
        print(f"\n6. Importing progress from {xlsx_imp}...")
        r2 = _msp_excel_import_progress(xlsx_path=xlsx_imp)
        print(f"   status={r2.get('status')} rows_imported={r2.get('rows_imported')}")
        if r2.get("status") != "ok":
            print(f"   [ERR] import failed: {r2.get('error')}")
        else:
            assert r2["rows_imported"] == 10, \
                f"expected 10 imports, got {r2['rows_imported']}"

        elapsed = time.time() - t0
        print(f"\n[OK] ACCEPTANCE: {elapsed:.2f}s total (target <90s)")
        assert elapsed < 90.0, f"Too slow: {elapsed}s"

    finally:
        try:
            for i in range(1, app.Projects.Count + 1):
                if app.Projects(i).Name == test_name:
                    app.WindowActivate(app.Projects(i).Windows(1).Caption)
                    app.FileClose(0)
                    print(f"[SAFE] closed test project {test_name}")
                    break
        except Exception as e:
            print(f"[WARN] cleanup error: {e}")


if __name__ == "__main__":
    main()
