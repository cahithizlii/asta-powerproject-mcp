"""Phase 5e XER native integration acceptance: end-to-end CAU XER pipeline.

Demonstrates the killer feature unlocked by Phase 5e: a single .xer
file_path flows through Phase 5b DCMA + Phase 5c Excel pipelines without
any conversion step.

NO MS Project COM required (XER pure-Python parse + DCMA pure-math + xlsx
all run without COM). Self-contained — builds fixture in temp dir.

Target: <30s wall clock.
"""
import functools
import os
import sys
import tempfile
import time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
print = functools.partial(print, flush=True)

from openpyxl import load_workbook

from msproject_mcp_core import (
    _msp_dcma_assess_all, _msp_dcma_summary, _msp_dcma_drill_down,
    _msp_excel_export_hakedis, _msp_excel_export_dcma,
)


SAMPLE_CAU_XER_CONTENT = """ERMHDR\t18.8\t2026-05-01\tcahit\tProject Management\tUSD
%T\tPROJECT
%F\tproj_id\tproj_short_name\tplan_start_date\tplan_end_date\tlast_recalc_date
%R\t1\tCAU\t2024-07-08 08:00\t2028-06-20 17:00\t2026-05-01 17:00
%T\tCALENDAR
%F\tclndr_id\tclndr_name\tday_hr_cnt\tweek_hr_cnt
%R\t1\tCAU 6x9\t9.0\t54.0
%T\tRSRC
%F\trsrc_id\trsrc_name\trsrc_short_name\trsrc_type\tmax_qty_per_hr
%R\t101\tConcrete Workers\tCOW\tRT_Labor\t10.0
%R\t102\tExtractors\tEXT\tRT_Labor\t5.0
%R\t103\tSteel\tSTL\tRT_Mat\t100.0
%R\t104\tCarpenters\tCAR\tRT_Labor\t8.0
%T\tTASK
%F\ttask_id\twbs_id\tproj_id\tclndr_id\ttask_code\ttask_name\ttask_type\ttarget_drtn_hr_cnt\ttarget_start_date\ttarget_end_date\tact_start_date\tact_end_date\tphys_complete_pct\ttotal_float_hr_cnt\tcstr_type\tstatus_code
%R\t1001\t1\t1\t1\tA1010\tFoundation\tTT_Task\t180.0\t2024-07-08 08:00\t2024-07-29 17:00\t2024-07-08 08:00\t2024-07-29 17:00\t100.0\t0.0\tCS_ASAP\tTK_Complete
%R\t1002\t1\t1\t1\tA1020\tFrame\tTT_Task\t360.0\t2024-07-30 08:00\t2024-09-09 17:00\t2024-07-30 08:00\t\t75.0\t0.0\tCS_ASAP\tTK_Active
%R\t1003\t1\t1\t1\tA1030\tWalls\tTT_Task\t180.0\t2024-09-10 08:00\t2024-10-01 17:00\t\t\t0.0\t72.0\tCS_ASAP\tTK_NotStart
%R\t1004\t1\t1\t1\tA1040\tRoof\tTT_Task\t180.0\t2024-10-02 08:00\t2024-10-23 17:00\t\t\t0.0\t72.0\tCS_ASAP\tTK_NotStart
%R\t1005\t1\t1\t1\tA1050\tInterior\tTT_Task\t360.0\t2024-10-24 08:00\t2024-12-04 17:00\t\t\t0.0\t72.0\tCS_ASAP\tTK_NotStart
%R\t1006\t1\t1\t1\tA1060\tHandover\tTT_FinMile\t0.0\t2024-12-15 17:00\t2024-12-15 17:00\t\t\t0.0\t81.0\tCS_MFO\tTK_NotStart
%T\tTASKPRED
%F\ttask_pred_id\ttask_id\tpred_task_id\tpred_type\tlag_hr_cnt
%R\t1\t1002\t1001\tPR_FS\t0.0
%R\t2\t1003\t1002\tPR_FS\t0.0
%R\t3\t1004\t1003\tPR_FS\t0.0
%R\t4\t1005\t1004\tPR_FS\t0.0
%R\t5\t1006\t1005\tPR_FS\t0.0
%T\tTASKRSRC
%F\ttaskrsrc_id\ttask_id\trsrc_id\ttarget_qty\tact_reg_qty\ttarget_cost\tact_reg_cost
%R\t1\t1001\t101\t180.0\t180.0\t180.0\t180.0
%R\t2\t1001\t103\t1000.0\t1000.0\t1000.0\t1000.0
%R\t3\t1002\t101\t360.0\t270.0\t360.0\t270.0
%R\t4\t1002\t104\t180.0\t135.0\t180.0\t135.0
%R\t5\t1003\t101\t180.0\t0.0\t180.0\t0.0
%R\t6\t1004\t104\t180.0\t0.0\t180.0\t0.0
%R\t7\t1005\t102\t360.0\t0.0\t360.0\t0.0
%E
"""


def main():
    out_dir = tempfile.mkdtemp(prefix="xer_integration_")
    xer_path = os.path.join(out_dir, "sample_cau.xer")
    with open(xer_path, "wb") as f:
        f.write(b"\xff\xfe")
        f.write(SAMPLE_CAU_XER_CONTENT.encode("utf-16-le"))
    print(f"[INFO] CAU XER fixture: {xer_path} ({os.path.getsize(xer_path)} bytes)")

    t0 = time.time()

    # 1. DCMA assess_all on .xer (Phase 5b through Phase 5e routing)
    print(f"\n1. msproject_health.assess_all(file_path=cau.xer)...")
    r = _msp_dcma_assess_all(file_path=xer_path)
    assert r["status"] == "ok", r
    assert len(r["rules"]) == 14
    print(f"   14 rules in {time.time()-t0:.3f}s")
    for rule in r["rules"]:
        ok = "OK  " if rule["status"] == "pass" else "FAIL"
        print(f"   [{ok}] Rule {rule['id']:2d}: {rule['name']:24s} "
              f"actual={rule.get('actual')}{rule.get('actual_unit', '')}")

    # 2. Summary
    s = _msp_dcma_summary(file_path=xer_path)
    print(f"\n2. Summary: {s['pass_count']}/14 pass, RAG={s['overall_rag'].upper()}")
    print(f"   {s['executive_text']}")

    # 3. Drill-down for first failed rule
    first_fail = next((rule for rule in r["rules"] if rule["status"] == "fail"), None)
    if first_fail:
        d = _msp_dcma_drill_down(file_path=xer_path, rule_id=first_fail["id"])
        print(f"\n3. Drill-down Rule {first_fail['id']} ({first_fail['name']}): "
              f"{d['failed_count']} failed tasks")
        for ft in d["failed_tasks"][:5]:
            print(f"   - Task {ft['id']}: {ft['name']}")

    # 4. Export DCMA-only xlsx
    dcma_xlsx = os.path.join(out_dir, "cau_dcma.xlsx")
    print(f"\n4. msproject_excel.export_dcma(file_path=cau.xer)...")
    r2 = _msp_excel_export_dcma(file_path=xer_path, xlsx_path=dcma_xlsx)
    assert r2["status"] == "ok"
    print(f"   {dcma_xlsx} ({os.path.getsize(dcma_xlsx)} bytes)")

    # 5. Full hakedis workbook
    hak_xlsx = os.path.join(out_dir, "cau_hakedis.xlsx")
    print(f"\n5. msproject_excel.export_hakedis(file_path=cau.xer)...")
    r3 = _msp_excel_export_hakedis(file_path=xer_path, xlsx_path=hak_xlsx)
    assert r3["status"] == "ok"
    print(f"   {hak_xlsx} ({os.path.getsize(hak_xlsx)} bytes)")
    print(f"   sheets: {r3['sheets_written']}")

    # 6. Verify hakedis structure
    wb = load_workbook(hak_xlsx, read_only=True)
    expected = {"Summary", "Tasks", "EVM_Compute", "EVM_TimePhased",
                "DCMA_Rules", "DCMA_Failed"}
    assert expected.issubset(set(wb.sheetnames)), \
        f"missing sheets: {expected - set(wb.sheetnames)}"
    print(f"\n6. Verified 6-sheet hakedis structure")
    wb.close()

    elapsed = time.time() - t0
    print(f"\n[OK] PHASE 5E ACCEPTANCE: {elapsed:.3f}s total (target <30s)")
    print(f"     Killer feature: CAU XER -> DCMA + Excel hakedis end-to-end")
    assert elapsed < 30.0, f"Too slow: {elapsed}s"


if __name__ == "__main__":
    main()
