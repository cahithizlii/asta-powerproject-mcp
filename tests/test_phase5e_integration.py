"""Test Phase 5e end-to-end: .xer file_path through DCMA + Excel pipelines (T110)."""
import os
import sys

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from msproject_mcp_core import (
    _evm_load_baseline_data,
    _msp_dcma_assess_all,
    _msp_dcma_summary,
    _msp_dcma_drill_down,
    _msp_excel_export_hakedis,
    _msp_excel_export_dcma,
)
from openpyxl import load_workbook


# ---- _evm_load_baseline_data XER routing ----

def test_evm_load_baseline_data_xer(sample_cau_xer):
    r = _evm_load_baseline_data(file_path=sample_cau_xer, baseline_number=0)
    assert r["status"] == "ok"
    assert "tasks" in r
    foundation = next(t for t in r["tasks"] if t["task_id"] == 1001)
    assert foundation["baseline_finish"] is not None
    assert foundation["baseline_work"] == 180.0


def test_evm_load_baseline_data_xer_excludes_summary(sample_cau_xer):
    r = _evm_load_baseline_data(file_path=sample_cau_xer)
    # CAU fixture: 6 real tasks, 0 summaries
    assert len(r["tasks"]) == 6


# ---- DCMA assess_all on .xer ----

def test_dcma_assess_all_xer(sample_cau_xer):
    r = _msp_dcma_assess_all(file_path=sample_cau_xer)
    assert r["status"] == "ok"
    assert len(r["rules"]) == 14
    assert "summary" in r


def test_dcma_summary_xer(sample_cau_xer):
    r = _msp_dcma_summary(file_path=sample_cau_xer)
    assert r["status"] == "ok"
    assert r["overall_rag"] in ("green", "amber", "red")


def test_dcma_drill_down_xer_rule_1(sample_cau_xer):
    """Rule 1 no_predecessor: Foundation has no preds."""
    r = _msp_dcma_drill_down(file_path=sample_cau_xer, rule_id=1)
    assert r["status"] == "ok"
    assert "failed_tasks" in r
    failed_ids = [t["id"] for t in r["failed_tasks"]]
    assert 1001 in failed_ids


def test_dcma_xer_rule_5_fs_link_pct(sample_cau_xer):
    """All 5 links FS → Rule 5 PASS (>90%)."""
    r = _msp_dcma_assess_all(file_path=sample_cau_xer)
    rule5 = next(rule for rule in r["rules"] if rule["id"] == 5)
    assert rule5["status"] == "pass"
    assert rule5["actual"] == 100.0


def test_dcma_xer_rule_13_critical_path(sample_cau_xer):
    """CAU XER chain has zero-slack tasks → critical_path > 0 → Rule 13 PASS."""
    r = _msp_dcma_assess_all(file_path=sample_cau_xer)
    rule13 = next(rule for rule in r["rules"] if rule["id"] == 13)
    assert rule13["status"] == "pass"
    assert rule13["actual"] >= 1


# ---- Excel exports on .xer ----

def test_excel_export_dcma_xer(sample_cau_xer, tmp_path):
    """export_dcma on XER produces 2-sheet xlsx."""
    xlsx = tmp_path / "dcma.xlsx"
    r = _msp_excel_export_dcma(file_path=sample_cau_xer, xlsx_path=str(xlsx))
    assert r["status"] == "ok"
    assert xlsx.exists()
    wb = load_workbook(str(xlsx), read_only=True)
    assert "DCMA_Rules" in wb.sheetnames
    assert "DCMA_Failed" in wb.sheetnames


def test_excel_export_hakedis_xer(sample_cau_xer, tmp_path):
    """export_hakedis on XER produces 6-sheet workbook."""
    xlsx = tmp_path / "hak.xlsx"
    r = _msp_excel_export_hakedis(file_path=sample_cau_xer, xlsx_path=str(xlsx))
    assert r["status"] == "ok"
    wb = load_workbook(str(xlsx), read_only=True)
    for s in ("Summary", "Tasks", "EVM_Compute", "EVM_TimePhased",
              "DCMA_Rules", "DCMA_Failed"):
        assert s in wb.sheetnames
