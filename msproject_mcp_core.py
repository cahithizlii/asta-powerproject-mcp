"""MS Project MCP Server - COM-based.

Hybrid speed strategy:
  1-5 items   -> COM direct (real-time, instant UI feedback)
  6-19 items  -> COM batch (Calculation manual + ScreenUpdating off)
  20+ items   -> MSPDI XML bulk import (~3-5s for 200 tasks)

Phase 1 tools: msproject_task, msproject_link, msproject_schedule.
"""
from __future__ import annotations
import atexit
import logging
import os
import threading
import time
from typing import Any, Callable, Dict, List, Optional, Tuple

import datetime as _dt

import pythoncom
import pywintypes
import win32com.client
from mcp.server.fastmcp import FastMCP

# ---------- LOGGING ----------
log_dir = os.path.expanduser("~/.claude/logs")
os.makedirs(log_dir, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.FileHandler(os.path.join(log_dir, "msproject_mcp.log"), encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("msproject_mcp")

# ---------- MCP SERVER ----------
mcp = FastMCP(
    "msproject_mcp",
    instructions=(
        "MS Project COM-based MCP server. Connects to running MS Project (Application='MSProject.Application'). "
        "Hybrid speed: 1-5 items COM direct, 6-19 batch, 20+ MSPDI bulk import. "
        "Tools: msproject_task, msproject_link, msproject_schedule, msproject_calendar, msproject_resource, msproject_baseline, msproject_progress."
    ),
)

# ---------- COM CONNECTION CACHE ----------
_app_lock = threading.RLock()
_app: Optional[Any] = None
_calc_modified = False  # track if we changed calc mode (need restore on exit)
_screenupdating_modified = False


def _connect_app() -> Any:
    """Connect to running MS Project. Cached singleton."""
    global _app
    with _app_lock:
        if _app is not None:
            try:
                _ = _app.Version  # ping
                return _app
            except Exception:
                _app = None  # invalidate
        pythoncom.CoInitialize()
        try:
            _app = win32com.client.GetActiveObject("MSProject.Application")
            logger.info(f"Connected to MS Project {_app.Version}")
        except Exception as e:
            raise RuntimeError(
                f"MS Project'e bağlanılamadı: {e}. "
                "(1) MS Project açık olduğundan emin olun. "
                "(2) Bir proje açık olmalı (boş Project bile yeterli). "
                "(3) Hala olmuyorsa MS Project'i yeniden başlatın."
            )
        return _app


def _validate_active_project() -> Any:
    """Validates ActiveProject is present."""
    app = _connect_app()
    if app.ActiveProject is None:
        raise RuntimeError("MS Project'te aktif proje yok. Boş bir proje açın veya File → New.")
    return app


def _route_operation(op_count: int) -> str:
    """Pick speed path based on operation count."""
    if op_count <= 5:
        return "com_direct"
    if op_count <= 19:
        return "com_batch"
    return "mspdi_bulk"


def _format_com_error(e: Exception) -> str:
    """Extract a human-readable message from pywintypes.com_error or fallback to str(e).

    pywintypes.com_error.args = (hresult, msg, excepinfo, argerr) where
    excepinfo = (wCode, source, description, helpFile, helpContext, scode).
    Description (excepinfo[2]) is the user-friendly message; fall back to
    msg (args[1]) or str(e) if unavailable.
    """
    try:
        if hasattr(e, "args") and len(e.args) >= 3 and isinstance(e.args[2], tuple):
            excepinfo = e.args[2]
            if len(excepinfo) >= 3 and excepinfo[2]:
                return str(excepinfo[2]).strip()
            if len(e.args) >= 2 and e.args[1]:
                return str(e.args[1]).strip()
    except Exception:
        pass
    return str(e)


def _enter_batch_mode():
    """Enter COM batch mode: disable screen update, manual calc, no events."""
    with _app_lock:
        global _calc_modified, _screenupdating_modified
        app = _connect_app()
        pj_manual = 0  # PjCalculation.pjManual
        if app.Calculation != pj_manual:
            app.Calculation = pj_manual
            _calc_modified = True
        if app.ScreenUpdating:
            app.ScreenUpdating = False
            _screenupdating_modified = True
        proj = app.ActiveProject
        try:
            proj.EventsEnabled = False
        except Exception:
            pass


def _exit_batch_mode():
    """Restore screen update + auto calc + events."""
    global _calc_modified, _screenupdating_modified
    with _app_lock:
        try:
            app = _connect_app()
            if _calc_modified:
                pj_auto = 1  # PjCalculation.pjAutomatic
                app.Calculation = pj_auto
            if _screenupdating_modified:
                app.ScreenUpdating = True
            proj = app.ActiveProject
            if proj:
                try:
                    proj.EventsEnabled = True
                except Exception:
                    pass
        except Exception as e:
            logger.warning(f"_exit_batch_mode error (non-fatal): {e}")
        finally:
            # Reset flags even if restore failed — otherwise stuck "modified"
            # state would prevent future entries from re-applying batch mode.
            _calc_modified = False
            _screenupdating_modified = False


@atexit.register
def _restore_on_exit():
    """Critical: ensure MS Project never left in manual/screen-off state."""
    _exit_batch_mode()


# ---------- TASK HELPERS ----------

def _parse_duration(d: str) -> int:
    """Convert '5d' -> minutes (assuming 8h/day for MSP)."""
    if not d:
        return 480
    s = d.strip()
    unit = s[-1].lower() if s[-1].isalpha() else "d"
    try:
        n = float(s.rstrip("dwhmDWHM"))
    except ValueError:
        n = 1.0
    return int({"d": 480, "w": 2400, "h": 60, "m": 1}.get(unit, 480) * n)


def _msp_task_add_single(name: str, duration: str = "1d",
                         start: Optional[str] = None,
                         finish: Optional[str] = None,
                         summary: bool = False,
                         milestone: bool = False,
                         notes: Optional[str] = None) -> Dict[str, Any]:
    """Add a single task via COM (Path 1)."""
    app = _validate_active_project()
    proj = app.ActiveProject
    try:
        new_task = proj.Tasks.Add(name)
        if milestone:
            new_task.Milestone = True
            new_task.Duration = 0
        elif duration:
            new_task.Duration = _parse_duration(duration)
        if start:
            new_task.Start = start
        if finish:
            new_task.Finish = finish
        if summary:
            try:
                new_task.Summary = True
            except Exception:
                pass  # Summary auto-set when task has children in MSP
        if notes:
            new_task.Notes = notes
        return {
            "status": "ok",
            "task_id": new_task.ID,
            "task_uid": new_task.UniqueID,
            "name": new_task.Name,
            "duration": duration,
            "milestone": bool(milestone),
        }
    except Exception as e:
        logger.error(f"_msp_task_add_single failed: {e}")
        return {"status": "error", "error": str(e)}


def _find_task_by_id(proj: Any, task_id: int) -> Optional[Any]:
    """Locate a task object by its ID. Returns None if not found.

    Defensive: `t.ID` access can raise pywintypes.com_error for invalid
    handles (e.g. tasks deleted/recycled mid-iteration). Such failures
    are treated as "not this task" and we continue scanning rather than
    aborting the whole lookup — Phase 5a TAIL fix (T84 acceptance script
    intermittent COM error on bulk progress loops).
    """
    for i in range(1, proj.Tasks.Count + 1):
        try:
            t = proj.Tasks(i)
            if t is not None and t.ID == task_id:
                return t
        except Exception:
            continue
    return None


# ---------- CALENDAR CONSTANTS ----------

UZBEK_HOLIDAYS_2026 = [
    ("Yılbaşı", 1, 1),
    ("Vatan Müdafaası Günü", 1, 14),
    ("Kadınlar Günü", 3, 8),
    ("Navruz", 3, 21),
    ("İşçi Bayramı", 5, 1),
    ("Hatıra ve Şeref Günü", 5, 9),
    ("Bağımsızlık Günü", 9, 1),
    ("Öğretmenler Günü", 10, 1),
    ("Anayasa Günü", 12, 8),
]

# Calendar dispatcher: actions that natively use 'calendar_name' and accept
# 'name' as alias.
_CALENDAR_NAME_ALIAS_ACTIONS = frozenset({
    "add_exception", "assign_to_task", "assign_to_resource",
    "list", "holidays_uzbek",
})
# Calendar dispatcher: actions that natively use 'name' and accept
# 'calendar_name' as alias.
_CALENDAR_NAME_NATIVE_ACTIONS = frozenset({"create", "update"})


# ---------- CALENDAR HELPERS ----------

def _find_calendar_by_name(proj: Any, name: str) -> Optional[Any]:
    """Locate a base calendar object in the project. Returns None if not found."""
    for i in range(1, proj.BaseCalendars.Count + 1):
        cal = proj.BaseCalendars(i)
        if cal is not None and cal.Name == name:
            return cal
    return None


def _msp_calendar_create(name: str, base_calendar: str = "Standard") -> Dict[str, Any]:
    """Create a new project-scoped base calendar copied from an existing one."""
    app = _validate_active_project()
    proj = app.ActiveProject
    # Pre-flight: name conflict
    if _find_calendar_by_name(proj, name) is not None:
        return {"status": "error", "error": f"Calendar '{name}' already exists"}
    # Pre-flight: base must exist
    if _find_calendar_by_name(proj, base_calendar) is None:
        return {"status": "error",
                "error": f"Base calendar '{base_calendar}' not found in project"}
    try:
        # app.BaseCalendarCreate creates a project-scoped base calendar
        app.BaseCalendarCreate(Name=name, FromName=base_calendar)
        cal = _find_calendar_by_name(proj, name)
        if cal is None:
            return {"status": "error",
                    "error": f"BaseCalendarCreate succeeded but '{name}' not found"}
        return {"status": "ok", "calendar_uid": cal.Guid, "name": name}
    except Exception as e:
        logger.error(f"_msp_calendar_create({name}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_calendar_update(name: str,
                         new_name: Optional[str] = None,
                         weekday_off: Optional[int] = None) -> Dict[str, Any]:
    """Rename a calendar and/or mark a weekday as non-working.

    weekday_off: 1=Sunday, 2=Monday, ..., 7=Saturday (MS Project convention).
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    cal = _find_calendar_by_name(proj, name)
    if cal is None:
        return {"status": "error", "error": f"Calendar '{name}' not found in project"}

    # Pre-flight: validate ALL inputs before any mutation (no partial writes)
    do_rename = new_name is not None and new_name != name
    if do_rename and _find_calendar_by_name(proj, new_name) is not None:
        return {"status": "error",
                "error": f"Calendar '{new_name}' already exists"}
    if weekday_off is not None and not (1 <= weekday_off <= 7):
        return {"status": "error",
                "error": "weekday_off must be 1-7 (1=Sunday, 7=Saturday)"}

    changes = []
    try:
        if do_rename:
            cal.Name = new_name
            changes.append("name")
        if weekday_off is not None:
            wd = cal.WeekDays(weekday_off)
            wd.Working = False
            changes.append("weekday_off")
        return {"status": "ok", "calendar_name": new_name or name, "changes": changes}
    except Exception as e:
        logger.error(f"_msp_calendar_update({name}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _parse_date(date_str: str) -> _dt.date:
    """Parse 'YYYY-MM-DD' to date."""
    return _dt.datetime.strptime(date_str, "%Y-%m-%d").date()


# PjExceptionType (Microsoft.Office.Interop.MSProject)
PJ_EXCEPTION_DAILY = 7              # single fixed-date or range (non-recurring)
PJ_EXCEPTION_RECUR_DAILY = 1        # Phase 10.2 — recurring daily
PJ_EXCEPTION_RECUR_WEEKLY = 2       # Phase 10.2 — recurring weekly
PJ_EXCEPTION_RECUR_MONTHLY = 4      # Phase 10.2 — monthly by day-of-month
PJ_EXCEPTION_RECUR_YEARLY = 5       # Phase 10.2 — yearly by day-of-month

# MSP DaysOfWeek bitmask (Phase 10.2)
_DAY_OF_WEEK_BITS = {
    "sun": 1, "sunday": 1,
    "mon": 2, "monday": 2,
    "tue": 4, "tuesday": 4,
    "wed": 8, "wednesday": 8,
    "thu": 16, "thursday": 16,
    "fri": 32, "friday": 32,
    "sat": 64, "saturday": 64,
}


def _parse_hhmm_time(s: str):
    """Parse 'HH:MM' or 'H:MM' string -> datetime.time. Phase 10.3."""
    parts = s.split(":")
    if len(parts) != 2:
        raise ValueError(f"expected HH:MM, got '{s}'")
    return _dt.time(int(parts[0]), int(parts[1]))


def _msp_calendar_add_exception(calendar_name: str, exception_name: str,
                                start: str,
                                finish: Optional[str] = None,
                                working: bool = False,
                                recurrence: Optional[str] = None,
                                days_of_week: Optional[List[str]] = None,
                                period: Optional[int] = None,
                                occurrences: Optional[int] = None,
                                working_hours_start: str = "08:00",
                                working_hours_finish: str = "17:00") -> Dict[str, Any]:
    """Add an exception to a calendar — Phase 2a + Phase 10.2 + Phase 10.3.

    Args:
        calendar_name, exception_name: targets.
        start: 'YYYY-MM-DD'.
        finish: optional 'YYYY-MM-DD' (defaults to start). For recurring
            exceptions, this is the recurrence END date.
        working: Phase 10.3 — if True, exception is a working window
            (overrides base calendar non-working). Default working
            window 08:00-17:00; override via working_hours_*.
        recurrence: Phase 10.2 — None (single date, default) or
            'daily'/'weekly'/'monthly'/'yearly'.
        days_of_week: Phase 10.2 — required for recurrence='weekly'.
            Day names: mon/tue/wed/thu/fri/sat/sun (case-insensitive).
        period: every N occurrences (e.g. 2 = every other week). Optional.
        occurrences: total recurrence count. Optional.
        working_hours_start/finish: Phase 10.3 — 'HH:MM' for working=True.

    Pre-Phase 10.2/10.3 callers continue to work — recurrence default
    None preserves single-date Type=7 behavior; working default False
    preserves non-working semantics.
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    cal = _find_calendar_by_name(proj, calendar_name)
    if cal is None:
        return {"status": "error",
                "error": f"Calendar '{calendar_name}' not found in project"}

    rec = recurrence.lower() if recurrence else None
    if rec is not None and rec not in ("daily", "weekly", "monthly", "yearly"):
        return {"status": "error",
                "error": (f"recurrence must be None/daily/weekly/monthly/"
                          f"yearly (got '{recurrence}')")}
    if rec == "weekly" and not days_of_week:
        return {"status": "error",
                "error": "recurrence='weekly' requires days_of_week list"}

    # Pre-flight validation BEFORE mutation
    try:
        start_d = _parse_date(start)
        finish_d = _parse_date(finish) if finish else start_d
    except ValueError as e:
        return {"status": "error",
                "error": f"Invalid date format (expected YYYY-MM-DD): {e}"}
    if finish_d < start_d:
        return {"status": "error",
                "error": "Start date must be <= finish date"}

    if working:
        try:
            wh_start_t = _parse_hhmm_time(working_hours_start)
            wh_finish_t = _parse_hhmm_time(working_hours_finish)
        except ValueError as e:
            return {"status": "error",
                    "error": f"Invalid working_hours_* format: {e}"}
    else:
        wh_start_t = wh_finish_t = None

    bitmask = 0
    if rec == "weekly":
        for d in days_of_week:
            bit = _DAY_OF_WEEK_BITS.get(str(d).lower())
            if bit is None:
                return {"status": "error",
                        "error": (f"Unknown day_of_week '{d}'. Valid: "
                                  f"mon/tue/wed/thu/fri/sat/sun")}
            bitmask |= bit

    type_map = {
        None: PJ_EXCEPTION_DAILY,
        "daily": PJ_EXCEPTION_RECUR_DAILY,
        "weekly": PJ_EXCEPTION_RECUR_WEEKLY,
        "monthly": PJ_EXCEPTION_RECUR_MONTHLY,
        "yearly": PJ_EXCEPTION_RECUR_YEARLY,
    }
    ex_type = type_map[rec]

    try:
        ex = cal.Exceptions.Add(
            Type=ex_type,
            Start=pywintypes.Time(start_d),
            Finish=pywintypes.Time(finish_d),
        )
        ex.Name = exception_name
        if rec == "weekly":
            try:
                ex.DaysOfWeek = bitmask
            except Exception as e:
                logger.warning(f"DaysOfWeek not settable: {e}")
        if rec in ("monthly", "yearly"):
            try:
                ex.MonthDay = start_d.day
            except Exception as e:
                logger.warning(f"MonthDay not settable: {e}")
            if rec == "yearly":
                try:
                    ex.Month = start_d.month
                except Exception as e:
                    logger.warning(f"Month not settable: {e}")
        if period is not None and period > 0:
            try:
                ex.Period = period
            except Exception as e:
                logger.warning(f"Period not settable: {e}")
        if occurrences is not None and occurrences > 0:
            try:
                ex.Occurrences = occurrences
            except Exception as e:
                logger.warning(f"Occurrences not settable: {e}")
        if working:
            # MSP 16: ex.Shift1.Start/Finish accepts COM Time. Combine
            # working hours with exception start_d for a valid datetime.
            wh_start_dt = _dt.datetime.combine(start_d, wh_start_t)
            wh_finish_dt = _dt.datetime.combine(start_d, wh_finish_t)
            try:
                ex.Shift1.Start = pywintypes.Time(wh_start_dt)
                ex.Shift1.Finish = pywintypes.Time(wh_finish_dt)
            except Exception as e:
                logger.warning(f"working hours (Shift1) not settable: {e}")
        return {"status": "ok",
                "calendar_name": calendar_name,
                "exception_name": exception_name,
                "start": start,
                "finish": finish or start,
                "recurrence": rec,
                "working": working}
    except Exception as e:
        logger.error(
            f"_msp_calendar_add_exception({calendar_name},{exception_name}) failed: {e}"
        )
        return {"status": "error", "error": _format_com_error(e)}


def _msp_calendar_assign_to_task(task_id: int, calendar_name: str) -> Dict[str, Any]:
    """Assign a base calendar to a specific task."""
    app = _validate_active_project()
    proj = app.ActiveProject
    cal = _find_calendar_by_name(proj, calendar_name)
    if cal is None:
        return {"status": "error",
                "error": f"Calendar '{calendar_name}' not found in project"}
    t = _find_task_by_id(proj, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}
    try:
        t.Calendar = calendar_name
        return {"status": "ok", "task_id": task_id, "calendar_name": calendar_name}
    except Exception as e:
        logger.error(f"_msp_calendar_assign_to_task({task_id},{calendar_name}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _find_resource_by_id(proj: Any, resource_id: int) -> Optional[Any]:
    """Locate a resource by ID. Phase 2a helper; reused/expanded in Phase 2b."""
    for i in range(1, proj.Resources.Count + 1):
        r = proj.Resources(i)
        if r is not None and r.ID == resource_id:
            return r
    return None


def _msp_calendar_assign_to_resource(resource_id: int, calendar_name: str) -> Dict[str, Any]:
    """Assign a base calendar to a resource."""
    app = _validate_active_project()
    proj = app.ActiveProject
    cal = _find_calendar_by_name(proj, calendar_name)
    if cal is None:
        return {"status": "error",
                "error": f"Calendar '{calendar_name}' not found in project"}
    res = _find_resource_by_id(proj, resource_id)
    if res is None:
        return {"status": "error", "error": f"Resource ID {resource_id} not found"}
    try:
        res.BaseCalendar = calendar_name
        return {"status": "ok", "resource_id": resource_id, "calendar_name": calendar_name}
    except Exception as e:
        logger.error(f"_msp_calendar_assign_to_resource({resource_id},{calendar_name}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_calendar_list() -> Dict[str, Any]:
    """List all base calendars in the active project with exception counts.

    Order: matches `proj.BaseCalendars` enumeration (typically insertion
    order, not sorted). If callers need lexicographic ordering, sort
    client-side on the `name` field.
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    out = []
    try:
        for i in range(1, proj.BaseCalendars.Count + 1):
            cal = proj.BaseCalendars(i)
            if cal is None:
                continue
            try:
                ex_count = cal.Exceptions.Count
            except Exception:
                ex_count = 0
            out.append({
                "calendar_uid": cal.Guid,
                "name": cal.Name,
                "exception_count": ex_count,
            })
        return {"status": "ok", "count": len(out), "calendars": out}
    except Exception as e:
        logger.error(f"_msp_calendar_list failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_calendar_holidays_uzbek(calendar_name: str, year: int = 2026) -> Dict[str, Any]:
    """Bulk-add 9 official Özbekistan public holidays to a calendar.

    Idempotent: name-based dedup. Re-running on a calendar that already
    has matching named exceptions skips them (returns them in `skipped`).
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    cal = _find_calendar_by_name(proj, calendar_name)
    if cal is None:
        return {"status": "error",
                "error": f"Calendar '{calendar_name}' not found in project"}

    # Pre-scan existing exception names for dedup
    existing_names = set()
    try:
        for i in range(1, cal.Exceptions.Count + 1):
            ex = cal.Exceptions(i)
            if ex is not None and ex.Name:
                existing_names.add(ex.Name)
    except Exception as e:
        logger.debug(f"holidays_uzbek pre-scan failed (treating calendar as empty): {_format_com_error(e)}")

    added = []
    skipped = []
    failures = []
    for name, month, day in UZBEK_HOLIDAYS_2026:
        date_str = f"{year:04d}-{month:02d}-{day:02d}"
        entry = {"name": name, "date": date_str, "month": month, "day": day}
        if name in existing_names:
            skipped.append({**entry, "reason": "already exists"})
            continue
        r = _msp_calendar_add_exception(
            calendar_name=calendar_name,
            exception_name=name,
            start=date_str,
        )
        if r.get("status") == "ok":
            added.append(entry)
        else:
            failures.append({**entry, "error": r.get("error")})

    if failures:
        status = "partial"
    elif not added and skipped:
        status = "already_done"
    else:
        status = "ok"

    if failures:
        logger.warning(f"holidays_uzbek partial: {len(added)} added, {len(skipped)} skipped, {len(failures)} failed")

    return {
        "status": status,
        "calendar_name": calendar_name,
        "year": year,
        "count": len(added),
        "skipped_count": len(skipped),
        "holidays": added,
        "skipped": skipped,
        "failures": failures,
    }


# ---------- RESOURCE CONSTANTS ----------

# COM PjResourceType enum: Work=0, Material=1, Cost=2
RESOURCE_TYPES = {"Work": 0, "Material": 1, "Cost": 2}
RESOURCE_TYPE_NAMES = {v: k for k, v in RESOURCE_TYPES.items()}


# ---------- RESOURCE HELPERS ----------

def _find_resource_by_name(proj: Any, name: str) -> Optional[Any]:
    """Locate a resource by name. Returns None if not found."""
    for i in range(1, proj.Resources.Count + 1):
        r = proj.Resources(i)
        if r is not None and r.Name == name:
            return r
    return None


def _parse_rate(raw: Any) -> float:
    """Parse a MS Project rate value to float.

    MS Project COM returns rates as locale-formatted strings (e.g. '$50.00/h',
    '₺50,00/hr', '50,00 €/sa'). Strip currency symbols + per-unit suffix and
    normalize decimal separator. Returns 0.0 for empty/unparseable values.
    """
    if raw is None or raw == "":
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s:
        return 0.0
    # Drop per-unit suffix after '/' (e.g. '/h', '/hr', '/sa', '/saat')
    if "/" in s:
        s = s.split("/", 1)[0]
    # Strip everything except digits, separators, and minus sign
    cleaned = "".join(ch for ch in s if ch.isdigit() or ch in ".,-")
    if not cleaned or cleaned in ("-", ".", ","):
        return 0.0
    # Locale-aware decimal separator: if both present, last one wins as decimal
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        # Comma-only: treat as decimal separator (TR, EU locales)
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _serialize_resource(res: Any) -> Dict[str, Any]:
    """Type-aware serialization. MaxUnits is COM-stored as fraction (1.0 = 100%);
    we expose as percentage for symmetry with assignment Units. Rates are
    locale-formatted strings from COM and are parsed to float via _parse_rate.

    Per-field reads are guarded so one bad COM value (corrupt/stale row) does
    not kill list iteration in callers like T35 resource_list. id/uid/name are
    intentionally unguarded — failure there means the resource is genuinely
    broken and the caller should know.
    """
    type_code = 0
    try:
        type_code = int(res.Type) if res.Type is not None else 0
    except Exception:
        pass
    type_name = RESOURCE_TYPE_NAMES.get(type_code, "Work")
    out: Dict[str, Any] = {
        "id": res.ID,
        "uid": res.UniqueID,
        "name": res.Name,
        "type": type_name,
    }
    # Type-specific properties — guarded so a single bad field doesn't break list iteration
    try:
        if type_name == "Work":
            out["max_units"] = float(res.MaxUnits) * 100.0  # 1.0 -> 100%
            out["standard_rate"] = _parse_rate(res.StandardRate)
            out["overtime_rate"] = _parse_rate(res.OvertimeRate)
        elif type_name == "Material":
            out["material_label"] = res.MaterialLabel or ""
            out["standard_rate"] = _parse_rate(res.StandardRate)
        elif type_name == "Cost":
            out["standard_rate"] = _parse_rate(res.StandardRate)
    except Exception:
        # Field read failed — return what we have (id/uid/name/type)
        pass
    return out


def _msp_resource_add(name: str, type: str = "Work",
                     max_units: Optional[float] = None,
                     standard_rate: Optional[float] = None,
                     overtime_rate: Optional[float] = None,
                     material_label: Optional[str] = None) -> Dict[str, Any]:
    """Add a resource. Type: 'Work' (default) | 'Material' | 'Cost'.

    max_units in % (100 = 1 person, 500 = 5-person crew). Stored in COM as
    fraction. MSP's documented ceiling is 6000% (60.0 fraction).
    standard_rate / overtime_rate in $/hour (Work) or $/unit (Material).
    material_label e.g. 'kg', 'm³', 'ton'.

    Atomic: if any post-Add property set fails, the orphan resource is
    deleted before returning error.
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    # Pre-flight validation
    if type not in RESOURCE_TYPES:
        return {"status": "error",
                "error": f"Invalid type '{type}'. Valid: Work/Material/Cost"}
    if _find_resource_by_name(proj, name) is not None:
        return {"status": "error", "error": f"Resource '{name}' already exists"}
    # Pre-flight value validation (Fix 2)
    if max_units is not None and max_units < 0:
        return {"status": "error", "error": "max_units must be >= 0"}
    if standard_rate is not None and standard_rate < 0:
        return {"status": "error", "error": "standard_rate must be >= 0"}
    if overtime_rate is not None and overtime_rate < 0:
        return {"status": "error", "error": "overtime_rate must be >= 0"}
    try:
        res = proj.Resources.Add(name)
    except Exception as e:
        logger.error(f"_msp_resource_add({name},{type}) Resources.Add failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}
    # Post-Add: any failure here must rollback the orphan
    try:
        res.Type = RESOURCE_TYPES[type]
        if type == "Work":
            res.MaxUnits = (max_units / 100.0) if max_units is not None else 1.0
            if standard_rate is not None:
                res.StandardRate = standard_rate
            if overtime_rate is not None:
                res.OvertimeRate = overtime_rate
        elif type == "Material":
            if material_label is not None:
                res.MaterialLabel = material_label
            if standard_rate is not None:
                res.StandardRate = standard_rate
        elif type == "Cost":
            if standard_rate is not None:
                res.StandardRate = standard_rate
        return {"status": "ok", "resource_id": res.ID, "resource_uid": res.UniqueID,
                "name": name, "type": type}
    except Exception as e:
        logger.error(f"_msp_resource_add({name},{type}) property-set failed (rolling back): {e}")
        try:
            res.Delete()
        except Exception as del_err:
            logger.warning(f"_msp_resource_add rollback delete failed: {del_err}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_resource_update(resource_id: int,
                        name: Optional[str] = None,
                        max_units: Optional[float] = None,
                        standard_rate: Optional[float] = None,
                        overtime_rate: Optional[float] = None,
                        material_label: Optional[str] = None) -> Dict[str, Any]:
    """Update a resource. Pre-flight validates ALL inputs before mutation
    (no partial writes — T20 lesson).

    max_units in % (100 = 1 person). standard_rate / overtime_rate in $/unit.
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    res = _find_resource_by_id(proj, resource_id)
    if res is None:
        return {"status": "error", "error": f"Resource ID {resource_id} not found"}

    # Pre-flight: validate ALL inputs before any mutation
    do_rename = name is not None and name != res.Name
    if do_rename and _find_resource_by_name(proj, name) is not None:
        return {"status": "error", "error": f"Resource '{name}' already exists"}
    if max_units is not None and max_units < 0:
        return {"status": "error", "error": "max_units must be >= 0"}
    if standard_rate is not None and standard_rate < 0:
        return {"status": "error", "error": "standard_rate must be >= 0"}
    if overtime_rate is not None and overtime_rate < 0:
        return {"status": "error", "error": "overtime_rate must be >= 0"}

    # Type-aware property validation: each property must match the resource's type
    res_type_code = int(res.Type) if res.Type is not None else 0
    res_type = RESOURCE_TYPE_NAMES.get(res_type_code, "Work")
    if res_type == "Work":
        if material_label is not None:
            return {"status": "error",
                    "error": f"material_label not applicable to Work resource (type={res_type})"}
    elif res_type == "Material":
        if max_units is not None:
            return {"status": "error",
                    "error": f"max_units not applicable to Material resource (type={res_type})"}
        if overtime_rate is not None:
            return {"status": "error",
                    "error": f"overtime_rate not applicable to Material resource (type={res_type})"}
    elif res_type == "Cost":
        if max_units is not None:
            return {"status": "error",
                    "error": f"max_units not applicable to Cost resource (type={res_type})"}
        if overtime_rate is not None:
            return {"status": "error",
                    "error": f"overtime_rate not applicable to Cost resource (type={res_type})"}
        if material_label is not None:
            return {"status": "error",
                    "error": f"material_label not applicable to Cost resource (type={res_type})"}

    # Empty-changes no-op intentionally returns ok (mirrors T20 calendar update);
    # caller can detect via empty `changes` list.
    changes = []
    try:
        if do_rename:
            res.Name = name; changes.append("name")
        if max_units is not None:
            res.MaxUnits = max_units / 100.0; changes.append("max_units")
        if standard_rate is not None:
            res.StandardRate = standard_rate; changes.append("standard_rate")
        if overtime_rate is not None:
            res.OvertimeRate = overtime_rate; changes.append("overtime_rate")
        if material_label is not None:
            res.MaterialLabel = material_label; changes.append("material_label")
        return {"status": "ok", "resource_id": resource_id, "changes": changes}
    except Exception as e:
        logger.error(f"_msp_resource_update({resource_id}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_resource_delete(resource_id: int) -> Dict[str, Any]:
    """Delete a resource by ID. Cascades: any assignments to this resource
    are silently removed by MS Project; the count is returned in the response.
    """
    app = _validate_active_project()
    res = _find_resource_by_id(app.ActiveProject, resource_id)
    if res is None:
        return {"status": "error", "error": f"Resource ID {resource_id} not found"}
    try:
        name = res.Name
        # Capture cascade-affected assignment count BEFORE delete
        try:
            assignments_removed = int(res.Assignments.Count)
        except Exception:
            assignments_removed = 0
        res.Delete()
        return {"status": "ok", "deleted_id": resource_id,
                "deleted_name": name, "assignments_removed": assignments_removed}
    except Exception as e:
        logger.error(f"_msp_resource_delete({resource_id}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_resource_list() -> Dict[str, Any]:
    """List all resources in the active project with type-aware properties + assignment counts.

    Order: matches `proj.Resources` enumeration (typically insertion order, not sorted).
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    out = []
    try:
        for i in range(1, proj.Resources.Count + 1):
            res = proj.Resources(i)
            if res is None:
                continue
            entry = _serialize_resource(res)
            try:
                entry["assignment_count"] = res.Assignments.Count
            except Exception:
                entry["assignment_count"] = 0
            out.append(entry)
        return {"status": "ok", "count": len(out), "resources": out}
    except Exception as e:
        logger.error(f"_msp_resource_list failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_resource_assign(task_id: int, resource_id: int,
                        units: Optional[float] = None,
                        work_hours: Optional[float] = None) -> Dict[str, Any]:
    """Assign a resource to a task. Units in % (100 = full-time, default).

    Uses MS Project's `task.Assignments.Add(TaskID, ResourceID, [Units])`.
    work_hours overrides the auto-calculated work duration.
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    # Pre-flight validation
    t = _find_task_by_id(proj, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}
    res = _find_resource_by_id(proj, resource_id)
    if res is None:
        return {"status": "error", "error": f"Resource ID {resource_id} not found"}
    if units is not None and units < 0:
        return {"status": "error", "error": "units must be >= 0"}
    if work_hours is not None and work_hours < 0:
        return {"status": "error", "error": "work_hours must be >= 0"}
    try:
        applied_units = units if units is not None else 100.0
        # task.Assignments.Add(TaskID, ResourceID, Units) — Units as fraction (1.0=100%)
        alloc = t.Assignments.Add(TaskID=task_id, ResourceID=resource_id,
                                 Units=applied_units / 100.0)
        warnings = []
        if work_hours is not None:
            try:
                # COM Work property accepts minutes
                alloc.Work = work_hours * 60.0
            except Exception as wh_err:
                warnings.append(f"work_hours not applied: {_format_com_error(wh_err)}")
        result = {"status": "ok",
                  "assignment_uid": alloc.UniqueID if alloc else None,
                  "task_id": task_id,
                  "resource_id": resource_id,
                  "units": applied_units}
        if warnings:
            result["warnings"] = warnings
        return result
    except Exception as e:
        logger.error(f"_msp_resource_assign({task_id},{resource_id}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_resource_assign_unsafe(task_obj: Any, res_obj: Any,
                               task_id: int, resource_id: int,
                               units: Optional[float] = None) -> Dict[str, Any]:
    """Internal fast-path for bulk: caller pre-resolved task_obj + res_obj.

    Skips _validate_active_project + _find_task_by_id + _find_resource_by_id —
    only safe when caller has just pre-built ID->object maps. NOT for public use.
    """
    try:
        applied_units = units if units is not None else 100.0
        if applied_units < 0:
            return {"status": "error", "error": "units must be >= 0",
                    "task_id": task_id, "resource_id": resource_id}
        # task.Assignments.Add is the only available API — proj.Assignments
        # does not exist on _IProjectDoc (probed 2026-04-28: AttributeError).
        # Measured ~9.6ms/call; not improvable via project-scoped collection.
        alloc = task_obj.Assignments.Add(TaskID=task_id, ResourceID=resource_id,
                                         Units=applied_units / 100.0)
        return {"status": "ok",
                "assignment_uid": alloc.UniqueID if alloc else None,
                "task_id": task_id,
                "resource_id": resource_id,
                "units": applied_units}
    except Exception as e:
        # Don't log per-call in bulk path — caller aggregates
        return {"status": "error",
                "task_id": task_id, "resource_id": resource_id,
                "error": _format_com_error(e)}


def _build_resource_id_map(proj: Any) -> Dict[int, Any]:
    """Pre-build resource_id -> Resource COM object map. O(N) one-time scan.

    Used by bulk_assign to avoid O(N×M) per-item lookup blow-up.
    """
    out: Dict[int, Any] = {}
    for i in range(1, proj.Resources.Count + 1):
        r = proj.Resources(i)
        if r is not None:
            out[r.ID] = r
    return out


def _build_task_id_map(proj: Any) -> Dict[int, Any]:
    """Pre-build task_id -> Task COM object map. O(N) one-time scan.

    Used by bulk_assign to avoid O(N×M) per-item lookup blow-up.
    """
    out: Dict[int, Any] = {}
    for i in range(1, proj.Tasks.Count + 1):
        t = proj.Tasks(i)
        if t is not None:
            out[t.ID] = t
    return out


def _msp_resource_bulk_assign_loop(items: List[Dict[str, Any]], path_label: str,
                                   task_map: Dict[int, Any],
                                   res_map: Dict[int, Any]) -> Dict[str, Any]:
    """Inner loop using pre-built maps + _assign_unsafe fast-path.

    Returns aggregated result with status / path / count / assignments / failures.
    """
    added: List[Dict[str, Any]] = []
    failures: List[Dict[str, Any]] = []
    for item in items:
        tid = item.get("task_id")
        rid = item.get("resource_id")
        t_obj = task_map.get(tid)
        r_obj = res_map.get(rid)
        if t_obj is None:
            failures.append({**item, "error": f"task_id {tid} not found"})
            continue
        if r_obj is None:
            failures.append({**item, "error": f"resource_id {rid} not found"})
            continue
        result = _msp_resource_assign_unsafe(
            task_obj=t_obj, res_obj=r_obj,
            task_id=tid, resource_id=rid,
            units=item.get("units"),
        )
        if result.get("status") == "ok":
            added.append({"task_id": tid, "resource_id": rid,
                         "assignment_uid": result.get("assignment_uid")})
        else:
            failures.append({**item, "error": result.get("error", "unknown")})
    status = "ok" if not failures else ("partial" if added else "error")
    return {"status": status, "path": path_label, "count": len(added),
            "assignments": added, "failures": failures}


def _msp_resource_bulk_assign(items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Hybrid bulk-assign: routes by item count (Phase 1 _route_operation pattern).

    Items: [{task_id, resource_id, [units]}, ...]
    Routing:
      - <=5 items   -> com_direct (no batch mode)
      - 6-19 items  -> com_batch  (batch mode + loop)
      - >=20 items  -> mspdi_bulk (Phase 2b: com_batch_fallback; true MSPDI merge = Phase 3+)

    Pre-builds task_id -> Task and resource_id -> Resource maps ONCE to avoid
    O(N×M) lookup blow-up on the HERO 14×200=2800 case.

    NOTE: ID->object maps are built fresh per call. Do not cache them across
    operations that mutate Tasks or Resources collections (add/delete) —
    map will be stale.

    Returns: {status, path, count, assignments, failures}
    """
    if not items:
        return {"status": "ok", "path": "noop", "count": 0,
                "assignments": [], "failures": []}
    app = _validate_active_project()
    proj = app.ActiveProject
    # Pre-build maps ONCE — avoids O(N×M) lookup
    task_map = _build_task_id_map(proj)
    res_map = _build_resource_id_map(proj)

    path = _route_operation(len(items))
    if path == "com_direct":
        return _msp_resource_bulk_assign_loop(items, "com_direct", task_map, res_map)
    elif path == "com_batch":
        _enter_batch_mode()
        try:
            return _msp_resource_bulk_assign_loop(items, "com_batch", task_map, res_map)
        finally:
            _exit_batch_mode()
    else:  # mspdi_bulk — Phase 2b uses com_batch_fallback (true MSPDI merge is Phase 3+)
        _enter_batch_mode()
        try:
            return _msp_resource_bulk_assign_loop(items, "mspdi_bulk", task_map, res_map)
        finally:
            _exit_batch_mode()


def _msp_resource_unassign(task_id: int, resource_id: int) -> Dict[str, Any]:
    """Remove the assignment of a resource from a task."""
    app = _validate_active_project()
    proj = app.ActiveProject
    t = _find_task_by_id(proj, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}
    # Find the matching assignment by ResourceID
    target_assignment = None
    try:
        for i in range(1, t.Assignments.Count + 1):
            a = t.Assignments(i)
            if a is not None and a.ResourceID == resource_id:
                target_assignment = a
                break
    except Exception:
        pass
    if target_assignment is None:
        return {"status": "error",
                "error": f"Resource {resource_id} not assigned to task {task_id}"}
    try:
        target_assignment.Delete()
        return {"status": "ok", "task_id": task_id, "resource_id": resource_id}
    except Exception as e:
        logger.error(f"_msp_resource_unassign({task_id},{resource_id}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


# ---------- BASELINE CONSTANTS ----------

# MSP supports 11 baseline slots: Baseline + Baseline1..Baseline10
BASELINE_NUMBERS = list(range(11))  # [0, 1, ..., 10]

# CRITICAL: app.BaselineSave's `Into` parameter uses OFFSET enum, not direct number.
# Baseline 0 → Into=0 (pjIntoBaseline); Baseline N (1-10) → Into=10+N (pjIntoBaselineN).
# (Verified from msproject_typelib.txt enum PjSaveBaselineTo.)
INTO_BASELINE_MAP = {n: (0 if n == 0 else 10 + n) for n in BASELINE_NUMBERS}

# Module-level session-level baseline name retention (TAIL #3).
# MSP COM doesn't natively persist baseline names, so we track them in-process
# for the save->list round-trip. Keys: (project_name, baseline_number).
# Values: name strings provided to _msp_baseline_save. Cleared on baseline
# clear / clear_all. Lost on Python process restart (best-effort metadata,
# matches the pattern established in T40).
_BASELINE_NAMES: Dict[Tuple[str, int], str] = {}


# ---------- BASELINE HELPERS ----------

def _baseline_property_name(field: str, baseline_number: int) -> str:
    """Map (field='Start', N=0) → 'BaselineStart'; (field='Start', N=3) → 'Baseline3Start'.

    Used to dynamically read task baseline properties for any of the 11 slots.
    """
    suffix = "" if baseline_number == 0 else str(baseline_number)
    return f"Baseline{suffix}{field}"


def _baseline_into_code(baseline_number: int) -> int:
    """Convert baseline_number (0-10) to pjIntoBaseline enum code for BaselineSave."""
    return INTO_BASELINE_MAP[baseline_number]


def _baseline_saved_date(proj: Any, baseline_number: int) -> Optional[Any]:
    """Return saved date of given baseline, or None if not saved.

    MSP returns various sentinels for unsaved baselines — normalize to Python None:
      - 0, None, "" (falsy)
      - "NA" string (MS Project 16.0 verified — most common case)
      - datetime with year < 1980 (some COM versions)
    """
    try:
        result = proj.BaselineSavedDate(Baseline=baseline_number)
        # MSP returns various falsy values for unsaved (0, datetime(year=0), None, "")
        if not result:
            return None
        # MS Project 16.0 returns the literal string "NA" for unsaved baselines
        if isinstance(result, str) and result.strip().upper() in ("NA", "N/A", ""):
            return None
        # Some COM versions return a datetime with year 1; treat year < 1980 as unsaved
        try:
            if hasattr(result, "year") and result.year < 1980:
                return None
        except Exception:
            pass
        return result
    except Exception:
        return None


def _msp_dt_or_none(v: Any) -> Optional[str]:
    """Normalize MSP baseline date sentinels ('NA', 'N/A', empty) to None.

    Tasks added AFTER a baseline was saved return the literal string 'NA' for
    BaselineNStart / BaselineNFinish. 'NA' is truthy → naive ``str(v) if v``
    leaks the sentinel downstream where date math silently yields 0 drift.

    pywintypes.datetime instances become ISO strings via ``str()``.
    """
    if not v:
        return None
    if isinstance(v, str):
        s = v.strip().upper()
        if s in ("NA", "N/A", ""):
            return None
        return v  # Already a string from MSP, return as-is
    return str(v)  # pywintypes.datetime → ISO string


def _read_task_baseline(task: Any, baseline_number: int) -> Dict[str, Any]:
    """Read a task's baseline values for the given baseline slot.

    Returns dict with start, finish, duration_h, work_h, cost.
    Unsaved baseline yields None/0 fallbacks. Each property read is guarded
    so a single bad COM read doesn't kill compare iteration.
    """
    out: Dict[str, Any] = {}
    for field, key, transform in [
        ("Start", "start", _msp_dt_or_none),
        ("Finish", "finish", _msp_dt_or_none),
        ("Duration", "duration_h", lambda v: float(v) / 60.0 if v else 0.0),
        ("Work", "work_h", lambda v: float(v) / 60.0 if v else 0.0),
        ("Cost", "cost", lambda v: _parse_rate(v)),
    ]:
        try:
            raw = getattr(task, _baseline_property_name(field, baseline_number))
            out[key] = transform(raw)
        except Exception:
            out[key] = None if field in ("Start", "Finish") else 0.0
    return out


def _msp_baseline_save(baseline_number: int = 0,
                      name: Optional[str] = None,
                      scope: str = "all",
                      roll_up_to_summary: bool = True) -> Dict[str, Any]:
    """Save current state as the specified baseline (0-10).

    scope: "all" (default — all tasks) or "selected" (currently-selected only)
    roll_up_to_summary: roll up to summary tasks (MSP default behavior)
    name: optional descriptive label (returned in metadata; MSP doesn't
          store baseline names natively in pre-2016 versions, captured here
          for caller's tracking)
    """
    if baseline_number not in BASELINE_NUMBERS:
        return {"status": "error",
                "error": f"baseline_number must be 0-10, got {baseline_number}"}
    if scope not in ("all", "selected"):
        return {"status": "error",
                "error": f"scope must be 'all' or 'selected', got '{scope}'"}
    app = _validate_active_project()
    proj = app.ActiveProject

    # Phase 1: COM mutation (atomic — abort with error on failure)
    try:
        copy_from = 0  # PjSaveBaselineFrom.pjCopyCurrent
        into = _baseline_into_code(baseline_number)
        # All=True (whole project) or False (selected); MSP COM expects bool
        all_param = (scope == "all")
        app.BaselineSave(All=all_param, Copy=copy_from, Into=into,
                        RollupToSummaryTasks=roll_up_to_summary)
    except Exception as e:
        logger.error(f"_msp_baseline_save({baseline_number}) BaselineSave failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}

    # TAIL #3: retain user-provided name for session-level list lookup.
    # If `name` is None, leave any prior retention untouched (re-saving the
    # same slot without a new name preserves the previous label — cheap memo).
    if name:
        try:
            _BASELINE_NAMES[(proj.Name, baseline_number)] = name
        except Exception:
            pass  # proj.Name read failure shouldn't fail save

    # Phase 2: metadata read-back (best-effort — save already succeeded)
    result: Dict[str, Any] = {"status": "ok",
                              "baseline_number": baseline_number,
                              "name": name}
    try:
        saved_date = _baseline_saved_date(proj, baseline_number)
        task_count = proj.Tasks.Count
        # Aggregate baseline totals (skip summary tasks)
        total_dur_min, total_work_min, total_cost = 0.0, 0.0, 0.0
        for i in range(1, task_count + 1):
            t = proj.Tasks(i)
            if t is None or t.Summary:
                continue
            data = _read_task_baseline(t, baseline_number)
            total_dur_min += (data["duration_h"] * 60) if data["duration_h"] else 0
            total_work_min += (data["work_h"] * 60) if data["work_h"] else 0
            total_cost += data["cost"] if data["cost"] else 0
        # 8h/day default — TODO: read from project calendar HoursPerDay (Phase 3b parity for CAU 9h/day)
        result.update({
            "saved_date": str(saved_date) if saved_date else None,
            "task_count": task_count,
            "total_duration_days": round(total_dur_min / 60 / 8, 2),
            "total_work_hours": round(total_work_min / 60, 2),
            "total_cost": round(total_cost, 2),  # Fix 3: rounded for caller-contract symmetry
        })
    except Exception as e:
        logger.warning(f"_msp_baseline_save({baseline_number}) metadata read-back failed (save itself succeeded): {e}")
        result["warning"] = f"metadata read failed: {_format_com_error(e)}"
    return result


def _msp_baseline_clear(baseline_number: int = 0) -> Dict[str, Any]:
    """Clear a single baseline (0-10). Idempotent: no-op if already empty.

    NOTE: BaselineClear's `From` parameter uses the PjSaveBaselineTo offset
    enum (same as BaselineSave's `Into`), NOT the direct 0-10 number.
    Empirically verified on MS Project 16.0 — passing 2 only clears baseline 0.
    Map via _baseline_into_code (B0=0, B1=11, ..., B10=20).
    """
    if baseline_number not in BASELINE_NUMBERS:
        return {"status": "error",
                "error": f"baseline_number must be 0-10, got {baseline_number}"}
    app = _validate_active_project()
    proj = app.ActiveProject
    was_saved = _baseline_saved_date(proj, baseline_number)
    try:
        from_code = _baseline_into_code(baseline_number)
        app.BaselineClear(All=True, From=from_code)
        # TAIL #3: evict retained name (if any) for this slot.
        try:
            _BASELINE_NAMES.pop((proj.Name, baseline_number), None)
        except Exception:
            pass
        return {"status": "ok",
                "baseline_number": baseline_number,
                "was_saved_date": str(was_saved) if was_saved else None}
    except Exception as e:
        logger.error(f"_msp_baseline_clear({baseline_number}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_baseline_clear_all() -> Dict[str, Any]:
    """Clear all 11 baselines that are currently saved. Returns list of cleared numbers.

    Uses the PjSaveBaselineTo offset enum for `From` (see _msp_baseline_clear note).
    Skips already-unsaved baselines (loop optimization, fewer COM calls).
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    cleared = []
    failures = []
    for n in BASELINE_NUMBERS:
        if _baseline_saved_date(proj, n) is None:
            continue
        try:
            from_code = _baseline_into_code(n)
            app.BaselineClear(All=True, From=from_code)
            cleared.append(n)
        except Exception as e:
            failures.append({"baseline_number": n, "error": _format_com_error(e)})
    # TAIL #3: evict retained names for ALL slots of this project.
    try:
        proj_name = proj.Name
        for key in [k for k in _BASELINE_NAMES if k[0] == proj_name]:
            _BASELINE_NAMES.pop(key, None)
    except Exception:
        pass
    return {"status": "ok" if not failures else "partial",
            "cleared": cleared,
            "count": len(cleared),
            "failures": failures}


def _msp_baseline_list() -> Dict[str, Any]:
    """List all 11 baseline slots; return only those currently saved with metadata.

    Iterates all 11 slots, checks saved date, includes task count + total stats
    only for saved ones. Returns sorted by baseline number (BASELINE_NUMBERS order).
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    out = []
    try:
        task_ct = proj.Tasks.Count
        for n in BASELINE_NUMBERS:
            saved = _baseline_saved_date(proj, n)
            if saved is None:
                continue
            # Aggregate per-task baseline totals (skip summaries)
            total_dur_min, total_work_min, total_cost = 0.0, 0.0, 0.0
            for i in range(1, task_ct + 1):
                t = proj.Tasks(i)
                if t is None or t.Summary:
                    continue
                data = _read_task_baseline(t, n)
                total_dur_min += (data["duration_h"] * 60) if data["duration_h"] else 0
                total_work_min += (data["work_h"] * 60) if data["work_h"] else 0
                total_cost += data["cost"] if data["cost"] else 0
            out.append({
                "number": n,
                # TAIL #3: session-level retention from _msp_baseline_save's
                # `name` arg. None when no name was ever supplied (or after
                # clear/clear_all/process restart).
                "name": _BASELINE_NAMES.get((proj.Name, n)),
                "saved_date": str(saved),
                "task_count": task_ct,
                "total_duration_days": round(total_dur_min / 60 / 8, 2),
                "total_work_hours": round(total_work_min / 60, 2),
                "total_cost": round(total_cost, 2),
            })
        return {"status": "ok", "count_saved": len(out), "baselines": out}
    except Exception as e:
        logger.error(f"_msp_baseline_list failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_baseline_get_task_baseline(task_id: int, baseline_number: int = 0) -> Dict[str, Any]:
    """Read a single task's baseline values without computing variance."""
    if baseline_number not in BASELINE_NUMBERS:
        return {"status": "error",
                "error": f"baseline_number must be 0-10, got {baseline_number}"}
    app = _validate_active_project()
    proj = app.ActiveProject
    t = _find_task_by_id(proj, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}
    try:
        data = _read_task_baseline(t, baseline_number)
        return {"status": "ok",
                "task_id": task_id,
                "baseline_number": baseline_number,
                "baseline": data}
    except Exception as e:
        logger.error(f"_msp_baseline_get_task_baseline({task_id},{baseline_number}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _datetime_diff_days(current_str: Optional[str], baseline_str: Optional[str]) -> float:
    """Compute calendar-day difference between two ISO datetime strings.
    Returns 0 if either is None/missing or unparseable (e.g. MSP 'NA' sentinel)."""
    if not current_str or not baseline_str:
        return 0.0
    try:
        from dateutil import parser
        c = parser.parse(current_str)
        b = parser.parse(baseline_str)
        return (c - b).total_seconds() / 86400.0
    except Exception:
        # Fallback: try strict ISO format
        try:
            c = _dt.datetime.fromisoformat(current_str.replace("+00:00", "").rstrip("Z"))
            b = _dt.datetime.fromisoformat(baseline_str.replace("+00:00", "").rstrip("Z"))
            return (c - b).total_seconds() / 86400.0
        except Exception:
            return 0.0


# Module-level constant for floating-point comparison (TAIL #4)
_VARIANCE_EPSILON = 1e-9  # ~picoseconds in days, ~picohours, ~picocents


def _read_task_current_state(task: Any) -> Dict[str, Any]:
    """Read live task fields in the same shape as `_read_task_baseline`.

    Adapter so `_compute_variance_set` can treat current state and baseline
    state uniformly. Uses `_msp_dt_or_none` for Start/Finish so 'NA' sentinels
    (impossible on live tasks but cheap defensive guard) become None.
    """
    return {
        "start": _msp_dt_or_none(task.Start) if task.Start else None,
        "finish": _msp_dt_or_none(task.Finish) if task.Finish else None,
        "duration_h": float(task.Duration) / 60.0 if task.Duration else 0.0,
        "work_h": float(task.Work) / 60.0 if task.Work else 0.0,
        "cost": _parse_rate(task.Cost) if task.Cost else 0.0,
    }


def _compute_variance_set(
    proj: Any,
    get_a: Callable[[Any], Dict[str, Any]],
    get_b: Callable[[Any], Dict[str, Any]],
    include_unchanged: bool,
    variance_threshold_days: float,
) -> Dict[str, Any]:
    """Compute per-task variance summary + list given two side-readers.

    Variance direction: `b - a` (positive = side B exceeds side A).
    For `_msp_baseline_compare`: a = saved baseline, b = current state →
    positive finish_var = slipped.
    For `_msp_baseline_compare_two`: a = baseline_a (e.g. Original),
    b = baseline_b (e.g. Revised) → positive finish_var = slip between
    revisions.

    Pre-builds the non-summary task list once (TAIL #1 perf) and uses
    EPSILON-based no_change check (TAIL #4 robustness).

    Returns {"summary": {...8 keys...}, "tasks": [...]}; caller wraps with
    status/baseline_number etc.
    """
    # Pre-build task cache ONCE (TAIL #1) — cuts ~50% COM dispatches on Summary check
    real_tasks: List[Any] = []
    for i in range(1, proj.Tasks.Count + 1):
        t = proj.Tasks(i)
        if t is not None and not t.Summary:
            real_tasks.append(t)

    tasks_var: List[Dict[str, Any]] = []
    slipped_ct = ahead_ct = on_time_ct = 0
    total_start_drift = total_finish_drift = 0.0
    total_dur_var_h = total_work_var_h = 0.0
    total_cost_var = 0.0

    for t in real_tasks:
        a = get_a(t)
        b = get_b(t)
        start_var = _datetime_diff_days(b["start"], a["start"])
        finish_var = _datetime_diff_days(b["finish"], a["finish"])
        dur_var = (b["duration_h"] or 0) - (a["duration_h"] or 0)
        work_var = (b["work_h"] or 0) - (a["work_h"] or 0)
        cost_var = (b["cost"] or 0) - (a["cost"] or 0)

        if finish_var > variance_threshold_days:
            status = "slipped"
            slipped_ct += 1
        elif finish_var < -variance_threshold_days:
            status = "ahead"
            ahead_ct += 1
        else:
            status = "on_time"
            on_time_ct += 1

        total_start_drift += start_var
        total_finish_drift += finish_var
        total_dur_var_h += dur_var
        total_work_var_h += work_var
        total_cost_var += cost_var

        # EPSILON-based no_change (TAIL #4) — robust to float residue
        no_change = (
            abs(start_var) < _VARIANCE_EPSILON
            and abs(finish_var) < _VARIANCE_EPSILON
            and abs(dur_var) < _VARIANCE_EPSILON
            and abs(work_var) < _VARIANCE_EPSILON
            and abs(cost_var) < _VARIANCE_EPSILON
        )
        if not include_unchanged and no_change:
            continue
        tasks_var.append({
            "id": t.ID,
            "name": t.Name,
            "start_var_days": round(start_var, 2),
            "finish_var_days": round(finish_var, 2),
            "duration_var_h": round(dur_var, 2),
            "work_var_h": round(work_var, 2),
            "cost_var": round(cost_var, 2),
            "status": status,
        })

    return {
        "summary": {
            "slipped_count": slipped_ct,
            "ahead_count": ahead_ct,
            "on_time_count": on_time_ct,
            "total_start_drift_days": round(total_start_drift, 2),
            "total_finish_drift_days": round(total_finish_drift, 2),
            "total_duration_var_h": round(total_dur_var_h, 2),
            "total_work_var_h": round(total_work_var_h, 2),
            "total_cost_var": round(total_cost_var, 2),
        },
        "tasks": tasks_var,
    }


def _msp_baseline_compare(baseline_number: int = 0,
                         include_unchanged: bool = False,
                         variance_threshold_days: float = 0.0) -> Dict[str, Any]:
    """Compare current state vs saved baseline. Returns summary + per-task variance.

    variance_threshold_days: tasks with |finish_var_days| <= threshold count as on_time.
    include_unchanged: if False, omit tasks with 0 variance from per-task list.
    """
    if baseline_number not in BASELINE_NUMBERS:
        return {"status": "error",
                "error": f"baseline_number must be 0-10, got {baseline_number}"}
    app = _validate_active_project()
    proj = app.ActiveProject
    if _baseline_saved_date(proj, baseline_number) is None:
        return {"status": "error",
                "error": f"Baseline {baseline_number} not saved (no baseline data to compare against)"}
    try:
        result = _compute_variance_set(
            proj,
            get_a=lambda t: _read_task_baseline(t, baseline_number),
            get_b=_read_task_current_state,
            include_unchanged=include_unchanged,
            variance_threshold_days=variance_threshold_days,
        )
        return {
            "status": "ok",
            "baseline_number": baseline_number,
            **result,
        }
    except Exception as e:
        logger.error(f"_msp_baseline_compare({baseline_number}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_baseline_compare_two(baseline_a: int,
                              baseline_b: int,
                              include_unchanged: bool = False,
                              variance_threshold_days: float = 0.0) -> Dict[str, Any]:
    """Compare two saved baselines as a delta (baseline_b - baseline_a).

    Variance is computed as (baseline_b - baseline_a): the change FROM
    baseline_a TO baseline_b. Use case: baseline_a=0 (Original Plan),
    baseline_b=1 (Revised Plan) → positive finish_var_days = task slipped
    between revisions; negative = task pulled ahead between revisions.

    Both baselines must be saved (have a baseline_save_date). Per-task fields
    with None on either side (e.g. tasks added after a save) yield 0 from
    _datetime_diff_days, matching the compare-action contract.

    Args:
        baseline_a: 0-10, the "earlier" baseline slot to subtract.
        baseline_b: 0-10, the "later" baseline slot.
        include_unchanged: if False, omit zero-variance tasks from the list.
        variance_threshold_days: tasks with |finish_var_days| <= threshold
            count as on_time.
    """
    for label, n in [("baseline_a", baseline_a), ("baseline_b", baseline_b)]:
        if n not in BASELINE_NUMBERS:
            return {"status": "error",
                    "error": f"{label} must be 0-10, got {n}"}
    app = _validate_active_project()
    proj = app.ActiveProject
    if _baseline_saved_date(proj, baseline_a) is None:
        return {"status": "error",
                "error": f"baseline_a ({baseline_a}) is not saved"}
    if _baseline_saved_date(proj, baseline_b) is None:
        return {"status": "error",
                "error": f"baseline_b ({baseline_b}) is not saved"}
    try:
        result = _compute_variance_set(
            proj,
            get_a=lambda t: _read_task_baseline(t, baseline_a),
            get_b=lambda t: _read_task_baseline(t, baseline_b),
            include_unchanged=include_unchanged,
            variance_threshold_days=variance_threshold_days,
        )
        return {
            "status": "ok",
            "baseline_a": baseline_a,
            "baseline_b": baseline_b,
            **result,
        }
    except Exception as e:
        logger.error(f"_msp_baseline_compare_two({baseline_a},{baseline_b}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_baseline_summary(baseline_number: int = 0) -> Dict[str, Any]:
    """Project-level RAG status from baseline comparison.

    RAG thresholds:
      green  : slipped_pct <= 5
      amber  : 5 < slipped_pct <= 20
      red    : slipped_pct > 20
    """
    cmp_result = _msp_baseline_compare(baseline_number=baseline_number,
                                       include_unchanged=True)
    if cmp_result.get("status") != "ok":
        return cmp_result  # propagate error
    s = cmp_result["summary"]
    total_tasks = s["slipped_count"] + s["ahead_count"] + s["on_time_count"]
    slipped_pct = (s["slipped_count"] / total_tasks * 100.0) if total_tasks > 0 else 0.0
    if slipped_pct <= 5:
        health = "green"
    elif slipped_pct <= 20:
        health = "amber"
    else:
        health = "red"
    return {
        "status": "ok",
        "baseline_number": baseline_number,
        "project": {
            "task_count": total_tasks,
            "slipped_count": s["slipped_count"],
            "slipped_pct": round(slipped_pct, 2),
            "start_drift_days": s["total_start_drift_days"],
            "finish_drift_days": s["total_finish_drift_days"],
            "schedule_health": health,
        },
    }


def _msp_baseline_set_active(baseline_number: int) -> Dict[str, Any]:
    """Set the active baseline for views/EVM calculations.

    The "active baseline" controls which Baseline*N* fields the Earned Value
    engine and baseline-aware views use. T47 probe (MSP 16.0) confirmed that
    proj.EarnedValueBaseline is the real read/write property — round-trips
    exactly. Other candidates (BaselineForEarnedValue, ActiveBaselineNumber)
    silently swallow setattr without persisting on this MSP build.

    NOTE: MSP COM API for this is version-dependent. Older MSP versions may
    not expose EarnedValueBaseline; in that case the function returns a
    graceful "not yet supported" error — saved baseline data is still readable
    via get_task_baseline / compare regardless of which is "active".
    """
    if baseline_number not in BASELINE_NUMBERS:
        return {"status": "error",
                "error": f"baseline_number must be 0-10, got {baseline_number}"}
    app = _validate_active_project()
    proj = app.ActiveProject
    # Probe-confirmed property: proj.EarnedValueBaseline (read/write int, round-trips).
    try:
        proj.EarnedValueBaseline = baseline_number
        readback = proj.EarnedValueBaseline
        if readback == baseline_number:
            return {"status": "ok", "active_baseline": baseline_number,
                    "method": "proj.EarnedValueBaseline"}
    except Exception as e:
        logger.debug(f"_msp_baseline_set_active: EarnedValueBaseline failed: {e}")
    return {"status": "error",
            "error": ("set_active is not yet supported on this MS Project version. "
                     "Use msproject_baseline compare/summary directly with the "
                     "baseline_number parameter — they don't require setting an active baseline.")}


# ==================== PHASE 3b — PROGRESS MANAGEMENT ====================
# Tool: msproject_progress (12 actions). Insertion point: between Phase 3a
# baseline section and Phase 1 _msp_task_update.
# Builds on: _validate_active_project, _format_com_error, _parse_rate,
# _find_task_by_id, _find_resource_by_id, _msp_dt_or_none, _route_operation,
# _enter_batch_mode, _exit_batch_mode, _build_task_id_map.

# ---------- PROGRESS CONSTANTS ----------

_PROGRESS_PCT_FIELDS = frozenset({
    "percent_complete",
    "percent_work_complete",
    "physical_pct",
})

_PROGRESS_WORK_FIELDS = frozenset({
    "actual_work_h",
    "remaining_work_h",
})

_PROGRESS_DURATION_FIELDS = frozenset({
    "actual_duration_h",
    "remaining_duration_h",
})

_PROGRESS_DATE_FIELDS = frozenset({
    "actual_start",
    "actual_finish",
    "stop",
    "resume",
})

# Probe-confirmed (msproject_typelib.txt + MSP 16.0 round-trip):
_TIMESCALE_UNIT_MAP = {
    "day": 4,    # pjTimescaleDays (probe-confirmed MSP 2024 — T60)
    "week": 3,   # pjTimescaleWeeks (probe-confirmed MSP 2024 — T60)
}

_PJ_TIMESCALED_ACTUAL_WORK = 10  # pjAssignmentTimescaledActualWork (probe-confirmed MSP 2024 — T60)


# ---------- PROGRESS HELPERS ----------

def _normalize_progress_pct(v: Any) -> float:
    """Validate + normalize a percentage value (0-100).

    Accepts int / float / str ('50', '50.5', '50%'). Raises ValueError on
    out-of-range or non-numeric input. Returns float rounded to 2 decimals.

    Note: ``bool`` is rejected explicitly. In Python ``bool`` is a subclass of
    ``int`` so ``True`` would otherwise coerce to 1.0 and ``False`` to 0.0 —
    JSON inputs like ``{"physical_pct": true}`` would silently produce 1%
    progress. This guard catches that class of caller bug at the boundary.
    """
    if v is None:
        raise ValueError("progress percentage cannot be None")
    if isinstance(v, bool):
        raise ValueError(
            f"progress percentage must be numeric, not bool: {v!r}"
        )
    if isinstance(v, str):
        s = v.strip().rstrip("%").strip()
        if not s:
            raise ValueError(f"progress percentage is empty string: {v!r}")
        try:
            f = float(s)
        except Exception:
            raise ValueError(f"progress percentage not numeric: {v!r}")
    else:
        try:
            f = float(v)
        except Exception:
            raise ValueError(f"progress percentage not numeric: {v!r}")
    if f < 0 or f > 100:
        raise ValueError(f"progress percentage must be 0-100, got {f}")
    return round(f, 2)


def _hours_to_minutes(h: float) -> int:
    """Convert public-API hours to MSP COM minutes (rounded int)."""
    return int(round(float(h) * 60))


def _minutes_to_hours(m: Any) -> float:
    """Convert MSP COM minutes to public-API hours (float, 2 decimals)."""
    if m is None:
        return 0.0
    try:
        return round(float(m) / 60.0, 2)
    except Exception:
        return 0.0


def _validate_actual_dates(start: Optional[str],
                          finish: Optional[str]) -> Optional[str]:
    """Verify actual_start <= actual_finish if both provided.

    Returns ``None`` if valid, or an error message string if invalid.

    Date format: ISO 8601 strict (``datetime.fromisoformat``-compatible).
    Falls back to ``dateutil.parser.parse`` with ``dayfirst=False`` for
    non-ISO inputs (e.g. the ``str()`` form of ``pywintypes.datetime``).

    Format assumption: ambiguous slash-separated dates like ``01/02/2026`` are
    interpreted as **US (Jan 2)**, never EU. Public callers should always pass
    ISO 8601 to avoid surprise.
    """
    if start is None or finish is None:
        return None

    def _parse(s: Any) -> "_dt.datetime":
        s = str(s).strip()
        if not s:
            raise ValueError("empty date string")
        try:
            return _dt.datetime.fromisoformat(
                s.replace("+00:00", "").rstrip("Z")
            )
        except Exception:
            from dateutil import parser
            return parser.parse(s, dayfirst=False)

    try:
        s_dt = _parse(start)
        f_dt = _parse(finish)
    except Exception as e:
        return f"could not parse dates ({start!r}, {finish!r}): {e}"
    if s_dt > f_dt:
        return f"actual_start ({start}) must be <= actual_finish ({finish})"
    return None


def _get_assignment_by_resource_id(task: Any, resource_id: int) -> Optional[Any]:
    """Find an assignment on the task matching the given resource_id.

    Iterates task.Assignments (1-indexed COM collection). Returns Assignment
    object or None.
    """
    try:
        for i in range(1, task.Assignments.Count + 1):
            try:
                asg = task.Assignments(i)
                if asg is None:
                    continue
                if int(asg.ResourceID) == int(resource_id):
                    return asg
            except Exception:
                continue
    except Exception as e:
        logger.debug(f"_get_assignment_by_resource_id iter failed: {e}")
    return None


def _read_task_progress_dict(task: Any) -> Dict[str, Any]:
    """Read all progress fields from a task → dict shaped for get_task_progress.

    Each property is guarded; failure on one field yields safe default.
    """
    out: Dict[str, Any] = {}
    # Defaults are 0.0 (float) for type stability — successful reads also use
    # float(v), so JSON output remains uniform whether the read succeeded or
    # fell back to the default.
    pct_pairs = [
        ("PercentComplete", "percent_complete", 0.0),
        ("PercentWorkComplete", "percent_work_complete", 0.0),
        ("PhysicalPercentComplete", "physical_pct", 0.0),
    ]
    for prop, key, default in pct_pairs:
        try:
            v = getattr(task, prop)
            out[key] = float(v) if v is not None else default
        except Exception:
            out[key] = default
    date_pairs = [
        ("ActualStart", "actual_start"),
        ("ActualFinish", "actual_finish"),
        ("Stop", "stop"),
        ("Resume", "resume"),
    ]
    for prop, key in date_pairs:
        try:
            out[key] = _msp_dt_or_none(getattr(task, prop))
        except Exception:
            out[key] = None
    work_pairs = [
        ("ActualWork", "actual_work_h"),
        ("RemainingWork", "remaining_work_h"),
        ("ActualDuration", "actual_duration_h"),
        ("RemainingDuration", "remaining_duration_h"),
    ]
    for prop, key in work_pairs:
        try:
            out[key] = _minutes_to_hours(getattr(task, prop))
        except Exception:
            out[key] = 0.0
    return out


def _to_pywintypes_date(v: Any) -> Any:
    """Convert public-API date input to MSP COM-compatible pywintypes.Time.

    MSP COM rejects raw strings on date properties (``task.ActualStart = "..."``
    raises an opaque COM error). Public callers JSON-encode dates as ISO
    strings, so this helper bridges the two.

    Accepts:
      * ``None`` → returns ``None`` (caller passes through to clear / no-op)
      * ``pywintypes.datetime`` / ``pywintypes.TimeType`` → returned as-is
      * ``datetime.datetime`` → wrapped via ``pywintypes.Time``
      * ISO 8601 string (``2026-04-15`` or ``2026-04-15 08:00:00``) parsed
        with ``datetime.fromisoformat``; trailing ``Z`` and ``+00:00`` stripped
      * Non-ISO strings → fall back to ``dateutil.parser.parse`` with
        ``dayfirst=False`` so ambiguous ``01/02/2026`` is deterministic (Jan 2)

    Raises ``ValueError`` on empty string, unparseable string, or unsupported
    type.
    """
    if v is None:
        return None
    # pywintypes.datetime / pywintypes.Time → pass through unchanged
    try:
        import pywintypes
        if isinstance(v, pywintypes.TimeType):
            return v
    except ImportError:
        pass
    # datetime.datetime → wrap in pywintypes.Time
    if isinstance(v, _dt.datetime):
        import pywintypes
        return pywintypes.Time(v)
    # ISO 8601 string (preferred)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            raise ValueError("date input is empty string")
        try:
            dt = _dt.datetime.fromisoformat(
                s.replace("+00:00", "").rstrip("Z")
            )
        except Exception:
            try:
                from dateutil import parser
                dt = parser.parse(s, dayfirst=False)
            except Exception as e:
                raise ValueError(f"could not parse date {v!r}: {e}")
        import pywintypes
        return pywintypes.Time(dt)
    raise ValueError(f"unsupported date type: {type(v).__name__} ({v!r})")


def _msp_task_set_progress_field(task: Any, field: str, value: Any) -> None:
    """Low-level setter that maps public field names → MSP COM properties.

    Raises on COM error; caller wraps in try/except per-field for partial-failure
    aggregation.
    """
    if field == "percent_complete":
        task.PercentComplete = _normalize_progress_pct(value)
    elif field == "percent_work_complete":
        task.PercentWorkComplete = _normalize_progress_pct(value)
    elif field == "physical_pct":
        task.PhysicalPercentComplete = _normalize_progress_pct(value)
    elif field == "actual_start":
        task.ActualStart = _to_pywintypes_date(value)
    elif field == "actual_finish":
        task.ActualFinish = _to_pywintypes_date(value)
    elif field == "actual_duration_h":
        task.ActualDuration = _hours_to_minutes(value)
    elif field == "actual_work_h":
        task.ActualWork = _hours_to_minutes(value)
    elif field == "remaining_work_h":
        task.RemainingWork = _hours_to_minutes(value)
    elif field == "remaining_duration_h":
        task.RemainingDuration = _hours_to_minutes(value)
    elif field == "stop":
        task.Stop = _to_pywintypes_date(value)
    elif field == "resume":
        task.Resume = _to_pywintypes_date(value)
    else:
        raise ValueError(f"Unknown progress field: {field}")


def _msp_progress_set_task(task_id: int,
                           percent_complete: Optional[float] = None,
                           percent_work_complete: Optional[float] = None,
                           actual_start: Optional[str] = None,
                           actual_finish: Optional[str] = None,
                           actual_duration_h: Optional[float] = None,
                           actual_work_h: Optional[float] = None,
                           remaining_work_h: Optional[float] = None,
                           remaining_duration_h: Optional[float] = None,
                           physical_pct: Optional[float] = None,
                           stop: Optional[str] = None,
                           resume: Optional[str] = None) -> Dict[str, Any]:
    """Set one or more progress fields on a task.

    Dual-mode: caller can pass `percent_complete` (duration-based) OR explicit
    actual_*/remaining_* values. Both modes coexist; MSP rebalances internally.

    Phase 3b — see design doc Section 6 Q1 (dual-track) and Q4 (PhysicalPercentComplete).
    Returns {status, task_id, changes: [field], readback: dict}.
    """
    # Build candidate field map; filter out None
    candidates = {
        "percent_complete": percent_complete,
        "percent_work_complete": percent_work_complete,
        "actual_start": actual_start,
        "actual_finish": actual_finish,
        "actual_duration_h": actual_duration_h,
        "actual_work_h": actual_work_h,
        "remaining_work_h": remaining_work_h,
        "remaining_duration_h": remaining_duration_h,
        "physical_pct": physical_pct,
        "stop": stop,
        "resume": resume,
    }
    fields_to_set = {k: v for k, v in candidates.items() if v is not None}
    if not fields_to_set:
        return {"status": "error", "error": "no progress fields provided"}

    # Pre-validate pct fields and date order
    for f in _PROGRESS_PCT_FIELDS:
        if f in fields_to_set:
            try:
                fields_to_set[f] = _normalize_progress_pct(fields_to_set[f])
            except ValueError as e:
                return {"status": "error", "error": str(e)}
    err = _validate_actual_dates(fields_to_set.get("actual_start"),
                                 fields_to_set.get("actual_finish"))
    if err:
        return {"status": "error", "error": err}
    err = _validate_actual_dates(fields_to_set.get("stop"),
                                 fields_to_set.get("resume"))
    if err:
        return {"status": "error", "error": "stop/resume order: " + err}

    app = _validate_active_project()
    proj = app.ActiveProject
    t = _find_task_by_id(proj, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}

    changes: List[str] = []
    failures: List[Dict[str, str]] = []
    for field, value in fields_to_set.items():
        try:
            _msp_task_set_progress_field(t, field, value)
            changes.append(field)
        except Exception as e:
            failures.append({"field": field, "error": _format_com_error(e)})
            logger.debug(f"set_task_progress({task_id}, {field}={value}) failed: {e}")

    try:
        readback = _read_task_progress_dict(t)
    except Exception:
        readback = None

    if not changes and failures:
        return {"status": "error", "task_id": task_id,
                "error": "all field writes failed", "failures": failures}
    status = "ok" if not failures else "partial"
    return {"status": status, "task_id": task_id,
            "changes": changes, "failures": failures,
            "readback": readback}


def _msp_progress_get_task(task_id: int) -> Dict[str, Any]:
    """Read all progress fields from a task → structured dict.

    Returns {status, task_id, progress: {percent_complete, percent_work_complete,
    actual_start, actual_finish, actual_duration_h, actual_work_h,
    remaining_work_h, remaining_duration_h, physical_pct, stop, resume}}.
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    t = _find_task_by_id(proj, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}
    try:
        progress = _read_task_progress_dict(t)
        return {"status": "ok", "task_id": task_id, "progress": progress}
    except Exception as e:
        logger.error(f"_msp_progress_get_task({task_id}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_progress_set_assignment(task_id: int,
                                 resource_id: int,
                                 actual_work_h: Optional[float] = None,
                                 actual_start: Optional[str] = None,
                                 actual_finish: Optional[str] = None,
                                 percent_work_complete: Optional[float] = None,
                                 remaining_work_h: Optional[float] = None,
                                 units: Optional[float] = None) -> Dict[str, Any]:
    """Per-resource man-hour write on a single assignment.

    Phase 3b — for hakediş workflows where each resource's hours are tracked
    separately (e.g., "T101: COW=24h, STL=18h, MSN=10h").

    MSP rolls up assignment-level writes to task.ActualWork automatically.
    Returns {status, task_id, resource_id, changes: [field]}.
    """
    candidates = {
        "actual_work_h": actual_work_h,
        "actual_start": actual_start,
        "actual_finish": actual_finish,
        "percent_work_complete": percent_work_complete,
        "remaining_work_h": remaining_work_h,
        "units": units,
    }
    fields = {k: v for k, v in candidates.items() if v is not None}
    if not fields:
        return {"status": "error", "error": "no assignment fields provided"}

    # Pre-validate pct
    if "percent_work_complete" in fields:
        try:
            fields["percent_work_complete"] = _normalize_progress_pct(
                fields["percent_work_complete"])
        except ValueError as e:
            return {"status": "error", "error": str(e)}
    err = _validate_actual_dates(fields.get("actual_start"),
                                 fields.get("actual_finish"))
    if err:
        return {"status": "error", "error": err}

    app = _validate_active_project()
    proj = app.ActiveProject
    t = _find_task_by_id(proj, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}
    asg = _get_assignment_by_resource_id(t, resource_id)
    if asg is None:
        return {"status": "error",
                "error": f"No assignment for resource_id {resource_id} on task {task_id}"}

    changes: List[str] = []
    failures: List[Dict[str, str]] = []
    for field, value in fields.items():
        try:
            if field == "actual_work_h":
                asg.ActualWork = _hours_to_minutes(value)
            elif field == "actual_start":
                asg.ActualStart = _to_pywintypes_date(value)  # T52 fix integration
            elif field == "actual_finish":
                asg.ActualFinish = _to_pywintypes_date(value)  # T52 fix integration
            elif field == "percent_work_complete":
                asg.PercentWorkComplete = value
            elif field == "remaining_work_h":
                asg.RemainingWork = _hours_to_minutes(value)
            elif field == "units":
                asg.Units = float(value)
            else:
                raise ValueError(f"unknown assignment field: {field}")
            changes.append(field)
        except Exception as e:
            failures.append({"field": field, "error": _format_com_error(e)})
            logger.debug(
                f"set_assignment_progress({task_id},{resource_id},{field}={value}) failed: {e}")

    if not changes and failures:
        return {"status": "error", "task_id": task_id, "resource_id": resource_id,
                "error": "all assignment field writes failed", "failures": failures}
    status = "ok" if not failures else "partial"
    return {"status": status, "task_id": task_id, "resource_id": resource_id,
            "changes": changes, "failures": failures}


def _msp_progress_get_assignments(task_id: int) -> Dict[str, Any]:
    """List all per-resource progress on a task.

    Returns {status, task_id, assignments: [{resource_id, resource_name,
    actual_work_h, percent_work_complete, remaining_work_h, units,
    actual_start, actual_finish}, ...]}.
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    t = _find_task_by_id(proj, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}
    out: List[Dict[str, Any]] = []
    try:
        for i in range(1, t.Assignments.Count + 1):
            try:
                asg = t.Assignments(i)
                if asg is None:
                    continue
                try:
                    res = asg.Resource
                    res_name = res.Name if res is not None else None
                except Exception:
                    res_name = None
                out.append({
                    "resource_id": int(asg.ResourceID),
                    "resource_name": res_name,
                    "actual_work_h": _minutes_to_hours(asg.ActualWork),
                    "actual_start": _msp_dt_or_none(asg.ActualStart),
                    "actual_finish": _msp_dt_or_none(asg.ActualFinish),
                    "percent_work_complete": float(asg.PercentWorkComplete or 0),
                    "remaining_work_h": _minutes_to_hours(asg.RemainingWork),
                    "units": float(asg.Units or 0),
                })
            except Exception as e:
                logger.debug(f"_msp_progress_get_assignments row {i} failed: {e}")
                continue
        return {"status": "ok", "task_id": task_id, "assignments": out}
    except Exception as e:
        logger.error(f"_msp_progress_get_assignments({task_id}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_progress_set_by_date(progress_date: Any,
                              scope: str = "all",
                              as_scheduled: bool = True) -> Dict[str, Any]:
    """Bulk-update progress to a given date (``app.UpdateProject``).

    Implements the "plan = actual" up-to-data-date assumption — fast retroactive
    backlog catch-up. Phase 3b — see design doc Section 6 (Q1) and Open
    Questions #2 for already-progressed task interaction.

    Probe-confirmed signature on MSP 16.0::

        UpdateProject(All, UpdateDate, action)

    where ``All`` (bool): True = all tasks / False = currently-selected only;
    ``UpdateDate`` (datetime.datetime; **not** pywintypes.Time — COM rejects);
    ``action`` (int): 1 = update progress as scheduled (default), 0 =
    reschedule incomplete only (no progress write), 2 = no-op.

    Args:
        progress_date: ISO 8601 string or ``datetime.datetime``. Empty / unparseable
                       string returns ``{status: error}``.
        scope: ``"all"`` (entire project) or ``"selected"`` (currently selected
               tasks in MSP UI).
        as_scheduled: True → action=1 (write progress as scheduled up to date);
                      False → action=0 (reschedule incomplete portion only).
                      Note: MSP COM ``UpdateProject`` does not expose a pure
                      "% complete only" mode. ``as_scheduled=False`` here means
                      reschedule-incomplete (lighter touch — no actuals written).

    Returns:
        ``{status, progress_date, mode, scope, task_count_affected}`` on success;
        ``{status: error, error}`` on parse / COM failure.
    """
    if scope not in ("all", "selected"):
        return {"status": "error",
                "error": f"scope must be 'all' or 'selected', got '{scope}'"}
    # Parse progress_date → plain datetime.datetime (NOT pywintypes — COM rejects)
    try:
        if isinstance(progress_date, _dt.datetime):
            pd = progress_date
        elif isinstance(progress_date, str):
            s = progress_date.strip()
            if not s:
                return {"status": "error",
                        "error": "progress_date is empty string"}
            try:
                pd = _dt.datetime.fromisoformat(
                    s.replace("+00:00", "").rstrip("Z")
                )
            except Exception:
                from dateutil import parser
                pd = parser.parse(s, dayfirst=False)
        else:
            return {"status": "error",
                    "error": f"unsupported progress_date type: "
                             f"{type(progress_date).__name__}"}
    except Exception as e:
        return {"status": "error",
                "error": f"could not parse progress_date {progress_date!r}: {e}"}

    app = _validate_active_project()
    proj = app.ActiveProject
    task_count_before = proj.Tasks.Count
    try:
        all_tasks_flag = (scope == "all")
        action_val = 1 if as_scheduled else 0  # 1 = update as scheduled; 0 = reschedule incomplete only
        # Probe-confirmed signature on MSP 16.0: positional only, plain datetime.
        app.UpdateProject(all_tasks_flag, pd, action_val)
        return {"status": "ok",
                "progress_date": pd.isoformat(),
                "mode": "as_scheduled" if as_scheduled else "reschedule_incomplete",
                "scope": scope,
                "task_count_affected": task_count_before}
    except Exception as e:
        logger.error(f"_msp_progress_set_by_date({progress_date}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_progress_set_status_date(status_date: str) -> Dict[str, Any]:
    """Set proj.StatusDate (CLAUDE.md RULE 5 'data_date').

    Used by EVM tooling (Phase 5) and time-phased PV/EV calculations.
    """
    try:
        from dateutil import parser
        sd = parser.parse(str(status_date))
    except Exception as e:
        return {"status": "error",
                "error": f"could not parse status_date {status_date!r}: {e}"}
    app = _validate_active_project()
    proj = app.ActiveProject
    try:
        previous = _msp_dt_or_none(proj.StatusDate)
    except Exception:
        previous = None
    try:
        proj.StatusDate = sd
        return {"status": "ok",
                "status_date": str(sd),
                "previous": previous}
    except Exception as e:
        logger.error(f"_msp_progress_set_status_date({status_date}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_progress_clear(task_id: int) -> Dict[str, Any]:
    """Reset a single task's progress to 0/None across all progress fields.

    Idempotent: tasks that are already at 0% return ok with cleared_fields=[].
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    t = _find_task_by_id(proj, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}
    cleared: List[str] = []
    try:
        # Setting PercentComplete = 0 cascades to ActualStart/Finish/Work clearing
        # in MSP, but we explicitly reset each for safety.
        if t.PercentComplete and t.PercentComplete > 0:
            t.PercentComplete = 0
            cleared.append("percent_complete")
        if t.ActualWork and t.ActualWork > 0:
            t.ActualWork = 0
            cleared.append("actual_work")
        # ActualStart / ActualFinish: write "NA" sentinel to clear (MSP convention)
        try:
            t.ActualStart = "NA"
            cleared.append("actual_start")
        except Exception:
            pass
        try:
            t.ActualFinish = "NA"
            cleared.append("actual_finish")
        except Exception:
            pass
        return {"status": "ok", "task_id": task_id, "cleared_fields": cleared}
    except Exception as e:
        logger.error(f"_msp_progress_clear({task_id}) failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_progress_clear_all() -> Dict[str, Any]:
    """Clear progress on every non-summary task in the project.

    Used to reset a project to "as-planned" state.
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    cleared_count = 0
    failures: List[Dict[str, Any]] = []
    try:
        _enter_batch_mode()
        try:
            for i in range(1, proj.Tasks.Count + 1):
                try:
                    t = proj.Tasks(i)
                    if t is None or t.Summary:
                        continue
                    if t.PercentComplete and t.PercentComplete > 0:
                        t.PercentComplete = 0
                        cleared_count += 1
                    elif t.ActualWork and t.ActualWork > 0:
                        t.ActualWork = 0
                        cleared_count += 1
                except Exception as e:
                    failures.append({"index": i, "error": _format_com_error(e)})
        finally:
            _exit_batch_mode()
        return {"status": "ok" if not failures else "partial",
                "cleared_count": cleared_count,
                "failures": failures}
    except Exception as e:
        logger.error(f"_msp_progress_clear_all failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_progress_time_phased_write(task_id: int,
                                    resource_id: int,
                                    periods: List[Dict[str, Any]],
                                    unit: str = "day") -> Dict[str, Any]:
    """Write per-period actual_work to an assignment via TimeScaleData.

    Phase 3b T60 — see design doc Section 6 Q2. Granularity: 'day' or 'week'.

    periods: [{start: ISO, end: ISO, actual_work_h: float}, ...]
    Each period maps to one (or more) TimeScaleValues slots; write fails per-
    slot if MSP doesn't have a matching cell. Failures aggregate into
    return['failures'] without raising. COM signature uses positional args
    (StartDate, EndDate, Type, TimeScaleUnit) — note capital S in TimeScaleUnit
    per probe-confirmed signature on MSP 2024.
    """
    if unit not in _TIMESCALE_UNIT_MAP:
        return {"status": "error",
                "error": f"unit must be 'day' or 'week', got '{unit}'"}
    if not isinstance(periods, list) or not periods:
        return {"status": "error", "error": "periods must be a non-empty list"}
    unit_code = _TIMESCALE_UNIT_MAP[unit]

    app = _validate_active_project()
    proj = app.ActiveProject
    t = _find_task_by_id(proj, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}
    asg = _get_assignment_by_resource_id(t, resource_id)
    if asg is None:
        return {"status": "error",
                "error": f"No assignment for resource_id {resource_id} on task {task_id}"}

    from dateutil import parser
    written = 0
    failures: List[Dict[str, Any]] = []
    for idx, p in enumerate(periods):
        try:
            ps = parser.parse(str(p["start"]))
            pe = parser.parse(str(p["end"]))
            hours = float(p["actual_work_h"])
        except Exception as e:
            failures.append({"index": idx, "period": p,
                             "error": f"parse failed: {e}"})
            continue
        try:
            # Positional call: (StartDate, EndDate, Type, TimeScaleUnit)
            tsv = asg.TimeScaleData(ps, pe,
                                    _PJ_TIMESCALED_ACTUAL_WORK,
                                    unit_code)
            if tsv.Count == 0:
                failures.append({"index": idx, "period": p,
                                 "error": "no time slots in range (assignment "
                                          "may not span this date)"})
                continue
            # Filter slots to those that fall WITHIN the requested period —
            # MSP returns boundary slots (e.g. the day before a 1-day query)
            # that should be skipped to honor caller intent.
            target_indices: List[int] = []
            for i in range(1, tsv.Count + 1):
                try:
                    item = tsv.Item(i)
                    item_start = item.StartDate
                    # Compare naive vs tz-aware safely by stripping tz
                    if hasattr(item_start, "tzinfo") and item_start.tzinfo is not None:
                        item_start_cmp = item_start.replace(tzinfo=None)
                    else:
                        item_start_cmp = item_start
                    ps_cmp = ps.replace(tzinfo=None) if ps.tzinfo else ps
                    pe_cmp = pe.replace(tzinfo=None) if pe.tzinfo else pe
                    if item_start_cmp >= ps_cmp and item_start_cmp < pe_cmp:
                        target_indices.append(i)
                except Exception:
                    continue
            # Fallback: if filter excluded everything, write to all slots so
            # something lands (out-of-range edge case).
            if not target_indices:
                target_indices = list(range(1, tsv.Count + 1))
            # Distribute hours across selected slots evenly
            minutes_total = _hours_to_minutes(hours)
            slot_count = len(target_indices)
            per_slot = minutes_total // slot_count
            remainder = minutes_total - (per_slot * slot_count)
            slot_failures = 0
            for n, i in enumerate(target_indices, start=1):
                try:
                    val = per_slot + (remainder if n == slot_count else 0)
                    tsv.Item(i).Value = val
                except Exception as e:
                    slot_failures += 1
                    failures.append({"index": idx, "slot": i,
                                     "error": _format_com_error(e)})
            # Only count as written if at least one slot succeeded
            if slot_failures < slot_count:
                written += 1
        except Exception as e:
            failures.append({"index": idx, "period": p,
                             "error": _format_com_error(e)})
    status = "ok" if not failures else ("partial" if written else "error")
    return {"status": status,
            "task_id": task_id, "resource_id": resource_id,
            "unit": unit,
            "written_count": written, "failures": failures}


def _msp_progress_time_phased_read(task_id: int,
                                   resource_id: int,
                                   start_date: str,
                                   end_date: str,
                                   unit: str = "day") -> Dict[str, Any]:
    """Read per-period actual_work from an assignment via TimeScaleData.

    Phase 3b T61. Returns {status, periods: [{period_start, period_end,
    actual_work_h}]}. Empty/blank slot values (no actual yet) return as 0.0
    hours, not omitted. COM signature uses positional args (StartDate,
    EndDate, Type, TimeScaleUnit) per probe-confirmed signature on MSP 2024.
    """
    if unit not in _TIMESCALE_UNIT_MAP:
        return {"status": "error",
                "error": f"unit must be 'day' or 'week', got '{unit}'"}
    unit_code = _TIMESCALE_UNIT_MAP[unit]

    try:
        from dateutil import parser
        ds = parser.parse(str(start_date))
        de = parser.parse(str(end_date))
    except Exception as e:
        return {"status": "error",
                "error": f"could not parse date range: {e}"}

    app = _validate_active_project()
    proj = app.ActiveProject
    t = _find_task_by_id(proj, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}
    asg = _get_assignment_by_resource_id(t, resource_id)
    if asg is None:
        return {"status": "error",
                "error": f"No assignment for resource_id {resource_id} on task {task_id}"}

    out: List[Dict[str, Any]] = []
    try:
        # Positional call: (StartDate, EndDate, Type, TimeScaleUnit)
        tsv = asg.TimeScaleData(ds, de,
                                _PJ_TIMESCALED_ACTUAL_WORK,
                                unit_code)
        for i in range(1, tsv.Count + 1):
            try:
                item = tsv.Item(i)
                out.append({
                    "period_start": _msp_dt_or_none(item.StartDate),
                    "period_end": _msp_dt_or_none(item.EndDate),
                    "actual_work_h": _minutes_to_hours(item.Value),
                })
            except Exception as e:
                logger.debug(f"_msp_progress_time_phased_read item {i} failed: {e}")
                continue
        return {"status": "ok",
                "task_id": task_id, "resource_id": resource_id,
                "unit": unit, "periods": out}
    except Exception as e:
        logger.error(f"_msp_progress_time_phased_read failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_progress_bulk_update_loop(items: List[Dict[str, Any]],
                                   path_label: str,
                                   task_map: Dict[int, Any]) -> Dict[str, Any]:
    """Inner loop using pre-built task map + per-item field application.

    Items: [{task_id, percent_complete?, actual_work_h?, percent_work_complete?,
             actual_start?, actual_finish?, remaining_work_h?, physical_pct?}, ...]
    """
    updated: List[Dict[str, Any]] = []
    failures: List[Dict[str, Any]] = []
    progress_field_keys = (_PROGRESS_PCT_FIELDS | _PROGRESS_WORK_FIELDS
                           | _PROGRESS_DURATION_FIELDS | _PROGRESS_DATE_FIELDS)
    for item in items:
        tid = item.get("task_id")
        t = task_map.get(tid) if tid is not None else None
        if t is None:
            failures.append({**item, "error": f"task_id {tid} not found"})
            continue
        # Apply each progress field present on this item
        applied: List[str] = []
        item_failures: List[str] = []
        # Pre-validate pct
        for pf in _PROGRESS_PCT_FIELDS:
            if pf in item and item[pf] is not None:
                try:
                    item[pf] = _normalize_progress_pct(item[pf])
                except ValueError as e:
                    item_failures.append(f"{pf}: {e}")
                    continue
        if item_failures:
            failures.append({"task_id": tid, "error": "; ".join(item_failures)})
            continue
        for field in progress_field_keys:
            if field in item and item[field] is not None:
                try:
                    _msp_task_set_progress_field(t, field, item[field])
                    applied.append(field)
                except Exception as e:
                    item_failures.append(f"{field}: {_format_com_error(e)}")
        if applied:
            updated.append({"task_id": tid, "applied": applied})
        if item_failures:
            failures.append({"task_id": tid, "error": "; ".join(item_failures),
                            "applied": applied})
    status = "ok" if not failures else ("partial" if updated else "error")
    return {"status": status, "path": path_label,
            "count": len(updated),
            "updated": updated, "failures": failures}


def _msp_progress_bulk_update(items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Hybrid bulk progress update: routes by item count (Phase 1 _route_operation).

    Mirrors Phase 2b T37 _msp_resource_bulk_assign:
      - <=5 items   -> com_direct (no batch mode)
      - 6-19 items  -> com_batch  (batch mode + loop)
      - >=20 items  -> mspdi_bulk (Phase 3b: com_batch_fallback;
                       true MSPDI progress merge = Phase 4+)

    Pre-builds task_id -> Task map ONCE to avoid O(N×M) lookup blow-up.

    Returns: {status, path, count, updated, failures}
    """
    if not items:
        return {"status": "ok", "path": "noop", "count": 0,
                "updated": [], "failures": []}
    app = _validate_active_project()
    proj = app.ActiveProject
    task_map = _build_task_id_map(proj)

    path = _route_operation(len(items))
    if path == "com_direct":
        return _msp_progress_bulk_update_loop(items, "com_direct", task_map)
    elif path == "com_batch":
        _enter_batch_mode()
        try:
            return _msp_progress_bulk_update_loop(items, "com_batch", task_map)
        finally:
            _exit_batch_mode()
    else:
        _enter_batch_mode()
        try:
            return _msp_progress_bulk_update_loop(items, "mspdi_bulk", task_map)
        finally:
            _exit_batch_mode()


def _msp_progress_summary() -> Dict[str, Any]:
    """Project-level progress aggregate (EVM foundation).

    Phase 3b — used as input to Phase 5 EVM tool. Returns hours (CLAUDE.md
    RULE 4: BAC = sum(target_qty); for cost-loaded projects RULE 3 applies
    and Phase 5 EVM tool replaces hours with $).

    Returns {status, project: {bac_h, acwp_h, total_actual_work_h,
    total_remaining_work_h, project_percent_complete, status_date,
    task_count, in_progress_count, completed_count, not_started_count}}.
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    bac_min = 0.0
    acwp_min = 0.0
    rem_min = 0.0
    task_count = 0
    in_progress = 0
    completed = 0
    not_started = 0
    try:
        for i in range(1, proj.Tasks.Count + 1):
            try:
                t = proj.Tasks(i)
                if t is None or t.Summary:
                    continue
                task_count += 1
                bac_min += float(t.Work or 0)
                acwp_min += float(t.ActualWork or 0)
                rem_min += float(t.RemainingWork or 0)
                pct = float(t.PercentComplete or 0)
                if pct >= 100:
                    completed += 1
                elif pct > 0:
                    in_progress += 1
                else:
                    not_started += 1
            except Exception as e:
                logger.debug(f"_msp_progress_summary task {i} skip: {e}")
                continue
        # Project-level pct
        try:
            pct = float(proj.PercentComplete or 0)
        except Exception:
            pct = (acwp_min / bac_min * 100.0) if bac_min > 0 else 0.0
        try:
            status_date = _msp_dt_or_none(proj.StatusDate)
        except Exception:
            status_date = None
        return {
            "status": "ok",
            "project": {
                "bac_h": round(_minutes_to_hours(bac_min), 2),
                "acwp_h": round(_minutes_to_hours(acwp_min), 2),
                "total_actual_work_h": round(_minutes_to_hours(acwp_min), 2),
                "total_remaining_work_h": round(_minutes_to_hours(rem_min), 2),
                "project_percent_complete": round(pct, 2),
                "status_date": status_date,
                "task_count": task_count,
                "in_progress_count": in_progress,
                "completed_count": completed,
                "not_started_count": not_started,
            },
        }
    except Exception as e:
        logger.error(f"_msp_progress_summary failed: {e}")
        return {"status": "error", "error": _format_com_error(e)}


def _msp_task_update(task_id: int, name: Optional[str] = None,
                     duration: Optional[str] = None,
                     start: Optional[str] = None, finish: Optional[str] = None,
                     percent_complete: Optional[float] = None,
                     notes: Optional[str] = None) -> Dict[str, Any]:
    app = _validate_active_project()
    t = _find_task_by_id(app.ActiveProject, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}
    changes = []
    try:
        if name is not None:
            t.Name = name; changes.append("name")
        if duration is not None:
            t.Duration = _parse_duration(duration); changes.append("duration")
        if start is not None:
            t.Start = start; changes.append("start")
        if finish is not None:
            t.Finish = finish; changes.append("finish")
        if percent_complete is not None:
            t.PercentComplete = percent_complete; changes.append("percent_complete")
        if notes is not None:
            t.Notes = notes; changes.append("notes")
        return {"status": "ok", "task_id": task_id, "changes": changes}
    except Exception as e:
        logger.error(f"_msp_task_update failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_task_delete(task_id: int) -> Dict[str, Any]:
    app = _validate_active_project()
    t = _find_task_by_id(app.ActiveProject, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}
    try:
        name = t.Name
        t.Delete()
        return {"status": "ok", "deleted_id": task_id, "deleted_name": name}
    except Exception as e:
        logger.error(f"_msp_task_delete failed: {e}")
        return {"status": "error", "error": str(e)}


def _serialize_task(t: Any) -> Dict[str, Any]:
    return {
        "id": t.ID,
        "uid": t.UniqueID,
        "name": t.Name,
        "duration": t.Duration,
        "start": str(t.Start) if t.Start else None,
        "finish": str(t.Finish) if t.Finish else None,
        "percent_complete": t.PercentComplete,
        "milestone": bool(t.Milestone),
        "summary": bool(t.Summary),
        "outline_level": t.OutlineLevel,
        "notes": t.Notes or "",
    }


def _msp_task_get(task_id: int) -> Dict[str, Any]:
    app = _validate_active_project()
    t = _find_task_by_id(app.ActiveProject, task_id)
    if t is None:
        return {"status": "error", "error": f"Task ID {task_id} not found"}
    return {"status": "ok", "task": _serialize_task(t)}


def _msp_task_list(include_summary: bool = True, limit: int = 100) -> Dict[str, Any]:
    app = _validate_active_project()
    proj = app.ActiveProject
    out = []
    for i in range(1, min(proj.Tasks.Count, limit) + 1):
        t = proj.Tasks(i)
        if t is None:
            continue
        if not include_summary and t.Summary:
            continue
        out.append(_serialize_task(t))
    return {"status": "ok", "total": proj.Tasks.Count, "returned": len(out), "tasks": out}


def _msp_task_add_summary(name: str, duration: str = "1d",
                          parent_task_id: Optional[int] = None) -> Dict[str, Any]:
    """Add a summary task. In MS Project summary is implicit (becomes summary when has children).

    TODO(T11+): Wire `parent_task_id` to indent the new task under the given parent
    via OutlineLevel / Task.OutlineIndent so it actually becomes a summary container.
    Currently the parameter is accepted for API stability but ignored — MSP only marks
    a task `Summary=True` when it has children, so summary-ness is implicit, not flagged.
    """
    return _msp_task_add_single(name=name, duration=duration, summary=True)


def _msp_task_add_milestone(name: str, date: Optional[str] = None) -> Dict[str, Any]:
    """Add a 0-duration milestone."""
    r = _msp_task_add_single(name=name, duration="0d", milestone=True)
    if r.get("status") == "ok" and date:
        _msp_task_update(task_id=r["task_id"], start=date, finish=date)
    return r


# ---------- BULK ADD HYBRID ROUTING ----------

def _msp_task_bulk_add_com_direct(items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Path 1: <=5 items, plain COM."""
    added = []
    for item in items:
        r = _msp_task_add_single(**item)
        if r.get("status") == "ok":
            added.append(r["task_id"])
    return {"status": "ok", "path": "com_direct", "count": len(added), "task_ids": added}


def _msp_task_bulk_add_com_batch(items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Path 2: 6-19 items, COM with batch mode (Calculation manual)."""
    _enter_batch_mode()
    try:
        added = []
        for item in items:
            r = _msp_task_add_single(**item)
            if r.get("status") == "ok":
                added.append(r["task_id"])
        return {"status": "ok", "path": "com_batch", "count": len(added), "task_ids": added}
    finally:
        _exit_batch_mode()


def _msp_task_bulk_add_mspdi(items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Path 3: 20+ items via MSPDI XML import.

    Strategy:
    1. Build temp MSPDI XML with all tasks via MsprojectBulkWriter
    2. Open as new project via app.FileOpen(temp.xml) — DisplayAlerts=False
       suppresses the Import Wizard dialog. The temp project's name is the
       XML filename stem (NOT the <Name> element), so we control it via the
       temp filename.
    3. SelectAll + EditCopy in temp project
    4. Switch back to target project, SelectTaskField row=last_count+1, EditPaste
    5. EditPaste sometimes inserts a leading None placeholder row; remove it
       via app.SelectRow(Row=1) + app.EditDelete (only if Tasks(1) is None)
    6. Close temp project without saving (FileClose 0)
    7. Cleanup temp file

    Performance target: 200 tasks in <5 seconds.

    On failure, falls back to COM-batch-per-task (slower but reliable).
    """
    import tempfile
    from msproject_bulk import MsprojectBulkWriter

    app = _validate_active_project()
    target_proj = app.ActiveProject
    target_name = target_proj.Name

    # Get current project start for new tasks (use as MSPDI StartDate)
    try:
        start_date_obj = target_proj.ProjectStart
        start_date = start_date_obj.strftime("%Y-%m-%dT%H:%M:%S") if start_date_obj else None
    except Exception:
        start_date = None

    # Use a controlled temp filename so we can find the loaded project by name
    # (FileOpen names the project after the file's stem, not the XML <Name>).
    tmp_dir = tempfile.gettempdir()
    tmp_stem = f"_BULK_TEMP_{int(time.time() * 1000)}"
    tmp_path = os.path.join(tmp_dir, f"{tmp_stem}.xml")

    # Build XML
    w = MsprojectBulkWriter(project_name=tmp_stem, start_date=start_date)
    w.bulk_add_tasks(items)
    w.save(tmp_path)

    # Pre-compute expected Duration per item (in MSP minutes) for post-paste
    # fix-up. MSP 16.0 silently drops <Duration> from MSPDI FileOpen imports
    # (Phase 2b TAIL — verified by probe across multiple XML variants:
    # adding <Manual>0</Manual>, full Calendar with WorkingTimes,
    # <MinutesPerDay>/<MinutesPerWeek> project settings, <Start>/<Finish>
    # — none restore Duration). Workaround: post-paste, re-set t.Duration
    # via COM in batch mode (~5 ms/task, ≈1s for 200 tasks).
    _expected_durations_min = [
        _parse_duration(item.get("duration", "1d")) for item in items
    ]

    # Save & flip DisplayAlerts (this is what suppresses the Import Wizard dialog)
    prev_alerts = True
    try:
        prev_alerts = app.DisplayAlerts
    except Exception:
        pass
    try:
        app.DisplayAlerts = False
    except Exception:
        pass

    _enter_batch_mode()
    try:
        # Open temp XML as new project
        app.FileOpen(tmp_path)

        # The loaded project's name is the file stem
        temp_proj = None
        for i in range(1, app.Projects.Count + 1):
            if app.Projects(i).Name == tmp_stem:
                temp_proj = app.Projects(i)
                break
        if temp_proj is None:
            raise RuntimeError(
                f"FileOpen didn't load {tmp_stem} (Projects: "
                f"{[app.Projects(i).Name for i in range(1, app.Projects.Count + 1)]})"
            )

        # Select all tasks in temp and copy
        try:
            app.SelectAll()
        except Exception:
            # Fallback: select via row range
            app.SelectTaskField(Row=1, Column="Name", Height=temp_proj.Tasks.Count)
        app.EditCopy()

        # Switch to target project window
        target_window_found = False
        for i in range(1, app.Projects.Count + 1):
            if app.Projects(i).Name == target_name:
                try:
                    app.WindowActivate(app.Projects(i).Windows(1).Caption)
                    target_window_found = True
                    break
                except Exception:
                    continue
        if not target_window_found:
            raise RuntimeError(f"Could not switch back to target project {target_name}")

        target_proj = app.ActiveProject
        last_count = target_proj.Tasks.Count
        paste_row = last_count + 1 if last_count > 0 else 1
        try:
            app.SelectTaskField(Row=paste_row, Column="Name")
        except Exception:
            app.SelectTaskField(Row=1, Column="Name")
        app.EditPaste()

        # Cleanup leading None placeholder rows that EditPaste sometimes inserts.
        # Use app.SelectRow(Row=1, RowRelative=False) + app.EditDelete to delete
        # the row by position (not by Tasks(N).Delete which fails on None).
        cleanup_safety = 50
        while (target_proj.Tasks.Count > 0
               and target_proj.Tasks(1) is None
               and cleanup_safety > 0):
            try:
                app.SelectRow(Row=1, RowRelative=False)
                app.EditDelete()
            except Exception as ce:
                logger.warning(f"None-row cleanup at row 1 failed: {ce}")
                break
            cleanup_safety -= 1

        # Cleanup trailing None rows too (defensive)
        cleanup_safety = 50
        while (target_proj.Tasks.Count > 0
               and target_proj.Tasks(target_proj.Tasks.Count) is None
               and cleanup_safety > 0):
            try:
                app.SelectRow(Row=target_proj.Tasks.Count, RowRelative=False)
                app.EditDelete()
            except Exception:
                break
            cleanup_safety -= 1

        # Close temp project without saving
        for i in range(1, app.Projects.Count + 1):
            if app.Projects(i).Name == tmp_stem:
                try:
                    app.WindowActivate(app.Projects(i).Windows(1).Caption)
                    app.FileClose(0)  # 0 = pjDoNotSave
                    break
                except Exception:
                    pass

        # Reactivate target window
        for i in range(1, app.Projects.Count + 1):
            if app.Projects(i).Name == target_name:
                try:
                    app.WindowActivate(app.Projects(i).Windows(1).Caption)
                    break
                except Exception:
                    pass

        # Collect IDs of newly added tasks AND restore Duration (MSPDI import
        # drops it — see workaround note above). The collection loop iterates
        # only valid (non-None) rows, in items order, so duration_idx maps
        # 1:1 to items[].
        target_proj = app.ActiveProject
        added_task_ids = []
        duration_set_failures = 0
        duration_idx = 0
        for i in range(last_count + 1, target_proj.Tasks.Count + 1):
            t = target_proj.Tasks(i)
            if t is None:
                continue
            if duration_idx < len(_expected_durations_min):
                try:
                    t.Duration = _expected_durations_min[duration_idx]
                except Exception as e:
                    duration_set_failures += 1
                    logger.debug(
                        f"post-paste Duration set failed at row {i} "
                        f"(idx={duration_idx}): {e}"
                    )
            added_task_ids.append(t.ID)
            duration_idx += 1

        return {
            "status": "ok",
            "path": "mspdi_bulk",
            "count": len(added_task_ids),
            "task_ids": added_task_ids,
            "method": "FileOpen + EditCopy + EditPaste + post-paste Duration set",
            "duration_set_failures": duration_set_failures,
        }
    except Exception as e:
        logger.error(f"MSPDI bulk path failed: {e}; falling back to COM batch")
        # Cleanup any orphan _BULK_TEMP_ project
        try:
            for i in range(1, app.Projects.Count + 1):
                if app.Projects(i).Name == tmp_stem:
                    app.WindowActivate(app.Projects(i).Windows(1).Caption)
                    app.FileClose(0)
                    break
        except Exception:
            pass
        # Switch back to target
        try:
            for i in range(1, app.Projects.Count + 1):
                if app.Projects(i).Name == target_name:
                    app.WindowActivate(app.Projects(i).Windows(1).Caption)
                    break
        except Exception:
            pass
        # Fallback to COM batch
        added = []
        for item in items:
            r = _msp_task_add_single(**item)
            if r.get("status") == "ok":
                added.append(r["task_id"])
        return {
            "status": "ok",
            "path": "mspdi_bulk_fallback_com",
            "count": len(added),
            "task_ids": added,
            "fallback_reason": str(e),
        }
    finally:
        _exit_batch_mode()
        try:
            app.DisplayAlerts = prev_alerts
        except Exception:
            pass
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


def _msp_task_bulk_add(items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Hybrid bulk add: routes by item count via _route_operation()."""
    if not items:
        return {"status": "ok", "path": "noop", "count": 0, "task_ids": []}
    path = _route_operation(len(items))
    if path == "com_direct":
        return _msp_task_bulk_add_com_direct(items)
    elif path == "com_batch":
        return _msp_task_bulk_add_com_batch(items)
    else:
        return _msp_task_bulk_add_mspdi(items)


# ---------- LINK HELPERS ----------

PJ_LINK_TYPES = {"FF": 0, "FS": 1, "SF": 2, "SS": 3}
PJ_LINK_REVERSE = {v: k for k, v in PJ_LINK_TYPES.items()}


def _msp_link_add(predecessor_id: int, successor_id: int,
                  type: str = "FS", lag: str = "0d") -> Dict[str, Any]:
    """Add a single predecessor link via COM (Path 1).

    Uses Predecessors string append: "5FS,3SS+2d" format.
    """
    app = _validate_active_project()
    proj = app.ActiveProject
    pred = _find_task_by_id(proj, predecessor_id)
    succ = _find_task_by_id(proj, successor_id)
    if pred is None:
        return {"status": "error", "error": f"Predecessor {predecessor_id} not found"}
    if succ is None:
        return {"status": "error", "error": f"Successor {successor_id} not found"}
    try:
        existing = succ.Predecessors or ""
        # Build new token: "5FS" or "5FS+2d" or "5FS-1d"
        type_str = type.upper() if type and type.upper() in PJ_LINK_TYPES else "FS"
        lag_str = ""
        if lag and lag != "0d":
            # Convert minutes to display format
            mins = _parse_duration(lag)
            if mins > 0:
                # Days if multiple of 480 mins
                if mins % 480 == 0:
                    lag_str = f"+{mins // 480}d"
                elif mins % 60 == 0:
                    lag_str = f"+{mins // 60}h"
                else:
                    lag_str = f"+{mins}m"
            elif mins < 0:
                if mins % 480 == 0:
                    lag_str = f"-{abs(mins) // 480}d"
                else:
                    lag_str = f"-{abs(mins)}m"
        new_token = f"{predecessor_id}{type_str}{lag_str}"
        succ.Predecessors = (existing + "," + new_token).strip(",") if existing else new_token
        return {
            "status": "ok",
            "predecessor_id": predecessor_id,
            "successor_id": successor_id,
            "type": type_str,
            "lag": lag,
        }
    except Exception as e:
        logger.error(f"_msp_link_add failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_link_delete(predecessor_id: int, successor_id: int) -> Dict[str, Any]:
    """Remove a specific predecessor link by rebuilding Predecessors string."""
    app = _validate_active_project()
    succ = _find_task_by_id(app.ActiveProject, successor_id)
    if succ is None:
        return {"status": "error", "error": f"Successor {successor_id} not found"}
    try:
        existing = succ.Predecessors or ""
        if not existing:
            return {"status": "ok", "message": "No predecessors to remove"}
        # Parse tokens (comma-separated): "5FS+2d", "3SS", "12FF-1d"
        tokens = [t.strip() for t in existing.split(",") if t.strip()]
        # Filter out matching predecessor (number prefix matches predecessor_id)
        kept = []
        removed_count = 0
        for tok in tokens:
            # Extract numeric prefix
            num_str = ""
            for ch in tok:
                if ch.isdigit():
                    num_str += ch
                else:
                    break
            try:
                tok_id = int(num_str) if num_str else -1
            except ValueError:
                tok_id = -1
            if tok_id == predecessor_id:
                removed_count += 1
            else:
                kept.append(tok)
        succ.Predecessors = ",".join(kept)
        if removed_count == 0:
            return {"status": "ok", "message": f"No link from {predecessor_id} to {successor_id} found"}
        return {"status": "ok", "predecessor_id": predecessor_id, "successor_id": successor_id,
                "removed_count": removed_count}
    except Exception as e:
        logger.error(f"_msp_link_delete failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_link_update(predecessor_id: int, successor_id: int,
                     new_type: Optional[str] = None,
                     new_lag: Optional[str] = None) -> Dict[str, Any]:
    """Update link properties via remove+re-add."""
    delete_result = _msp_link_delete(predecessor_id=predecessor_id, successor_id=successor_id)
    if delete_result.get("status") != "ok":
        return delete_result
    add_result = _msp_link_add(
        predecessor_id=predecessor_id,
        successor_id=successor_id,
        type=new_type or "FS",
        lag=new_lag or "0d",
    )
    return add_result


def _msp_link_bulk_add(items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Hybrid routing for bulk links (uses _route_operation)."""
    if not items:
        return {"status": "ok", "path": "noop", "count": 0}
    path = _route_operation(len(items))
    if path == "com_batch" or path == "mspdi_bulk":
        _enter_batch_mode()
    try:
        added = 0
        errors = []
        for it in items:
            r = _msp_link_add(**it)
            if r.get("status") == "ok":
                added += 1
            else:
                errors.append(r)
        result = {"status": "ok", "path": path, "count": added}
        if errors:
            result["errors"] = errors[:10]  # cap to avoid huge response
        return result
    finally:
        if path in ("com_batch", "mspdi_bulk"):
            _exit_batch_mode()


def _msp_link_chain(task_ids: List[int], type: str = "FS",
                    lag: str = "0d") -> Dict[str, Any]:
    """Chain N tasks: T1->T2->T3->...->TN with given link type."""
    if len(task_ids) < 2:
        return {"status": "error", "error": "At least 2 tasks required for chain"}
    added = 0
    errors = []
    for i in range(len(task_ids) - 1):
        r = _msp_link_add(
            predecessor_id=task_ids[i],
            successor_id=task_ids[i+1],
            type=type, lag=lag,
        )
        if r.get("status") == "ok":
            added += 1
        else:
            errors.append({"index": i, "error": r.get("error")})
    return {
        "status": "ok",
        "links_added": added,
        "chain_length": len(task_ids),
        "errors": errors[:10] if errors else [],
    }


# ---------- SCHEDULE HELPERS ----------

def _to_com_date(date_str: str) -> Any:
    """Convert ISO date string ('2026-04-30') to a pywintypes.Time COM date.

    MS Project's StatusDate property requires a real VT_DATE value; passing a
    plain string raises 'argument value is not valid'. Accepts both 'YYYY-MM-DD'
    and 'YYYY-MM-DD HH:MM:SS'.
    """
    s = date_str.strip()
    fmts = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y")
    last_err: Optional[Exception] = None
    for f in fmts:
        try:
            dt = _dt.datetime.strptime(s, f)
            return pywintypes.Time(dt)
        except Exception as e:
            last_err = e
    raise ValueError(f"Unrecognized date format: {date_str!r} ({last_err})")


def _msp_schedule_reschedule(report_date: Optional[str] = None) -> Dict[str, Any]:
    """Recalculate the project (CalculateProject)."""
    app = _validate_active_project()
    try:
        if report_date:
            app.ActiveProject.StatusDate = _to_com_date(report_date)
        app.CalculateProject()
        return {"status": "ok", "message": "Project rescheduled",
                "report_date": report_date or "unchanged"}
    except Exception as e:
        logger.error(f"_msp_schedule_reschedule failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_schedule_level(within_slack: bool = False) -> Dict[str, Any]:
    """Level resources (Application.LevelNow)."""
    app = _validate_active_project()
    try:
        # MS Project COM uses LevelNow(All=True) — no LevelAll method exists.
        # LevelingOptions controls within_slack behavior.
        if within_slack:
            try:
                app.LevelingOptions(DelayInSlack=True)
            except Exception as e:
                logger.warning(f"LevelingOptions(DelayInSlack=True) failed (non-fatal): {e}")
        app.LevelNow(All=True)
        return {"status": "ok", "message": "Resources leveled",
                "within_slack": within_slack}
    except Exception as e:
        logger.error(f"_msp_schedule_level failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_schedule_set_data_date(date: str) -> Dict[str, Any]:
    """Set status date / data date."""
    app = _validate_active_project()
    try:
        app.ActiveProject.StatusDate = _to_com_date(date)
        return {"status": "ok", "status_date": date}
    except Exception as e:
        logger.error(f"_msp_schedule_set_data_date failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_schedule_protect_actuals(enable: bool = True) -> Dict[str, Any]:
    """Protect actuals from being recalculated.

    In MS Project these are Project-level properties that govern whether the
    scheduling engine moves end-of-completed and start-of-remaining parts when
    rescheduling. enable=True locks them down (no move).

      MoveCompleted          Move end of completed parts after status date back to status date
      AndMoveRemaining       And move start of remaining parts back to status date
      MoveRemaining          Move start of remaining parts before status date forward to status date
      AndMoveCompleted       And move end of completed parts forward to status date
    """
    app = _validate_active_project()
    try:
        proj = app.ActiveProject
        # When 'enable=True' = protect (don't move any completed/remaining parts)
        proj.MoveCompleted = not enable
        proj.AndMoveRemaining = not enable
        proj.MoveRemaining = not enable
        proj.AndMoveCompleted = not enable
        return {"status": "ok", "actuals_protected": bool(enable)}
    except Exception as e:
        logger.error(f"_msp_schedule_protect_actuals failed: {e}")
        return {"status": "error", "error": str(e)}


# ---------- TOOL DISPATCHERS ----------

@mcp.tool(
    name="msproject_task",
    annotations={"title": "MS Project Task Operations", "readOnlyHint": False},
)
async def msproject_task(params: dict) -> str:
    """Manage tasks in active MS Project (COM-based, hybrid speed routing).

    Actions:
    - add: Add task. Params: name, duration (e.g. "5d"), [start, finish, summary, milestone, notes]
    - update: Update task. Params: task_id, [name, duration, start, finish, percent_complete, notes]
    - delete: Delete task. Params: task_id
    - add_summary: Add summary task. Params: name, duration, [parent_task_id]
    - add_milestone: Add 0-duration milestone. Params: name, [date]
    - get: Get task details. Params: task_id
    - list: List tasks. Params: [include_summary=true, limit=100]
    - bulk_add: Bulk add tasks. Params: items=[{name, duration, ...}, ...]
                Routes: 1-5=COM direct, 6-19=COM batch, 20+=MSPDI bulk

    Returns JSON-encoded result. Hybrid speed: 200 task bulk in ~3-5 sec.
    """
    import json
    action = params.get("action", "")
    p = {k: v for k, v in params.items() if k != "action"}
    try:
        if action == "add":
            r = _msp_task_add_single(**p)
        elif action == "update":
            r = _msp_task_update(**p)
        elif action == "delete":
            r = _msp_task_delete(**p)
        elif action == "add_summary":
            r = _msp_task_add_summary(**p)
        elif action == "add_milestone":
            r = _msp_task_add_milestone(**p)
        elif action == "get":
            r = _msp_task_get(**p)
        elif action == "list":
            r = _msp_task_list(**p)
        elif action == "bulk_add":
            r = _msp_task_bulk_add(**p)
        else:
            r = {"status": "error",
                 "error": f"Unknown action '{action}'. Valid: add/update/delete/add_summary/add_milestone/get/list/bulk_add"}
    except Exception as e:
        logger.error(f"msproject_task({action}) failed: {e}")
        r = {"status": "error", "error": str(e)}
    return json.dumps(r, default=str, ensure_ascii=False)


@mcp.tool(
    name="msproject_link",
    annotations={"title": "MS Project Link Operations", "readOnlyHint": False},
)
async def msproject_link(params: dict) -> str:
    """Manage predecessor/successor links.

    Actions:
    - add: Add link. Params: predecessor_id, successor_id, [type='FS', lag='0d']
    - delete: Remove link. Params: predecessor_id, successor_id
    - update: Update link (type/lag). Params: predecessor_id, successor_id, [new_type, new_lag]
    - bulk_add: Bulk links. Params: items=[{predecessor_id, successor_id, type, lag}, ...]
    - chain: Chain N tasks T1->T2->...->TN. Params: task_ids=[1,2,3,...], [type='FS', lag='0d']
    """
    import json
    action = params.get("action", "")
    p = {k: v for k, v in params.items() if k != "action"}
    try:
        if action == "add":
            r = _msp_link_add(**p)
        elif action == "delete":
            r = _msp_link_delete(**p)
        elif action == "update":
            r = _msp_link_update(**p)
        elif action == "bulk_add":
            r = _msp_link_bulk_add(**p)
        elif action == "chain":
            r = _msp_link_chain(**p)
        else:
            r = {"status": "error",
                 "error": f"Unknown action '{action}'. Valid: add/delete/update/bulk_add/chain"}
    except Exception as e:
        logger.error(f"msproject_link({action}) failed: {e}")
        r = {"status": "error", "error": str(e)}
    return json.dumps(r, default=str, ensure_ascii=False)


@mcp.tool(
    name="msproject_schedule",
    annotations={"title": "MS Project Schedule Operations", "readOnlyHint": False},
)
async def msproject_schedule(params: dict) -> str:
    """Schedule operations.

    Actions:
    - reschedule: Recalculate (CalculateProject). Params: [report_date='YYYY-MM-DD']
    - level: Resource leveling (LevelNow). Params: [within_slack=False]
    - set_data_date: Set status_date. Params: date='YYYY-MM-DD'
    - protect_actuals: Lock actuals. Params: [enable=True]
    """
    import json
    action = params.get("action", "")
    p = {k: v for k, v in params.items() if k != "action"}
    try:
        if action == "reschedule":
            r = _msp_schedule_reschedule(**p)
        elif action == "level":
            r = _msp_schedule_level(**p)
        elif action == "set_data_date":
            r = _msp_schedule_set_data_date(**p)
        elif action == "protect_actuals":
            r = _msp_schedule_protect_actuals(**p)
        else:
            r = {"status": "error",
                 "error": f"Unknown action '{action}'. Valid: reschedule/level/set_data_date/protect_actuals"}
    except Exception as e:
        logger.error(f"msproject_schedule({action}) failed: {e}")
        r = {"status": "error", "error": str(e)}
    return json.dumps(r, default=str, ensure_ascii=False)


@mcp.tool(
    name="msproject_calendar",
    annotations={"title": "MS Project Calendar Operations", "readOnlyHint": False},
)
async def msproject_calendar(params: dict) -> str:
    """Manage project calendars in active MS Project (COM-based).

    Actions:
    - create: New base calendar from existing one. Params: name, [base_calendar="Standard"]
    - update: Rename or set weekday off. Params: name, [new_name, weekday_off=1-7]
    - add_exception: Non-working day/range. Params: calendar_name, exception_name, start (YYYY-MM-DD), [finish, working=False]
    - assign_to_task: Apply calendar to a task. Params: task_id, calendar_name
    - assign_to_resource: Apply calendar to a resource. Params: resource_id, calendar_name
    - list: List all base calendars + exception counts. Params: (none)
    - holidays_uzbek: Bulk-add 9 Ozbekistan 2026 official holidays (idempotent, name-based dedup). Params: calendar_name, [year=2026]

    Phase 2a (27 Apr 2026). Resource integration arrives in Phase 2b.
    """
    import json
    action = params.get("action", "")
    p = {k: v for k, v in params.items() if k != "action"}
    # Alias: accept 'name' / 'calendar_name' interchangeably across actions.
    # Reject ambiguous calls where both keys are provided rather than silently
    # dropping one — surfaces caller bugs instead of hiding them.
    if "name" in p and "calendar_name" in p:
        return json.dumps(
            {"status": "error",
             "error": "Specify either 'name' or 'calendar_name', not both"},
            default=str, ensure_ascii=False,
        )
    if action in _CALENDAR_NAME_ALIAS_ACTIONS and "name" in p and "calendar_name" not in p:
        p["calendar_name"] = p.pop("name")
    elif action in _CALENDAR_NAME_NATIVE_ACTIONS and "calendar_name" in p and "name" not in p:
        p["name"] = p.pop("calendar_name")
    try:
        if action == "create":
            r = _msp_calendar_create(**p)
        elif action == "update":
            r = _msp_calendar_update(**p)
        elif action == "add_exception":
            r = _msp_calendar_add_exception(**p)
        elif action == "assign_to_task":
            r = _msp_calendar_assign_to_task(**p)
        elif action == "assign_to_resource":
            r = _msp_calendar_assign_to_resource(**p)
        elif action == "list":
            r = _msp_calendar_list(**p)
        elif action == "holidays_uzbek":
            r = _msp_calendar_holidays_uzbek(**p)
        else:
            r = {"status": "error",
                 "error": f"Unknown action '{action}'. Valid: create/update/add_exception/assign_to_task/assign_to_resource/list/holidays_uzbek"}
    except Exception as e:
        logger.error(f"msproject_calendar({action}) failed: {e}")
        r = {"status": "error", "error": _format_com_error(e)}
    return json.dumps(r, default=str, ensure_ascii=False)


@mcp.tool(
    name="msproject_resource",
    annotations={"title": "MS Project Resource Operations", "readOnlyHint": False},
)
async def msproject_resource(params: dict) -> str:
    """Manage resources + assignments in active MS Project (COM-based, hybrid bulk).

    Actions:
    - add: Add resource. Params: name, type=Work|Material|Cost, [max_units, standard_rate, overtime_rate, material_label]
    - update: Update. Params: resource_id, [name, max_units, standard_rate, overtime_rate, material_label]
    - delete: Params: resource_id
    - list: List all resources + types + assignment counts
    - assign: Single. Params: task_id, resource_id, [units=100, work_hours]
    - unassign: Params: task_id, resource_id
    - bulk_assign: Hybrid (1-5 COM, 6-19 batch, 20+ MSPDI fallback). Params: items=[{task_id, resource_id, [units]}, ...]

    Phase 2b (28 Apr 2026). Note: bulk_assign perf is ~10ms/call due to MS Project
    COM intrinsic limit; true MSPDI assignment merge is Phase 3+.
    """
    import json
    action = params.get("action", "")
    p = {k: v for k, v in params.items() if k != "action"}
    try:
        if action == "add":
            r = _msp_resource_add(**p)
        elif action == "update":
            r = _msp_resource_update(**p)
        elif action == "delete":
            r = _msp_resource_delete(**p)
        elif action == "list":
            r = _msp_resource_list(**p)
        elif action == "assign":
            r = _msp_resource_assign(**p)
        elif action == "unassign":
            r = _msp_resource_unassign(**p)
        elif action == "bulk_assign":
            r = _msp_resource_bulk_assign(**p)
        else:
            r = {"status": "error",
                 "error": f"Unknown action '{action}'. Valid: add/update/delete/list/assign/unassign/bulk_assign"}
    except Exception as e:
        logger.error(f"msproject_resource({action}) failed: {e}")
        r = {"status": "error", "error": _format_com_error(e)}
    return json.dumps(r, default=str, ensure_ascii=False)


@mcp.tool(
    name="msproject_baseline",
    annotations={"title": "MS Project Baseline Operations", "readOnlyHint": False},
)
async def msproject_baseline(params: dict) -> str:
    """Multi-baseline (Baseline + Baseline1..Baseline10) management + variance reporting.

    Actions:
    - save: Save current as baseline. Params: [baseline_number=0, name, scope='all', roll_up_to_summary=True]
    - clear: Clear single baseline. Params: [baseline_number=0]
    - clear_all: Clear all 11 saved baselines.
    - list: List all saved baselines with metadata.
    - get_task_baseline: Read one task's baseline values. Params: task_id, [baseline_number=0]
    - compare: Variance current vs baseline. Params: [baseline_number=0, include_unchanged=False, variance_threshold_days=0]
    - compare_two: Delta between two baselines. Params: baseline_a, baseline_b, [include_unchanged, variance_threshold_days]
    - summary: Project-level RAG status. Params: [baseline_number=0]
    - set_active: Set active baseline for views (version-dependent). Params: baseline_number

    Phase 3a (28 Apr 2026). RAG thresholds: green<=5%, amber<=20%, red>20% slipped.
    """
    import json
    action = params.get("action", "")
    p = {k: v for k, v in params.items() if k != "action"}
    try:
        if action == "save":
            r = _msp_baseline_save(**p)
        elif action == "clear":
            r = _msp_baseline_clear(**p)
        elif action == "clear_all":
            r = _msp_baseline_clear_all(**p)
        elif action == "list":
            r = _msp_baseline_list(**p)
        elif action == "get_task_baseline":
            r = _msp_baseline_get_task_baseline(**p)
        elif action == "compare":
            r = _msp_baseline_compare(**p)
        elif action == "compare_two":
            r = _msp_baseline_compare_two(**p)
        elif action == "summary":
            r = _msp_baseline_summary(**p)
        elif action == "set_active":
            r = _msp_baseline_set_active(**p)
        else:
            r = {"status": "error",
                 "error": f"Unknown action '{action}'. Valid: save/clear/clear_all/list/get_task_baseline/compare/compare_two/summary/set_active"}
    except Exception as e:
        logger.error(f"msproject_baseline({action}) failed: {e}")
        r = {"status": "error", "error": _format_com_error(e)}
    return json.dumps(r, default=str, ensure_ascii=False)


@mcp.tool(
    name="msproject_progress",
    annotations={"title": "MS Project Progress Operations", "readOnlyHint": False},
)
async def msproject_progress(params: dict) -> str:
    """Task + assignment progress, time-phased actuals, status date, EVM-ready summary.

    Actions:
    - set_task_progress: Task-level set. Params: task_id, [percent_complete, percent_work_complete, actual_start, actual_finish, actual_duration_h, actual_work_h, remaining_work_h, remaining_duration_h, physical_pct, stop, resume]
    - get_task_progress: Task-level read. Params: task_id
    - set_assignment_progress: Per-resource man-hour set. Params: task_id, resource_id, [actual_work_h, actual_start, actual_finish, percent_work_complete, remaining_work_h, units]
    - get_assignment_progress: Per-task assignments list. Params: task_id
    - set_progress_by_date: Bulk update via app.UpdateProject. Params: progress_date, [scope='all', as_scheduled=True]
    - set_status_date: Set proj.StatusDate (data_date). Params: status_date
    - clear_progress: Single-task progress reset. Params: task_id
    - clear_all_progress: Project-wide progress reset.
    - time_phased_actual_write: Per-period actual_work write. Params: task_id, resource_id, periods, [unit='day']
    - time_phased_actual_read: Per-period actual_work read. Params: task_id, resource_id, start_date, end_date, [unit='day']
    - bulk_progress_update: Hybrid bulk path. Params: items (list of {task_id, ...progress fields})
    - summary: Project-level EVM-ready aggregate.

    Phase 3b (29 Apr 2026). DCMA-aligned: physical_pct exposed for EV input
    (independent of percent_complete). Foundation for Phase 5 msproject_evm.
    """
    import json
    action = params.get("action", "")
    p = {k: v for k, v in params.items() if k != "action"}
    try:
        if action == "set_task_progress":
            r = _msp_progress_set_task(**p)
        elif action == "get_task_progress":
            r = _msp_progress_get_task(**p)
        elif action == "set_assignment_progress":
            r = _msp_progress_set_assignment(**p)
        elif action == "get_assignment_progress":
            r = _msp_progress_get_assignments(**p)
        elif action == "set_progress_by_date":
            r = _msp_progress_set_by_date(**p)
        elif action == "set_status_date":
            r = _msp_progress_set_status_date(**p)
        elif action == "clear_progress":
            r = _msp_progress_clear(**p)
        elif action == "clear_all_progress":
            r = _msp_progress_clear_all(**p)
        elif action == "time_phased_actual_write":
            r = _msp_progress_time_phased_write(**p)
        elif action == "time_phased_actual_read":
            r = _msp_progress_time_phased_read(**p)
        elif action == "bulk_progress_update":
            r = _msp_progress_bulk_update(**p)
        elif action == "summary":
            r = _msp_progress_summary(**p)
        else:
            r = {"status": "error",
                 "error": f"Unknown action '{action}'. Valid: set_task_progress/"
                          "get_task_progress/set_assignment_progress/"
                          "get_assignment_progress/set_progress_by_date/"
                          "set_status_date/clear_progress/clear_all_progress/"
                          "time_phased_actual_write/time_phased_actual_read/"
                          "bulk_progress_update/summary"}
    except Exception as e:
        logger.error(f"msproject_progress({action}) failed: {e}")
        r = {"status": "error", "error": _format_com_error(e)}
    return json.dumps(r, default=str, ensure_ascii=False)


# ============================================================================
# PHASE 4 - FILE MCP (msproject_file)
# ============================================================================
# Phase 4 (T65 foundations) — adds file-mode helpers separate from the COM
# code above. Phase 1-3 helpers DO NOT reference anything below this line.
# Imports/state added here:
#   * MspdiProject (native MSPDI parser, reused from Asta side)
#   * _jvm_started bool (lazy MPXJ JVM lifecycle)
#   * _detect_msp_xml_schema, _get_msp_file_manager, MspMppFileManager
# ----------------------------------------------------------------------------

# Native MSPDI parser (zero Java dependency, reuse from Asta).
from mspdi_parser import MspdiProject

# JVM pre-start for MPXJ (lazy — only if .mpp encountered).
_jvm_started = False


def _ensure_jvm_started() -> None:
    """Start JVM lazily on first MPP request. Idempotent.

    MPXJ requires JPype1+JVM to read .mpp binaries. We pay this cost only
    when the user actually opens a .mpp file. Re-invocation is a no-op.
    """
    global _jvm_started
    if _jvm_started:
        return
    try:
        import mpxj
        if not mpxj.jpype.isJVMStarted():
            mpxj.jpype.startJVM()
            logger.info("JVM started for MPXJ")
        _jvm_started = True
    except ImportError as e:
        raise RuntimeError(
            "MPXJ not installed. For .mpp support: pip install mpxj jpype1"
        ) from e


def _detect_msp_xml_schema(file_path: str) -> bool:
    """Read first 512 bytes; check for MS Project MSPDI namespace.

    Returns True if MS Project XML (schemas.microsoft.com/project), False
    if Asta or unknown schema. Raises FileNotFoundError if path missing.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    with open(file_path, 'rb') as f:
        head = f.read(512).decode('utf-8', errors='replace')
    return 'schemas.microsoft.com/project' in head


def _mpxj_duration_to_hours(d) -> float:
    """Convert MPXJ Duration object to hours (float). Handles None safely."""
    if d is None:
        return 0.0
    try:
        from org.mpxj import TimeUnit
        n = float(d.getDuration())
        unit = d.getUnits()
        if unit == TimeUnit.HOURS:
            return n
        elif unit == TimeUnit.DAYS:
            return n * 8.0
        elif unit == TimeUnit.WEEKS:
            return n * 40.0
        elif unit == TimeUnit.MINUTES:
            return n / 60.0
        else:
            return n  # fallback assume hours
    except Exception as e:
        logger.warning(f"_mpxj_duration_to_hours failed (returning 0.0): {e}")
        return 0.0


class MspMppFileManager:
    """Read-only manager for .mpp files via MPXJ + JVM.

    Adapted from Asta asta_mcp_file.py AstaFileManager (drops .pp support,
    MPP only). MPP write is not supported by MPXJ — Phase 4 intentionally
    keeps this class read-only.

    Lazy load: __init__ does NOT touch MPXJ/JVM. The first read_*() call
    triggers JVM start + UniversalProjectReader. This makes init cheap
    (and importable on systems without Java).
    """

    def __init__(self, file_path: str):
        self.file_path = file_path.replace("\\", "/")
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"File not found: {self.file_path}")
        self._project = None  # lazy load on first read

    def _load(self):
        if self._project is not None:
            return self._project
        _ensure_jvm_started()
        from org.mpxj.reader import UniversalProjectReader
        reader = UniversalProjectReader()
        try:
            self._project = reader.read(self.file_path)
            logger.info(f"MPP loaded: {self.file_path}")
            return self._project
        except Exception as e:
            raise RuntimeError(f"MPXJ read failed for {self.file_path}: {e}") from e

    def read_tasks(self) -> List[Dict[str, Any]]:
        proj = self._load()
        out: List[Dict[str, Any]] = []
        for t in proj.getTasks():
            if t is None or t.getID() == 0:
                continue
            out.append({
                "id": int(t.getID()),
                "name": str(t.getName() or ""),
                "duration_h": _mpxj_duration_to_hours(t.getDuration()),
                "start": str(t.getStart()) if t.getStart() else None,
                "finish": str(t.getFinish()) if t.getFinish() else None,
                "percent_complete": float(t.getPercentageComplete() or 0),
                "summary": bool(t.getSummary()),
            })
        return out

    def read_links(self) -> List[Dict[str, Any]]:
        # T-future: cache predecessor lookups if MPP perf becomes an issue
        # (per-task Java->Python crossings are O(N*M); MPP read is cold-path
        # for typical Phase 4 use, most users hit XML)
        proj = self._load()
        out: List[Dict[str, Any]] = []
        for t in proj.getTasks():
            if t is None or t.getID() == 0:
                continue
            for rel in (t.getPredecessors() or []):
                lag = rel.getLag()
                out.append({
                    "from_id": int(rel.getTargetTask().getID()),
                    "to_id": int(t.getID()),
                    "type": str(rel.getType()),
                    "lag_days": _mpxj_duration_to_hours(lag) / 8.0 if lag else 0.0,
                })
        return out

    def read_resources(self) -> List[Dict[str, Any]]:
        proj = self._load()
        out: List[Dict[str, Any]] = []
        for r in proj.getResources():
            if r is None or r.getID() == 0:
                continue
            out.append({
                "id": int(r.getID()),
                "name": str(r.getName() or ""),
                "type": str(r.getType() or "Work"),
                "max_units": float(r.getMaxUnits() or 1.0),
            })
        return out

    def read_assignments(self) -> List[Dict[str, Any]]:
        proj = self._load()
        out: List[Dict[str, Any]] = []
        for a in proj.getResourceAssignments():
            t = a.getTask()
            r = a.getResource()
            if t is None or r is None:
                continue
            out.append({
                "task_id": int(t.getID()),
                "resource_id": int(r.getID()),
                "units": float(a.getUnits() or 1.0),
                "work_h": _mpxj_duration_to_hours(a.getWork()),
            })
        return out

    def read_calendars(self) -> List[Dict[str, Any]]:
        proj = self._load()
        out: List[Dict[str, Any]] = []
        for cal in proj.getCalendars():
            out.append({
                "name": str(cal.getName() or ""),
                "is_base": bool(cal.getParent() is None),
            })
        return out

    def read_baselines(self, baseline_number: int = 0) -> Dict[str, Any]:
        return {
            "baseline_number": baseline_number,
            "saved_date": None,
            "tasks": [],
            "note": "MPP baseline read via MPXJ - limited fields",
        }

    def read_progress(self) -> Dict[str, Any]:
        proj = self._load()
        tasks: List[Dict[str, Any]] = []
        for t in proj.getTasks():
            if t is None or t.getID() == 0 or t.getSummary():
                continue
            tasks.append({
                "id": int(t.getID()),
                "percent_complete": float(t.getPercentageComplete() or 0),
                "actual_work_h": _mpxj_duration_to_hours(t.getActualWork()),
            })
        try:
            props = proj.getProjectProperties()
            status_date = props.getStatusDate() if props else None
        except Exception as e:
            logger.debug(f"MPP status_date read failed: {e}")
            status_date = None
        return {
            "status_date": str(status_date) if status_date else None,
            "tasks": tasks,
        }


def _get_msp_file_manager(file_path: str):
    """Factory: returns MspdiProject for .xml/.mspdi, MspMppFileManager for .mpp.

    Performs schema check for XML — refuses non-MSPDI XML with clear error
    pointing at asta_powerproject_file MCP for Asta files.

    Note on error precedence: extension is validated FIRST (cheap, no I/O)
    so callers passing an unsupported extension always see ValueError, even
    if the path doesn't exist. Existence check follows.
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in ('.xml', '.mspdi', '.mpp'):
        raise ValueError(
            f"Unsupported extension '{ext}'. Phase 4 supports: .xml, .mspdi, .mpp"
        )
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    if ext in ('.xml', '.mspdi'):
        if not _detect_msp_xml_schema(file_path):
            raise ValueError(
                f"Not a MS Project XML - appears to be Asta or unknown schema. "
                f"For Asta files use asta_powerproject_file MCP. File: {file_path}"
            )
        return MspdiProject(file_path)
    # ext == '.mpp'
    return MspMppFileManager(file_path)


# ---------- PHASE 4 ACTION HELPERS ----------
#
# T66 PROBE FINDINGS (mspdi_parser.MspdiProject API):
#   * get_all_tasks() returns list of dicts with keys:
#       id, unique_id, name, duration (str "1d"), start, finish,
#       percent_complete, critical, milestone, summary, total_float,
#       notes, predecessors, successors
#   * predecessors / successors are lists of dicts:
#       {"task_id": int, "type": "FS"|"SS"|"FF"|"SF", "lag": "0d"}
#     (task_id resolved from UID by _id_by_uid; type/lag pre-formatted)
#   * NO get_links() method. get_link_chain(from_pat, to_pat) takes 2
#     regex strings - not suitable for "all links". Use task walk.
#   * MspMppFileManager (added T65) already returns the unified contract
#     {id, name, duration_h, start, finish, percent_complete, summary}
#     for tasks and {from_id, to_id, type, lag_days} for links.
#
# Adapter strategy: when manager is MspdiProject, normalize task dicts
# (rename unique_id, parse duration string -> hours) and walk
# predecessors lists to build the unified link list.


def _parse_duration_h(d) -> float:
    """Parse duration representations to hours.

    Accepts numeric (assume hours), ISO 8601 (PT8H0M0S), or
    Asta/MSPDI-style strings ('1d', '3w', '8h')."""
    if d is None:
        return 0.0
    if isinstance(d, (int, float)):
        return float(d)
    s = str(d).strip()
    if s.startswith('PT'):
        # ISO 8601: PT[H][M][S] — handle decimals (PT8.5H) and seconds (PT0H30M45S)
        import re
        h = re.search(r'(\d+(?:\.\d+)?)H', s)
        m = re.search(r'(\d+(?:\.\d+)?)M', s)
        sec = re.search(r'(\d+(?:\.\d+)?)S', s)
        return ((float(h.group(1)) if h else 0.0)
                + (float(m.group(1)) / 60.0 if m else 0.0)
                + (float(sec.group(1)) / 3600.0 if sec else 0.0))
    if s.endswith('d'):
        try:
            return float(s[:-1]) * 8.0
        except ValueError:
            return 0.0
    if s.endswith('w'):
        try:
            return float(s[:-1]) * 40.0
        except ValueError:
            return 0.0
    if s.endswith('h'):
        try:
            return float(s[:-1])
        except ValueError:
            return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _normalize_mspdi_task(raw: Dict[str, Any]) -> Dict[str, Any]:
    """Translate MspdiProject get_all_tasks() dict to unified Phase 4 task dict.

    Probe-confirmed keys: id, name, duration (str), start, finish,
    percent_complete, summary.
    """
    return {
        "id": raw["id"],
        "name": raw["name"],
        "duration_h": _parse_duration_h(raw.get("duration")),
        "start": raw.get("start"),
        "finish": raw.get("finish"),
        "percent_complete": float(raw.get("percent_complete", 0)),
        "summary": bool(raw.get("summary", False)),
    }


def _msp_file_read_tasks(file_path: str,
                         filters: Optional[Dict[str, Any]] = None,
                         limit: Optional[int] = None) -> Dict[str, Any]:
    """Read all tasks from a MS Project file. Format auto-detected by extension.

    Excludes summary tasks (root project + WBS summaries) for cleaner results.

    filters: simple equality match dict, e.g. ``{"name": "Foundation"}``.
        Limited to exact value match — for complex expressions use the
        ``query`` action (T69, supports operators and AND/OR).
        NOTE: summary tasks are stripped BEFORE filters apply, so
        ``filters={"summary": True}`` will always return 0 results.
    limit: cap returned task count after filtering.

    Phase 5f additive routing: file_path ending '.xer' delegates to
    XerFile.read_tasks (Phase 5d reader). Existing .xml/.mpp paths
    unchanged.
    """
    if file_path and isinstance(file_path, str) and file_path.lower().endswith(".xer"):
        try:
            from xer_parser import XerFile
            xer = XerFile(file_path)
            cals = xer.read_calendars()
            day_hr_cnt = cals[0]["day_hr_cnt"] if cals else 8.0
            tasks = xer.read_tasks(day_hr_cnt=day_hr_cnt)
            tasks = [t for t in tasks if not t.get("summary", False)]
            if filters:
                for k, v in filters.items():
                    tasks = [t for t in tasks if t.get(k) == v]
            if limit and limit > 0:
                tasks = tasks[:limit]
            return {"status": "ok", "count": len(tasks), "tasks": tasks}
        except FileNotFoundError as e:
            return {"status": "error", "error": str(e)}
        except Exception as e:
            logger.error(f"_msp_file_read_tasks({file_path}) XER failed: {e}")
            return {"status": "error", "error": str(e)}
    try:
        mgr = _get_msp_file_manager(file_path)
        if isinstance(mgr, MspdiProject):
            raw_tasks = mgr.get_all_tasks()
            tasks = [_normalize_mspdi_task(t) for t in raw_tasks]
        else:
            tasks = mgr.read_tasks()
        # Filter out summary tasks (root + WBS)
        tasks = [t for t in tasks if not t.get("summary", False)]
        # Apply optional field filters
        if filters:
            for k, v in filters.items():
                tasks = [t for t in tasks if t.get(k) == v]
        if limit and limit > 0:
            tasks = tasks[:limit]
        return {"status": "ok", "count": len(tasks), "tasks": tasks}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.error(f"_msp_file_read_tasks({file_path}) failed: {e}")
        return {"status": "error", "error": str(e)}


def _extract_links_from_mspdi(mgr: 'MspdiProject') -> List[Dict[str, Any]]:
    """Extract links from MspdiProject by walking tasks for predecessor info.

    Each task's predecessors list (probe-confirmed shape) is:
        [{"task_id": int, "type": "FS"|"SS"|"FF"|"SF", "lag": "0d"}, ...]
    where task_id is the predecessor's task ID. Returns unified contract
    {from_id, to_id, type, lag_days}.
    """
    links: List[Dict[str, Any]] = []
    raw_tasks = mgr.get_all_tasks()
    for t in raw_tasks:
        preds = t.get("predecessors") or []
        if not preds:
            continue
        to_id = t["id"]
        for p in preds:
            links.append({
                "from_id": p["task_id"],
                "to_id": to_id,
                "type": p.get("type", "FS"),
                # mspdi_parser._format_lag emits "Xd" working-day strings.
                # _parse_duration_h converts to hours via *8.0; /8.0 here
                # round-trips back to working days. The math cancels — this
                # is NOT a 5x8 calendar assumption per CLAUDE.md RULE 1.
                "lag_days": _parse_duration_h(p.get("lag", "0d")) / 8.0,
            })
    return links


def _msp_file_read_links(file_path: str) -> Dict[str, Any]:
    """Read all task predecessor/successor links from a MS Project file.

    For MspdiProject (XML): walks tasks and extracts predecessor entries.
    For MspMppFileManager (MPP): delegates to its read_links().

    Phase 5f additive routing: file_path ending '.xer' delegates to
    XerFile.read_links. Existing .xml/.mpp paths unchanged.
    """
    if file_path and isinstance(file_path, str) and file_path.lower().endswith(".xer"):
        try:
            from xer_parser import XerFile
            links = XerFile(file_path).read_links()
            return {"status": "ok", "count": len(links), "links": links}
        except FileNotFoundError as e:
            return {"status": "error", "error": str(e)}
        except Exception as e:
            logger.error(f"_msp_file_read_links({file_path}) XER failed: {e}")
            return {"status": "error", "error": str(e)}
    try:
        mgr = _get_msp_file_manager(file_path)
        if isinstance(mgr, MspdiProject):
            links = _extract_links_from_mspdi(mgr)
        else:
            links = mgr.read_links()
        return {"status": "ok", "count": len(links), "links": links}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.error(f"_msp_file_read_links({file_path}) failed: {e}")
        return {"status": "error", "error": str(e)}


def _normalize_mspdi_resource(raw: Dict[str, Any]) -> Dict[str, Any]:
    """Translate MspdiProject get_resources() dict to unified Phase 4 resource dict.

    Probe T67 confirms keys: id, unique_id, name, type, max_units,
    standard_rate, cost, calendar. id=0 = system resource (filtered upstream).
    """
    return {
        "id": raw["id"],
        "name": raw["name"],
        "type": raw.get("type", "Work"),
        "max_units": float(raw.get("max_units", 1.0)),
    }


def _normalize_mspdi_assignment(raw: Dict[str, Any]) -> Dict[str, Any]:
    """Translate MspdiProject get_resource_assignments() dict to unified Phase 4 assignment.

    Probe T67 confirms keys: task_id, task_name, resource_id, resource_name,
    units, work (Asta-style str like "1d"), cost.
    """
    return {
        "task_id": raw["task_id"],
        "resource_id": raw["resource_id"],
        "units": float(raw.get("units", 1.0)),
        "work_h": _parse_duration_h(raw.get("work")),
    }


def _normalize_mspdi_calendar(raw: Dict[str, Any]) -> Dict[str, Any]:
    """Translate MspdiProject get_calendars() dict to unified Phase 4 calendar dict.

    Probe-confirmed keys: id, name, is_base (T67 review fix exposed is_base
    via mspdi_parser widening — was previously stripped).
    """
    return {
        "name": raw["name"],
        "is_base": bool(raw.get("is_base", False)),
    }


def _msp_file_read_resources(file_path: str) -> Dict[str, Any]:
    """Read all resources from a MS Project file.

    Excludes system resource (id=0) for cleaner output. Returns count + list
    of {id, name, type, max_units}.

    Phase 5f additive routing: file_path ending '.xer' delegates to
    XerFile.read_resources. Existing .xml/.mpp paths unchanged.
    """
    if file_path and isinstance(file_path, str) and file_path.lower().endswith(".xer"):
        try:
            from xer_parser import XerFile
            resources = XerFile(file_path).read_resources()
            return {"status": "ok", "count": len(resources), "resources": resources}
        except FileNotFoundError as e:
            return {"status": "error", "error": str(e)}
        except Exception as e:
            logger.error(f"_msp_file_read_resources({file_path}) XER failed: {e}")
            return {"status": "error", "error": str(e)}
    try:
        mgr = _get_msp_file_manager(file_path)
        if isinstance(mgr, MspdiProject):
            raw_resources = mgr.get_resources()
            resources = [_normalize_mspdi_resource(r) for r in raw_resources]
        else:
            resources = mgr.read_resources()
        # Exclude system resource (id=0) if present (MspMppFileManager already
        # filters; MspdiProject does not).
        resources = [r for r in resources if r.get("id", 0) != 0]
        return {"status": "ok", "count": len(resources), "resources": resources}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.error(f"_msp_file_read_resources({file_path}) failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_file_read_assignments(file_path: str,
                               task_id: Optional[int] = None) -> Dict[str, Any]:
    """Read all task-resource assignments. Optional task_id filter.

    Phase 5f additive routing: file_path ending '.xer' delegates to
    XerFile.read_assignments. Existing .xml/.mpp paths unchanged.
    """
    if file_path and isinstance(file_path, str) and file_path.lower().endswith(".xer"):
        try:
            from xer_parser import XerFile
            assignments = XerFile(file_path).read_assignments()
            if task_id is not None:
                assignments = [a for a in assignments if a["task_id"] == task_id]
            return {"status": "ok", "count": len(assignments), "assignments": assignments}
        except FileNotFoundError as e:
            return {"status": "error", "error": str(e)}
        except Exception as e:
            logger.error(f"_msp_file_read_assignments({file_path}) XER failed: {e}")
            return {"status": "error", "error": str(e)}
    try:
        mgr = _get_msp_file_manager(file_path)
        if isinstance(mgr, MspdiProject):
            raw_asgs = mgr.get_resource_assignments()
            assignments = [_normalize_mspdi_assignment(a) for a in raw_asgs]
        else:
            assignments = mgr.read_assignments()
        if task_id is not None:
            # Normalizer guarantees task_id key (probe-confirmed required field)
            assignments = [a for a in assignments if a["task_id"] == task_id]
        return {"status": "ok", "count": len(assignments), "assignments": assignments}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.error(f"_msp_file_read_assignments({file_path}) failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_file_read_calendars(file_path: str) -> Dict[str, Any]:
    """Read all calendars defined in a MS Project file.

    Phase 5f additive routing: file_path ending '.xer' delegates to
    XerFile.read_calendars. Existing .xml/.mpp paths unchanged.
    """
    if file_path and isinstance(file_path, str) and file_path.lower().endswith(".xer"):
        try:
            from xer_parser import XerFile
            calendars = XerFile(file_path).read_calendars()
            return {"status": "ok", "count": len(calendars), "calendars": calendars}
        except FileNotFoundError as e:
            return {"status": "error", "error": str(e)}
        except Exception as e:
            logger.error(f"_msp_file_read_calendars({file_path}) XER failed: {e}")
            return {"status": "error", "error": str(e)}
    try:
        mgr = _get_msp_file_manager(file_path)
        if isinstance(mgr, MspdiProject):
            raw_cals = mgr.get_calendars()
            calendars = [_normalize_mspdi_calendar(c) for c in raw_cals]
        else:
            calendars = mgr.read_calendars()
        return {"status": "ok", "count": len(calendars), "calendars": calendars}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.error(f"_msp_file_read_calendars({file_path}) failed: {e}")
        return {"status": "error", "error": str(e)}


def _none_if_na(v: Any) -> Any:
    """Normalize MSPDI 'N/A' date sentinel to None.

    mspdi_parser._parse_date returns the literal string 'N/A' for missing
    dates. Convert to None for downstream EVM/query consistency (CLAUDE.md
    RULE 5 — date comparisons must not see 'N/A' strings).
    """
    return None if v == "N/A" else v


def _normalize_mspdi_progress(raw: Dict[str, Any]) -> Dict[str, Any]:
    """Translate MspdiProject task dict to unified Phase 4 progress dict.

    Probe T68 confirmed: get_all_tasks() exposes only `percent_complete` in
    the original parser. The widened parser (T68) now also exposes
    `actual_start`, `actual_finish`, `actual_work`, `remaining_work` from
    the underlying _tasks store (data was already parsed, just stripped by
    _task_to_list_dict — same pattern as T67 calendar `is_base` fix).

    Date fields ('N/A' sentinel from mspdi_parser) are normalized to None
    so downstream queries can use `is None` comparisons (matches MPP path
    semantics where missing dates are absent keys).
    """
    return {
        "id": raw["id"],
        "percent_complete": float(raw.get("percent_complete", 0)),
        "actual_work_h": _parse_duration_h(raw.get("actual_work")),
        "actual_start": _none_if_na(raw.get("actual_start")),
        "actual_finish": _none_if_na(raw.get("actual_finish")),
        "remaining_work_h": _parse_duration_h(raw.get("remaining_work")),
    }


def _msp_file_read_baselines(file_path: str, baseline_number: int = 0) -> Dict[str, Any]:
    """Read saved baseline data from a MS Project file (Phase 3a file integration).

    XML path: MspdiProject does not currently expose baseline parsing
    (probe T68 confirmed no baseline-related methods). Returns minimal
    contract per spec. Phase 5 may extend mspdi_parser to parse <Baseline>
    XML elements when baselines are saved in real fixtures.
    MPP path: delegates to MspMppFileManager.read_baselines (returns
    placeholder per T65 — MPXJ baseline access is limited).

    baseline_number: 0-10 (Baseline + Baseline1..Baseline10).
    """
    if baseline_number not in BASELINE_NUMBERS:
        return {"status": "error",
                "error": f"baseline_number must be 0-10, got {baseline_number}"}
    # Phase 5f additive routing: .xer baseline = target schedule (CAU pattern).
    if file_path and isinstance(file_path, str) and file_path.lower().endswith(".xer"):
        try:
            from xer_parser import XerFile
            xer = XerFile(file_path)
            proj = xer.read_project()
            cals = xer.read_calendars()
            day_hr_cnt = cals[0]["day_hr_cnt"] if cals else 8.0
            raw_tasks = xer.read_tasks(day_hr_cnt=day_hr_cnt)
            baseline_tasks = []
            for t in raw_tasks:
                if t.get("summary", False):
                    continue
                baseline_tasks.append({
                    "task_id": t["id"],
                    "id": t["id"],
                    "name": t.get("name", ""),
                    "start": t.get("start"),
                    "finish": t.get("finish"),
                    "baseline_start": t.get("start"),
                    "baseline_finish": t.get("finish"),
                    "work_h": float(t.get("duration_h") or 0),
                    "baseline_work": float(t.get("duration_h") or 0),
                })
            return {
                "status": "ok",
                "baseline_number": baseline_number,
                "saved_date": (proj or {}).get("last_recalc_date"),
                "tasks": baseline_tasks,
            }
        except FileNotFoundError as e:
            return {"status": "error", "error": str(e)}
        except Exception as e:
            logger.error(f"_msp_file_read_baselines({file_path}) XER failed: {e}")
            return {"status": "error", "error": str(e)}
    try:
        mgr = _get_msp_file_manager(file_path)
        if isinstance(mgr, MspdiProject):
            # MspdiProject has no get_baselines API (probe T68 confirmed).
            # Phase 4 returns minimal contract; Phase 5 may extend parser.
            return {
                "status": "ok",
                "baseline_number": baseline_number,
                "saved_date": None,
                "tasks": [],
                "note": "No baseline data parsed for XML files in Phase 4; "
                        "for full baseline read, use the msproject_baseline "
                        "COM tool against an open project.",
            }
        else:
            data = mgr.read_baselines(baseline_number)
            return {"status": "ok", **data}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.error(f"_msp_file_read_baselines({file_path}) failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_file_read_progress(file_path: str,
                            include_assignments: bool = False) -> Dict[str, Any]:
    """Read progress fields from a MS Project file (Phase 3b file integration).

    XML path: walks MspdiProject.get_all_tasks() and normalizes progress
    fields (percent_complete, actual_start, actual_finish, actual_work_h,
    remaining_work_h). status_date is read from get_project_summary().
    Summary tasks are excluded for clean output.
    MPP path: delegates to MspMppFileManager.read_progress.

    include_assignments: reserved parameter for future Phase 4 extension.
        Phase 4 currently parses task-level progress only; per-resource
        assignment progress is available via the separate read_assignments
        action. Accepted for forward compat with the dispatcher contract.

    Returns {status, status_date, tasks: [...]}.
    """
    _ = include_assignments  # parameter is reserved (forward compat)
    # Phase 5f additive routing: .xer delegates to XerFile.read_progress.
    if file_path and isinstance(file_path, str) and file_path.lower().endswith(".xer"):
        try:
            from xer_parser import XerFile
            prog = XerFile(file_path).read_progress()
            return {"status": "ok",
                    "status_date": prog["status_date"],
                    "tasks": prog["tasks"]}
        except FileNotFoundError as e:
            return {"status": "error", "error": str(e)}
        except Exception as e:
            logger.error(f"_msp_file_read_progress({file_path}) XER failed: {e}")
            return {"status": "error", "error": str(e)}
    try:
        mgr = _get_msp_file_manager(file_path)
        if isinstance(mgr, MspdiProject):
            tasks: List[Dict[str, Any]] = []
            for raw in mgr.get_all_tasks():
                if raw.get("summary"):
                    continue
                tasks.append(_normalize_mspdi_progress(raw))
            status_date = None
            try:
                summary = mgr.get_project_summary()
                if isinstance(summary, dict):
                    status_date = summary.get("status_date")
            except Exception as e:
                logger.debug(f"get_project_summary failed: {e}")
            return {"status": "ok", "status_date": status_date, "tasks": tasks}
        else:
            data = mgr.read_progress()
            return {"status": "ok", **data}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.error(f"_msp_file_read_progress({file_path}) failed: {e}")
        return {"status": "error", "error": str(e)}


def _safe_eval_filter(expression: str, row: Dict[str, Any],
                     code_cache: Optional[Dict[str, Any]] = None) -> bool:
    """Evaluate a simple filter expression against a row dict.

    Supports: == != < <= > >= AND OR. String literals in single/double quotes.
    Field names are dict keys (e.g., 'name', 'duration_h', 'percent_complete').

    Restricted eval — no builtins, no module access, no function calls,
    no attribute access. Only literal comparisons + boolean combinators.

    THREAT MODEL: trusted MCP user input, NOT internet-exposed. Error
    messages may include Python interpreter detail (e.g., NameError) —
    acceptable for trusted contexts only. If this helper ever becomes
    reachable from untrusted callers, sanitize errors to a generic
    'Invalid expression' first and consider upgrading the substring
    forbidden-token check to an AST-walker whitelist.

    code_cache: optional shared dict for compiled expressions. Pass the
    same dict across multiple calls to amortize compile cost (T69 review:
    O(N) compile -> O(1) per query).
    """
    # Normalize boolean operators (case-insensitive AND/OR)
    expr = expression
    for kw in (" AND ", " and "):
        expr = expr.replace(kw, " and ")
    for kw in (" OR ", " or "):
        expr = expr.replace(kw, " or ")
    # Reject dangerous patterns up front (defense in depth — empty builtins
    # already prevent most attacks but explicit reject is faster + clearer).
    # NOTE: substring match — `__` is intentionally broad (blocks all dunder
    # routes). If a future MSPDI field name contains `import`/`exec`/etc. as
    # a substring, upgrade to a word-boundary regex.
    forbidden = ("__", "import", "exec", "eval", "open(", "globals(",
                 "locals(", "compile(", "lambda", ";", ":=")
    for f in forbidden:
        if f in expression:
            raise ValueError(f"Expression contains forbidden token '{f}'")
    safe_globals = {"__builtins__": {}}
    safe_locals = {k: row.get(k) for k in row}
    try:
        # Compile once per unique expression; reuse compiled code object.
        if code_cache is not None:
            code = code_cache.get(expr)
            if code is None:
                code = compile(expr, "<filter>", "eval")
                code_cache[expr] = code
        else:
            code = compile(expr, "<filter>", "eval")
        return bool(eval(code, safe_globals, safe_locals))
    except SyntaxError as e:
        raise ValueError(f"Invalid expression syntax: {e}") from e
    except Exception as e:
        # NameError on unknown field, TypeError on bad comparison, etc.
        raise ValueError(f"Expression eval failed: {e}") from e


def _msp_file_query(file_path: str,
                    expression: str,
                    limit: Optional[int] = None,
                    include_summaries: bool = False) -> Dict[str, Any]:
    """Run an ad-hoc filter expression against tasks in a project file.

    Returns matching task list with the standard task contract fields.
    Expression syntax: simple Python-like comparisons with AND/OR operators.
    Field names are task keys (id, name, duration_h, start, finish,
    percent_complete, summary).

    AND/OR must be SPACE-DELIMITED (e.g. ``"a AND b"`` not ``"a AND b"``)
    or use lowercase Python ``and``/``or``. The forbidden-token check uses
    substring match, so field names containing ``__``/``import``/``exec``/
    etc. as substrings are rejected (current MSPDI/MPXJ field set is
    safe — none collide).

    Examples:
      "duration_h > 8 AND name == 'T2'"
      "percent_complete < 100 OR id < 100"

    include_summaries: by default, summary tasks (root project + WBS
        rollups) are excluded from the candidate set BEFORE the filter
        runs (matches ``read_tasks`` behavior). Set True to query over
        summaries too — useful when you specifically want
        ``"summary == True"``.

    Restricted eval — no function calls, imports, attribute access, or
    builtins. Use _msp_file_read_tasks for unfiltered reads. Threat model:
    trusted MCP user input. Compile cost amortized via per-call code cache.
    """
    try:
        mgr = _get_msp_file_manager(file_path)
        if isinstance(mgr, MspdiProject):
            raw_tasks = mgr.get_all_tasks()
            tasks = [_normalize_mspdi_task(t) for t in raw_tasks]
        else:
            tasks = mgr.read_tasks()
        if not include_summaries:
            tasks = [t for t in tasks if not t.get("summary", False)]
        # Compile expression once across all rows
        code_cache: Dict[str, Any] = {}
        results = []
        for t in tasks:
            try:
                if _safe_eval_filter(expression, t, code_cache=code_cache):
                    results.append(t)
            except ValueError as e:
                return {"status": "error",
                        "error": f"Invalid expression: {e}"}
        if limit and limit > 0:
            results = results[:limit]
        return {"status": "ok", "count": len(results), "results": results}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.exception(f"_msp_file_query({file_path}) failed: {e}")
        return {"status": "error", "error": str(e)}


# ---------- PHASE 4 WRITE ACTION HELPERS (T70+) ----------

def _ensure_xml_write_target(mgr) -> None:
    """Raise ValueError if mgr is not an MspdiProject (XML write only).

    .mpp is Microsoft proprietary binary — no Python library can write it.
    Convert to .xml first via MS Project Save As, or use COM tools for live ops.
    """
    if not isinstance(mgr, MspdiProject):
        raise ValueError(
            ".mpp write not supported (Microsoft proprietary binary format). "
            "Convert to .xml first via MS Project Save As, or use COM tools."
        )


def _auto_sync_to_open_msp_available() -> bool:
    """T70 stub: True if _auto_sync_to_open_msp helper is wired (T72).

    Returns False until T72 lands. Write helpers call this to decide
    whether to attempt auto-sync.
    """
    return ('_auto_sync_to_open_msp' in globals() and
            callable(globals().get('_auto_sync_to_open_msp')))


def _maybe_auto_sync(file_path: str) -> Dict[str, Any]:
    """Conditional auto-sync wrapper.

    Returns {auto_imported, reschedule_ok} populated by T72's helper if
    available, else {auto_imported: False} stub.
    """
    if _auto_sync_to_open_msp_available():
        return _auto_sync_to_open_msp(file_path)  # type: ignore[name-defined]
    return {"auto_imported": False}


def _msp_file_add_tasks(file_path: str, items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Bulk add tasks to a MS Project XML file.

    items: list of {name, duration, [start, ...]} dicts. Required: name, duration.
    MspdiProject.add_task signature uses duration_str (not duration); adapter
    translates the unified contract.

    Returns {status, count, task_ids, auto_imported, ...}.
    """
    if not items:
        return {"status": "ok", "count": 0, "task_ids": [],
                "auto_imported": False}
    try:
        mgr = _get_msp_file_manager(file_path)
        _ensure_xml_write_target(mgr)
        task_ids: List[int] = []
        for item in items:
            kwargs = {k: v for k, v in item.items()
                      if k not in ("name", "duration")}
            res = mgr.add_task(name=item["name"],
                               duration_str=item.get("duration", "1d"),
                               **kwargs)
            # mspdi_parser.add_task returns dict {task_id, uid, name, ...}
            task_ids.append(res["task_id"] if isinstance(res, dict) else int(res))
        # Overwrite original file (mspdi_parser.save() default writes a
        # timestamped sibling; we want in-place edit semantics).
        mgr.save(output_path=file_path)
        sync = _maybe_auto_sync(file_path)
        return {"status": "ok", "count": len(task_ids),
                "task_ids": task_ids, **sync}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except ValueError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.exception(f"_msp_file_add_tasks failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_file_add_links(file_path: str, items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Bulk add predecessor links between tasks.

    items: list of {from_id, to_id, [type='FS', lag='0d']} dicts.
    MspdiProject.add_link uses predecessor_id/successor_id/link_type/lag_str
    keyword names; this adapter translates from the unified contract.
    """
    if not items:
        return {"status": "ok", "count": 0, "auto_imported": False}
    try:
        mgr = _get_msp_file_manager(file_path)
        _ensure_xml_write_target(mgr)
        added = 0
        for item in items:
            mgr.add_link(predecessor_id=item["from_id"],
                         successor_id=item["to_id"],
                         link_type=item.get("type", "FS"),
                         lag_str=item.get("lag"))
            added += 1
        # Overwrite original file (mspdi_parser.save() default writes a
        # timestamped sibling; we want in-place edit semantics).
        mgr.save(output_path=file_path)
        sync = _maybe_auto_sync(file_path)
        return {"status": "ok", "count": added, **sync}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except ValueError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.exception(f"_msp_file_add_links failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_file_add_resources(file_path: str, items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Bulk add resources to a MS Project XML file.

    items: list of {name, [type='Work', max_units=1.0, standard_rate]} dicts.
    Uses MspdiProject.add_resource (T70 extension to mspdi_parser).
    """
    if not items:
        return {"status": "ok", "count": 0, "resource_ids": [],
                "auto_imported": False}
    try:
        mgr = _get_msp_file_manager(file_path)
        _ensure_xml_write_target(mgr)
        res_ids: List[int] = []
        for item in items:
            rid = mgr.add_resource(name=item["name"],
                                   type=item.get("type", "Work"),
                                   max_units=float(item.get("max_units", 1.0)),
                                   standard_rate=item.get("standard_rate"))
            res_ids.append(rid)
        # Overwrite original file (mspdi_parser.save() default writes a
        # timestamped sibling; we want in-place edit semantics).
        mgr.save(output_path=file_path)
        sync = _maybe_auto_sync(file_path)
        return {"status": "ok", "count": len(res_ids),
                "resource_ids": res_ids, **sync}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except ValueError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.exception(f"_msp_file_add_resources failed: {e}")
        return {"status": "error", "error": str(e)}


# T71 — fields key translation: unified contract -> mspdi_parser kwarg names
_TASK_UPDATE_FIELD_MAP = {
    "duration": "duration_str",
    "name": "name",
    "percent_complete": "percent_complete",
    "notes": "notes",
    "start": "start_date",
    "finish": "finish_date",
}

# Phase 9.3 — baseline fields routed through MspdiProject.write_baseline
# instead of update_task (different XML element under <Task>).
_TASK_UPDATE_BASELINE_FIELDS = (
    "baseline_start",
    "baseline_finish",
    "baseline_duration_h",
    "baseline_work_h",
)


def _msp_file_update_task(file_path: str, task_id: int,
                          fields: Dict[str, Any],
                          baseline_number: int = 0) -> Dict[str, Any]:
    """Update a single task's fields in an XML file.

    Args:
        fields: unified-contract dict. Schedule keys (duration, name,
            percent_complete, notes, start, finish) routed through
            MspdiProject.update_task. Phase 9.3 — baseline keys
            (baseline_start, baseline_finish, baseline_duration_h,
            baseline_work_h) routed through MspdiProject.write_baseline
            for the same task_id.
        baseline_number: 0=primary baseline, 1-10=numbered (Phase 9.3).

    Returns:
        {status, task_id, schedule_updated, baseline_written,
         auto_imported, reschedule_ok}.
    """
    if not isinstance(fields, dict) or not fields:
        return {"status": "error", "error": "fields dict is required"}
    try:
        mgr = _get_msp_file_manager(file_path)
        _ensure_xml_write_target(mgr)
        schedule_kwargs: Dict[str, Any] = {}
        baseline_entry: Dict[str, Any] = {}
        valid_keys = (set(_TASK_UPDATE_FIELD_MAP.keys())
                      | set(_TASK_UPDATE_BASELINE_FIELDS))
        for k, v in fields.items():
            if k in _TASK_UPDATE_FIELD_MAP:
                schedule_kwargs[_TASK_UPDATE_FIELD_MAP[k]] = v
            elif k in _TASK_UPDATE_BASELINE_FIELDS:
                baseline_entry[k] = v
            else:
                return {"status": "error",
                        "error": (f"Unknown field '{k}'. Valid: "
                                  f"{sorted(valid_keys)}")}
        # Apply schedule update first (mirrors prior behavior)
        schedule_updated = False
        if schedule_kwargs:
            result = mgr.update_task(task_id=task_id, **schedule_kwargs)
            if isinstance(result, dict) and "error" in result:
                return {"status": "error", "error": result["error"]}
            schedule_updated = True
        # Phase 9.3 — apply baseline update via write_baseline
        baseline_written = 0
        if baseline_entry:
            uid = mgr._uid_by_id.get(task_id)
            if uid is None:
                return {"status": "error",
                        "error": f"task_id {task_id} not found"}
            baseline_entry["task_uid"] = uid
            baseline_written = mgr.write_baseline(baseline_number,
                                                   [baseline_entry])
        if not schedule_updated and baseline_written == 0:
            return {"status": "error",
                    "error": "no fields applied (task_id may not exist)"}
        # Phase 10.1 — read-back the task's baseline after write
        baseline_after = None
        if baseline_written > 0:
            for bl in mgr.read_baselines(baseline_number):
                if bl.get("task_id") == task_id:
                    baseline_after = bl
                    break
        mgr.save(output_path=file_path)
        sync = _maybe_auto_sync(file_path)
        return {"status": "ok", "task_id": task_id,
                "schedule_updated": schedule_updated,
                "baseline_written": baseline_written,
                "baseline_after": baseline_after,
                **sync}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except ValueError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.exception(f"_msp_file_update_task failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_file_save_as(file_path: str, output_path: str) -> Dict[str, Any]:
    """Save an XML project to a new path. Source file unchanged.

    output_path must end in .xml or .mspdi.
    """
    ext = os.path.splitext(output_path)[1].lower()
    if ext not in ('.xml', '.mspdi'):
        return {"status": "error",
                "error": f"output_path must end in .xml or .mspdi (got '{ext}')"}
    try:
        mgr = _get_msp_file_manager(file_path)
        _ensure_xml_write_target(mgr)
        mgr.save(output_path=output_path)
        size = os.path.getsize(output_path)
        return {"status": "ok", "output_path": output_path, "size_bytes": size}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except ValueError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.exception(f"_msp_file_save_as failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_file_write_baseline(file_path: str = None,
                              baseline_number: int = 0,
                              baseline_data: List[Dict[str, Any]] = None,
                              output_path: str = None) -> Dict[str, Any]:
    """Phase 8.2 + Phase 9.1 — Write baseline elements to MSPDI XML.

    Exposes Phase 6.3 MspdiProject.write_baseline + save() through the
    file MCP using the standard Phase 4 manager pattern. Requires
    output_path (no in-place write — caller controls naming).

    Phase 9.1: integrates `_maybe_auto_sync(output_path)` so that if
    MS Project is open and a project's FullName matches output_path,
    the modified file is auto-reloaded (FileClose + FileOpen +
    Reschedule). This is default behavior, not opt-in — matches the
    file_mcp_auto_sync feedback rule.

    Args:
        file_path: source MSPDI (.xml/.mspdi). .mpp not supported
            (Microsoft proprietary binary).
        baseline_number: 0=primary baseline, 1-10=numbered baselines.
        baseline_data: list of {task_uid (required), baseline_start,
            baseline_finish, baseline_duration_h, baseline_work_h}.
            Tasks with unknown UID are skipped silently.
        output_path: target .xml/.mspdi path.

    Returns:
        {status, tasks_written, output_path, baseline_number,
         auto_imported, reschedule_ok}.
    """
    if not file_path:
        return {"status": "error", "error": "file_path required"}
    if not output_path:
        return {"status": "error", "error": "output_path required"}
    if not baseline_data:
        return {"status": "error",
                "error": "baseline_data required (non-empty list)"}
    src_ext = os.path.splitext(file_path)[1].lower()
    if src_ext not in (".xml", ".mspdi"):
        return {"status": "error",
                "error": (f"write_baseline supports .xml/.mspdi only "
                          f"(got '{src_ext}'). .mpp/.xer not supported.")}
    out_ext = os.path.splitext(output_path)[1].lower()
    if out_ext not in (".xml", ".mspdi"):
        return {"status": "error",
                "error": (f"output_path must end in .xml or .mspdi "
                          f"(got '{out_ext}')")}
    try:
        mgr = _get_msp_file_manager(file_path)
        _ensure_xml_write_target(mgr)
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except ValueError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"failed to open: {e}"}
    try:
        n = mgr.write_baseline(baseline_number, baseline_data)
    except Exception as e:
        return {"status": "error", "error": f"write_baseline failed: {e}"}
    try:
        actual_path = mgr.save(output_path)
    except Exception as e:
        return {"status": "error", "error": f"save failed: {e}"}
    sync = _maybe_auto_sync(actual_path)
    return {
        "status": "ok",
        "tasks_written": n,
        "output_path": actual_path,
        "baseline_number": baseline_number,
        "auto_imported": sync.get("auto_imported", False),
        "reschedule_ok": sync.get("reschedule_ok"),
    }


def _msp_file_bulk_add_assignments(file_path: str,
                                   items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """🚀 T73 HERO — bulk write task-resource assignments to MSPDI XML.

    Phase 4 success gate: 2800 assignments in <5s via single XML write
    pass + auto-sync. Pure-Python path; no per-call COM crossing.

    items: list of {task_id, resource_id, [units]} dicts.
    Returns {status, count, elapsed_s, auto_imported, ...}.
    """
    import time
    if not items:
        return {"status": "ok", "count": 0, "elapsed_s": 0.0,
                "auto_imported": False}
    start = time.time()
    try:
        mgr = _get_msp_file_manager(file_path)
        _ensure_xml_write_target(mgr)
        added = mgr.bulk_add_assignments(items)
        mgr.save(output_path=file_path)
        elapsed = time.time() - start
        sync = _maybe_auto_sync(file_path)
        return {"status": "ok", "count": added,
                "elapsed_s": round(elapsed, 3), **sync}
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e),
                "elapsed_s": round(time.time() - start, 3)}
    except ValueError as e:
        return {"status": "error", "error": str(e),
                "elapsed_s": round(time.time() - start, 3)}
    except Exception as e:
        logger.exception(f"_msp_file_bulk_add_assignments failed: {e}")
        return {"status": "error", "error": str(e),
                "elapsed_s": round(time.time() - start, 3)}


def _auto_sync_to_open_msp(modified_xml_path: str) -> Dict[str, Any]:
    """Auto-sync a modified MSPDI XML into the open MS Project active project.

    SAFETY semantics (Phase 4 conservative):
      - If MSP COM unavailable → return {auto_imported: False, msg: ...}.
        Caller's XML is on disk; user can manually open later.
      - If MSP open but no project matches file_path (by FullName) →
        skip with auto_imported=False. Never touch unrelated projects.
      - If MSP open AND a project's FullName matches file_path →
        FileClose(0) + FileOpen(file_path) for clean reload, then
        ActiveProject.Reschedule().

    This avoids the EditCopy/EditPaste merge complexity of
    _msp_task_bulk_add_mspdi (which is for live COM operations on the
    active project) — file MCP semantics are simpler: the file changed,
    so MSP should reload it if it has the same file open.

    Memory: feedback_file_mcp_auto_sync.md — write -> open -> auto import.
    """
    if not os.path.exists(modified_xml_path):
        return {"auto_imported": False,
                "error": f"XML not found: {modified_xml_path}"}
    try:
        pythoncom.CoInitialize()
    except Exception:
        pass
    try:
        try:
            app = win32com.client.GetActiveObject('MSProject.Application')
        except Exception as e:
            return {"auto_imported": False,
                    "msg": f"MSP closed; XML saved at {modified_xml_path}",
                    "error": str(e)}
        try:
            normalized = modified_xml_path.replace("\\", "/").lower()
            matching_proj = None
            try:
                count = app.Projects.Count
            except Exception as e:
                logger.debug(f"app.Projects.Count failed: {e}")
                count = 0
            for i in range(1, count + 1):
                try:
                    proj = app.Projects(i)
                    if proj is None:
                        continue
                    full_name = (proj.FullName or "").replace("\\", "/").lower()
                    if full_name == normalized:
                        matching_proj = proj
                        break
                except Exception as e:
                    logger.debug(f"Project enumeration failed at {i}: {e}")
                    continue
            if matching_proj is None:
                return {"auto_imported": False,
                        "msg": ("No matching project open in MSP; XML saved at "
                                f"{modified_xml_path}. SAFETY: file MCP never "
                                "merges into unrelated projects.")}
            # Match found: close (no save — file already on disk) + reopen
            try:
                app.WindowActivate(matching_proj.Windows(1).Caption)
                app.FileClose(0)  # 0 = pjDoNotSave
            except Exception as e:
                return {"auto_imported": False,
                        "error": f"Could not close matching project: {e}"}
            try:
                app.FileOpen(modified_xml_path)
            except Exception as e:
                return {"auto_imported": False,
                        "error": f"FileOpen of modified XML failed: {e}"}
            reschedule_ok = False
            try:
                if app.ActiveProject is not None:
                    app.ActiveProject.Reschedule()
                    reschedule_ok = True
            except Exception as e:
                logger.warning(f"Reschedule failed after auto-sync: {e}")
            return {"auto_imported": True, "reschedule_ok": reschedule_ok}
        except Exception as e:
            logger.exception(f"_auto_sync_to_open_msp failed: {e}")
            return {"auto_imported": False, "error": str(e)}
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


@mcp.tool(
    name="msproject_file",
    annotations={"title": "MS Project File-Based Operations", "readOnlyHint": False},
)
async def msproject_file(params: dict) -> str:
    """File-based read+write for MS Project files (.xml/.mspdi/.mpp).

    Actions:
    Read (8): read_tasks, read_links, read_resources, read_assignments,
              read_calendars, read_baselines, read_progress, query
    Write (6): add_tasks, add_links, add_resources, bulk_add_assignments,
               update_task, save_as

    All actions require file_path. .xml/.mspdi via native Python parser
    (zero Java); .mpp via MPXJ + JVM (lazy init).
    Write actions: if MSP open AND a project's FullName matches file_path,
    auto FileClose+FileOpen+Reschedule for clean reload (default — not
    opt-in). MSP closed or unrelated projects → XML on disk only.

    Phase 4 (30 Apr 2026). HERO: bulk_add_assignments 2800 in <5s.
    """
    import json
    action = params.get("action", "")
    p = {k: v for k, v in params.items() if k != "action"}
    try:
        if action == "read_tasks":
            r = _msp_file_read_tasks(**p)
        elif action == "read_links":
            r = _msp_file_read_links(**p)
        elif action == "read_resources":
            r = _msp_file_read_resources(**p)
        elif action == "read_assignments":
            r = _msp_file_read_assignments(**p)
        elif action == "read_calendars":
            r = _msp_file_read_calendars(**p)
        elif action == "read_baselines":
            r = _msp_file_read_baselines(**p)
        elif action == "read_progress":
            r = _msp_file_read_progress(**p)
        elif action == "query":
            r = _msp_file_query(**p)
        elif action == "add_tasks":
            r = _msp_file_add_tasks(**p)
        elif action == "add_links":
            r = _msp_file_add_links(**p)
        elif action == "add_resources":
            r = _msp_file_add_resources(**p)
        elif action == "bulk_add_assignments":
            r = _msp_file_bulk_add_assignments(**p)
        elif action == "update_task":
            r = _msp_file_update_task(**p)
        elif action == "save_as":
            r = _msp_file_save_as(**p)
        elif action == "write_baseline":
            r = _msp_file_write_baseline(**p)
        else:
            r = {"status": "error",
                 "error": (f"Unknown action '{action}'. Valid: read_tasks/"
                           "read_links/read_resources/read_assignments/"
                           "read_calendars/read_baselines/read_progress/"
                           "query/add_tasks/add_links/add_resources/"
                           "bulk_add_assignments/update_task/save_as/"
                           "write_baseline")}
    except TypeError as e:
        r = {"status": "error", "error": f"Invalid params for {action}: {e}"}
    except Exception as e:
        logger.error(f"msproject_file({action}) failed: {e}")
        r = {"status": "error", "error": str(e)}
    return json.dumps(r, default=str, ensure_ascii=False)


# ============================================================================
# PHASE 5A — EVM TOOL
# ============================================================================
import datetime as _dt5
from evm_math import (
    compute_metrics as _evm_compute,
    forecast as _evm_forecast,
    earned_schedule as _evm_earned_schedule,
    time_phased_pv as _evm_tp_pv,
    time_phased_ev as _evm_tp_ev,
    time_phased_ac as _evm_tp_ac,
    time_phased_ac_increments as _evm_tp_ac_inc,
    period_delta as _evm_period_delta,
    progress_data_quality as _evm_pdq,
    rag_status as _evm_rag,
)


def _parse_iso_date(s):
    """Parse '2026-01-01...' or 'YYYY-MM-DD'-prefix string to date.

    Returns None for None/empty/'N/A'/unparseable input.
    """
    if not s or s == "N/A":
        return None
    try:
        return _dt5.date.fromisoformat(str(s)[:10])
    except (ValueError, TypeError):
        return None


def _evm_load_task_data(file_path=None):
    """Hybrid: file_path -> Phase 4 file path; None -> Phase 1 COM path.

    Returns {status, tasks: [...], resources: [...], assignments: [...],
             status_date, project_name, project_file}.
    Each task dict has at least: id, name, duration_h, baseline_start,
    baseline_finish, baseline_work, percent_complete, actual_work, summary.

    Phase 5e additive routing: file_path ending '.xer' delegates to
    _xer_to_evm_task_shape (Phase 5e adapter, defined later in module).
    Existing .xml/.mpp/COM paths unchanged.
    """
    try:
        if file_path and isinstance(file_path, str) and file_path.lower().endswith(".xer"):
            from xer_parser import XerFile
            return _xer_to_evm_task_shape(XerFile(file_path))
        if file_path:
            tr = _msp_file_read_tasks(file_path=file_path)
            if tr.get("status") != "ok":
                return tr
            rr = _msp_file_read_resources(file_path=file_path)
            ar = _msp_file_read_assignments(file_path=file_path)
            pr = _msp_file_read_progress(file_path=file_path)
            # Merge baseline fields and progress fields into task dicts
            tasks = tr.get("tasks", [])
            # Read baseline 0 for default PV
            br = _msp_file_read_baselines(file_path=file_path, baseline_number=0)
            baseline_tasks = {bt.get("task_id"): bt
                             for bt in (br.get("tasks", []) if br.get("status") == "ok" else [])}
            progress_tasks = {pt.get("id"): pt
                             for pt in (pr.get("tasks", []) if pr.get("status") == "ok" else [])}
            for t in tasks:
                tid = t.get("id")
                bt = baseline_tasks.get(tid, {})
                pt = progress_tasks.get(tid, {})
                # Baseline fields: try a few common names
                t.setdefault("baseline_start", bt.get("start") or bt.get("baseline_start"))
                t.setdefault("baseline_finish", bt.get("finish") or bt.get("baseline_finish"))
                t.setdefault("baseline_work", bt.get("work_h") or bt.get("baseline_work") or t.get("duration_h", 0))
                t.setdefault("percent_complete", t.get("percent_complete") or pt.get("percent_complete") or 0)
                t.setdefault("actual_work", pt.get("actual_work_h") or 0)
            return {
                "status": "ok",
                "tasks": tasks,
                "resources": rr.get("resources", []) if rr.get("status") == "ok" else [],
                "assignments": ar.get("assignments", []) if ar.get("status") == "ok" else [],
                "status_date": pr.get("status_date") if pr.get("status") == "ok" else None,
                "project_file": file_path,
            }
        # COM path
        app = _validate_active_project()
        proj = app.ActiveProject
        tasks = []
        for i in range(1, proj.Tasks.Count + 1):
            t = proj.Tasks(i)
            if t is None:
                continue
            tasks.append({
                "id": t.ID,
                "name": t.Name or "",
                "duration_h": float(t.Duration or 0) / 60.0,
                "start": str(t.Start) if t.Start else None,
                "finish": str(t.Finish) if t.Finish else None,
                "percent_complete": float(t.PercentComplete or 0),
                "summary": bool(t.Summary),
                "baseline_start": str(t.BaselineStart) if t.BaselineStart else None,
                "baseline_finish": str(t.BaselineFinish) if t.BaselineFinish else None,
                "baseline_work": float(t.BaselineWork or 0) / 60.0,
                "actual_work": float(t.ActualWork or 0) / 60.0,
            })
        resources = []
        for i in range(1, proj.Resources.Count + 1):
            r = proj.Resources(i)
            if r is None:
                continue
            resources.append({
                "id": r.ID,
                "name": r.Name or "",
                "type": "Work",
                "max_units": float(r.MaxUnits or 1.0),
            })
        try:
            status_date = str(proj.StatusDate) if proj.StatusDate else None
        except Exception:
            status_date = None
        return {
            "status": "ok",
            "tasks": [t for t in tasks if not t["summary"]],
            "resources": resources,
            "assignments": [],  # COM path skips for perf
            "status_date": status_date,
            "project_name": proj.Name,
        }
    except Exception as e:
        logger.exception(f"_evm_load_task_data failed: {e}")
        return {"status": "error", "error": str(e)}


def _evm_load_progress_data(file_path=None):
    """Read progress fields (percent_complete, actual_work, status_date).

    file_path -> Phase 4 _msp_file_read_progress
    None      -> Phase 1 _msp_progress_summary (existing helper)
    """
    if file_path:
        return _msp_file_read_progress(file_path=file_path)
    # COM path — Phase 3b summary already provides BAC/ACWP/StatusDate
    return _msp_progress_summary()


def _evm_load_baseline_data(file_path=None, baseline_number=0):
    """Phase 5e additive routing: file_path ending '.xer' delegates to
    _xer_to_evm_baseline_shape (Phase 5e adapter, defined after Phase 5d).
    Existing .xml/.mpp/COM paths unchanged.
    """
    if file_path and isinstance(file_path, str) and file_path.lower().endswith(".xer"):
        from xer_parser import XerFile
        return _xer_to_evm_baseline_shape(XerFile(file_path), baseline_number)
    return _evm_load_baseline_data_impl(file_path, baseline_number)


def _evm_load_baseline_data_impl(file_path=None, baseline_number=0):
    """Read baseline data per Phase 3a. Validates baseline_number 0-10."""
    if baseline_number not in BASELINE_NUMBERS:
        return {"status": "error",
                "error": f"baseline_number must be 0-10, got {baseline_number}"}
    if file_path:
        return _msp_file_read_baselines(file_path=file_path,
                                       baseline_number=baseline_number)
    # COM path
    app = _validate_active_project()
    proj = app.ActiveProject
    saved = _baseline_saved_date(proj, baseline_number)
    tasks_baseline = []
    for i in range(1, proj.Tasks.Count + 1):
        t = proj.Tasks(i)
        if t is None or t.Summary:
            continue
        b = _read_task_baseline(t, baseline_number)
        b["task_id"] = t.ID
        tasks_baseline.append(b)
    return {
        "status": "ok",
        "baseline_number": baseline_number,
        "saved_date": str(saved) if saved else None,
        "tasks": tasks_baseline,
    }


def _evm_detect_currency_mode(tasks, resources):
    """RULE 3 — hours vs cost loading (legacy 2-mode return).

    Phase 6.1: delegates to currency_validator pure module. Maps 4-mode
    output -> 2-mode for backward compat with existing dispatcher action
    `detect_currency_mode` and any caller expecting 'cost'/'hours' only:
        cost / mixed       -> 'cost' (cost data present)
        hours / uncertain  -> 'hours' (no cost data)
    """
    from currency_validator import detect_mode_from_tasks_resources
    mode = detect_mode_from_tasks_resources(tasks, resources)
    return "cost" if mode in ("cost", "mixed") else "hours"


def _evm_compute_pv_ev_ac(load_data, baseline_load):
    """Aggregate BAC/EV/AC + linear-distributed PV at data_date.

    BAC = sum(baseline_work) across non-summary tasks
    PV  = linear distribution at data_date per RULE 5
    EV  = sum(baseline_work x percent_complete / 100)
    AC  = sum(actual_work)

    Currency-agnostic: caller chose units (hours or cost) upstream.
    """
    tasks = load_data.get("tasks", []) or []
    bac = sum(float(t.get("baseline_work") or 0) for t in tasks)
    ev = sum(float(t.get("baseline_work") or 0) *
             float(t.get("percent_complete") or 0) / 100.0
             for t in tasks)
    ac = sum(float(t.get("actual_work") or 0) for t in tasks)
    # PV at data_date - use linear distribution per task
    sd_str = load_data.get("status_date")
    data_date = _parse_iso_date(sd_str) if sd_str else _dt5.date.today()
    enriched = []
    for t in tasks:
        bs = _parse_iso_date(t.get("baseline_start"))
        bf = _parse_iso_date(t.get("baseline_finish"))
        if bs is None or bf is None:
            continue
        enriched.append({
            "baseline_start": bs,
            "baseline_finish": bf,
            "baseline_work": float(t.get("baseline_work") or 0),
        })
    if enriched:
        pv = _evm_tp_pv(enriched, [(_dt5.date.min, data_date)])[0]
    else:
        pv = 0.0
    return bac, pv, ev, ac


def _msp_evm_compute_metrics(file_path=None, baseline_number=0):
    """Action 1: compute_metrics - SPI/CPI/SV/CV (RULE 4)."""
    load = _evm_load_task_data(file_path=file_path)
    if load.get("status") != "ok":
        return load
    bload = _evm_load_baseline_data(file_path=file_path, baseline_number=baseline_number)
    if bload.get("status") != "ok":
        return bload
    bac, pv, ev, ac = _evm_compute_pv_ev_ac(load, bload)
    metrics = _evm_compute(bac=bac, pv=pv, ev=ev, ac=ac)
    return {"status": "ok", "baseline_number": baseline_number, **metrics}


def _msp_evm_forecast(file_path=None, baseline_number=0):
    """Action 2: forecast - EAC1/2/3, ETC, VAC, TCPI (RULE 9)."""
    cm = _msp_evm_compute_metrics(file_path=file_path, baseline_number=baseline_number)
    if cm.get("status") != "ok":
        return cm
    fc = _evm_forecast(bac=cm["bac"], ev=cm["ev"], ac=cm["ac"],
                      cpi=cm.get("cpi"), spi=cm.get("spi"))
    return {"status": "ok", "baseline_number": baseline_number, **fc}


def _msp_evm_summary(file_path=None, baseline_number=0):
    """Action 4: summary - RAG (RULE 12) + executive."""
    cm = _msp_evm_compute_metrics(file_path=file_path, baseline_number=baseline_number)
    if cm.get("status") != "ok":
        return cm
    completion_pct = (cm["ev"] / cm["bac"] * 100.0) if cm["bac"] > 0 else 0.0
    rag = _evm_rag(spi=cm.get("spi"), completion_pct=completion_pct)
    return {
        "status": "ok",
        "baseline_number": baseline_number,
        "rag": rag,
        "completion_pct": round(completion_pct, 2),
        "spi": cm.get("spi"),
        "cpi": cm.get("cpi"),
        "schedule_health": rag,
    }


def _evm_build_pv_curve(tasks, project_start, project_finish, bucket="week"):
    """Build cumulative PV curve points across project duration.

    tasks: list of task dicts with baseline_start/finish/work
    project_start, project_finish: date or datetime
    bucket: 'day'|'week'|'month' (default 'week')

    Returns list of (date, cumulative_pv) tuples.
    """
    enriched = []
    for t in tasks:
        bs = _parse_iso_date(t.get("baseline_start"))
        bf = _parse_iso_date(t.get("baseline_finish"))
        if bs is None or bf is None:
            continue
        enriched.append({
            "baseline_start": bs, "baseline_finish": bf,
            "baseline_work": float(t.get("baseline_work") or 0),
        })
    delta = _dt5.timedelta(days=7) if bucket == "week" else \
            _dt5.timedelta(days=1) if bucket == "day" else \
            _dt5.timedelta(days=30)
    points = []
    if hasattr(project_start, "date"):
        project_start = project_start.date()
    if hasattr(project_finish, "date"):
        project_finish = project_finish.date()
    d = project_start
    while d <= project_finish:
        d += delta
        if enriched:
            pv_now = _evm_tp_pv(enriched, [(_dt5.date.min, d)])[0]
        else:
            pv_now = 0.0
        points.append((d, pv_now))
    return points


def _evm_derive_project_bounds(tasks):
    """Derive (project_start, project_finish) from baseline or current dates.

    Falls back to start/finish if baseline not available.
    Returns (None, None) if no usable dates anywhere.
    """
    starts = []
    finishes = []
    for t in tasks:
        bs = _parse_iso_date(t.get("baseline_start") or t.get("start"))
        bf = _parse_iso_date(t.get("baseline_finish") or t.get("finish"))
        if bs:
            starts.append(bs)
        if bf:
            finishes.append(bf)
    if not starts or not finishes:
        return None, None
    return min(starts), max(finishes)


def _msp_evm_earned_schedule(file_path=None, baseline_number=0, bucket="week"):
    """Action 3: earned_schedule (RULE 8 Lipke 2003)."""
    load = _evm_load_task_data(file_path=file_path)
    if load.get("status") != "ok":
        return load
    tasks = load.get("tasks", []) or []
    if not tasks:
        return {"status": "error", "error": "No tasks loaded"}
    project_start, project_finish = _evm_derive_project_bounds(tasks)
    if project_start is None or project_finish is None:
        return {"status": "error", "error": "Cannot determine project bounds"}
    sd_str = load.get("status_date")
    data_date = _parse_iso_date(sd_str) if sd_str else _dt5.date.today()
    if data_date is None:
        data_date = _dt5.date.today()
    # Compute current EV
    ev = sum(float(t.get("baseline_work") or 0) *
             float(t.get("percent_complete") or 0) / 100.0
             for t in tasks)
    # Build PV curve
    pv_curve = _evm_build_pv_curve(tasks, project_start, project_finish, bucket)
    es = _evm_earned_schedule(pv_curve=pv_curve, ev_now=ev,
                              project_start=project_start, data_date=data_date)
    return {"status": "ok", "baseline_number": baseline_number,
            "bucket": bucket, **es}


def _msp_evm_progress_data_quality(file_path=None, baseline_number=0):
    """Action 7: progress_data_quality (RULE 7)."""
    load = _evm_load_task_data(file_path=file_path)
    if load.get("status") != "ok":
        return load
    cm = _msp_evm_compute_metrics(file_path=file_path, baseline_number=baseline_number)
    es = _msp_evm_earned_schedule(file_path=file_path,
                                  baseline_number=baseline_number)
    spi_h = cm.get("spi") if cm.get("status") == "ok" else None
    spi_t = es.get("spi_t") if es.get("status") == "ok" else None
    completion_pct = (cm.get("ev", 0) / cm["bac"] * 100.0) if cm.get("bac", 0) > 0 else 0
    has_resources = len(load.get("resources", []) or []) > 0
    warnings = _evm_pdq(spi_h=spi_h, spi_t=spi_t,
                       completion_pct=completion_pct, has_resources=has_resources)
    return {"status": "ok", "warnings": warnings,
            "spi_h": spi_h, "spi_t": spi_t,
            "completion_pct": round(completion_pct, 2)}


def _msp_evm_detect_currency_mode(file_path=None):
    """Action 13: detect_currency_mode (RULE 3)."""
    load = _evm_load_task_data(file_path=file_path)
    if load.get("status") != "ok":
        return load
    mode = _evm_detect_currency_mode(load.get("tasks", []),
                                    load.get("resources", []))
    return {"status": "ok", "mode": mode}


def _msp_evm_validate_currency_mode(file_path=None):
    """Phase 6.1 Action 14: validate_currency_mode (RULE 3, multi-source).

    Cross-validates currency mode across sources:
        - tasks + resources cost fields (all formats)
        - XER TASKRSRC assignments target_cost/target_qty pattern (XER only)
        - XER ERMHDR.currency code (XER only)

    4-mode primary output: 'cost'|'hours'|'mixed'|'uncertain'.

    Returns:
        {
          status,
          primary_mode,
          currency_code: 'USD'|None,
          cross_validation: {consensus_mode, confidence, conflicts,
                             warnings, source_counts},
          sources: {tasks_resources, xer_assignments, currency_header},
        }
    """
    from currency_validator import (
        detect_mode_from_xer_assignments,
        detect_mode_from_tasks_resources,
        extract_currency_code,
        cross_validate_modes,
    )
    load = _evm_load_task_data(file_path=file_path)
    if load.get("status") != "ok":
        return load
    sources = []
    # Source 1: tasks + resources (all formats)
    tr_mode = detect_mode_from_tasks_resources(
        load.get("tasks", []), load.get("resources", []))
    sources.append(("tasks_resources", tr_mode))
    # Source 2: XER assignments RULE 3 pattern (XER only)
    xer_assignments_mode = None
    currency_code = None
    is_xer = (file_path and isinstance(file_path, str)
              and file_path.lower().endswith(".xer"))
    if is_xer:
        from xer_parser import XerFile
        try:
            xf = XerFile(file_path)
            assignments = xf.read_assignments()
            xer_assignments_mode = detect_mode_from_xer_assignments(assignments)
            sources.append(("xer_assignments", xer_assignments_mode))
            currency_code = extract_currency_code(xf.header_fields)
        except Exception as e:
            logger.warning(f"XER currency validation skipped: {e}")
    cv = cross_validate_modes(sources)
    primary_mode = cv["consensus_mode"]
    # Fallback: if consensus uncertain but tr_mode is decisive, use it
    if primary_mode == "uncertain" and tr_mode != "uncertain":
        primary_mode = tr_mode
    return {
        "status": "ok",
        "primary_mode": primary_mode,
        "currency_code": currency_code,
        "cross_validation": cv,
        "sources": {
            "tasks_resources": tr_mode,
            "xer_assignments": xer_assignments_mode,
            "currency_header": currency_code,
        },
    }


def _evm_bucket_to_delta(bucket):
    """Map bucket name to timedelta. Returns None for invalid."""
    if bucket == "day":
        return _dt5.timedelta(days=1)
    if bucket == "week":
        return _dt5.timedelta(days=7)
    if bucket == "month":
        return _dt5.timedelta(days=30)
    return None


def _msp_evm_time_phased_evm(file_path=None, baseline_number=0, bucket="week"):
    """Action 5: time_phased_evm — PV/EV/AC per period.

    Returns {status, bucket, buckets: [{period_start, period_end, pv, ev, ac}]}.
    Validates bucket: day/week/month only.

    AC simplification: total AC distributed evenly across past buckets up to
    data_date. True per-period AC requires Phase 3b time_phased_actual_read
    for COM path or per-assignment work distribution for file path —
    deferred to Phase 6.
    """
    delta = _evm_bucket_to_delta(bucket)
    if delta is None:
        return {"status": "error",
                "error": f"bucket must be day/week/month, got '{bucket}'"}
    load = _evm_load_task_data(file_path=file_path)
    if load.get("status") != "ok":
        return load
    tasks = load.get("tasks", []) or []
    if not tasks:
        return {"status": "ok", "bucket": bucket, "buckets": []}
    project_start, project_finish = _evm_derive_project_bounds(tasks)
    if project_start is None or project_finish is None:
        return {"status": "ok", "bucket": bucket, "buckets": []}
    sd_str = load.get("status_date")
    data_date = _parse_iso_date(sd_str) if sd_str and sd_str != "N/A" else None
    if data_date is None:
        data_date = _dt5.date.today()
    # Build buckets and compute PV/EV
    buckets = []
    d = project_start
    while d <= project_finish:
        next_d = d + delta
        buckets.append((d, min(next_d, project_finish + _dt5.timedelta(days=1))))
        d = next_d
    enriched = []
    for t in tasks:
        bs = _parse_iso_date(t.get("baseline_start"))
        bf = _parse_iso_date(t.get("baseline_finish"))
        if bs is None or bf is None:
            continue
        enriched.append({
            "baseline_start": bs, "baseline_finish": bf,
            "baseline_work": float(t.get("baseline_work") or 0),
            "percent_complete": float(t.get("percent_complete") or 0),
            # Phase 6.2 — actual fields for per-task AC distribution
            "actual_start": _parse_iso_date(t.get("actual_start")),
            "actual_finish": _parse_iso_date(t.get("actual_finish")),
            "actual_work": float(t.get("actual_work") or 0),
        })
    pv = _evm_tp_pv(enriched, [(s, e) for (s, e) in buckets])
    ev = _evm_tp_ev(enriched, [(s, e) for (s, e) in buckets], data_date=data_date)
    # Phase 6.2 — true per-task AC distribution (replaces uniform total/past)
    ac = _evm_tp_ac(enriched, [(s, e) for (s, e) in buckets], data_date=data_date)
    # Phase 9.2 — per-bucket AC delta (non-cumulative)
    ac_inc = _evm_tp_ac_inc(enriched, [(s, e) for (s, e) in buckets],
                             data_date=data_date)
    out = []
    for i, (s, e) in enumerate(buckets):
        out.append({
            "period_start": s.isoformat(),
            "period_end": e.isoformat(),
            "pv": round(pv[i], 2),
            "ev": round(ev[i], 2),
            "ac": round(ac[i], 2),
            "ac_increment": round(ac_inc[i], 2),
        })
    return {"status": "ok", "bucket": bucket, "buckets": out}


def _msp_evm_period_delta(file_path=None, baseline_number=0, snapshot_path=None):
    """Action 6: period_delta vs prev snapshot (RULE 6).

    Loads previous snapshot from snapshot_path JSON; computes delta vs
    current state. If snapshot_path missing or no prev snapshots,
    returns first-period semantics (period_* = current cum values).
    """
    cm = _msp_evm_compute_metrics(file_path=file_path, baseline_number=baseline_number)
    if cm.get("status") != "ok":
        return cm
    snap_now = {"pv": cm["pv"], "ev": cm["ev"], "ac": cm["ac"], "bac": cm["bac"]}
    snap_prev = None
    if snapshot_path and os.path.exists(snapshot_path):
        try:
            import json as _json
            with open(snapshot_path, "r", encoding="utf-8") as f:
                data = _json.load(f)
            snaps = sorted(data.get("snapshots", []),
                          key=lambda s: s.get("saved_at", ""))
            if snaps:
                snap_prev = snaps[-1].get("metrics", {})
        except Exception as e:
            logger.warning(f"period_delta: failed to load prev snapshot: {e}")
    delta = _evm_period_delta(snap_now, snap_prev)
    return {"status": "ok", "current": snap_now, **delta}


def _msp_evm_variance_to_baseline(file_path=None, baseline_number=0):
    """Action 8: variance_to_baseline.

    Wraps compute_metrics with explicit baseline_number argument.
    Phase 3a integration — supports baselines 0-10.
    """
    if baseline_number not in BASELINE_NUMBERS:
        return {"status": "error",
                "error": f"baseline_number must be 0-10, got {baseline_number}"}
    return _msp_evm_compute_metrics(file_path=file_path,
                                    baseline_number=baseline_number)


def _msp_evm_compare_baselines_evm(file_path=None,
                                   baseline_a=0, baseline_b=1):
    """Action 9: compare_baselines_evm — B_a vs B_b EVM delta.

    Phase 3a compare_two pattern: shows EVM impact of revision plan.
    Returns {status, baseline_a, baseline_b, delta} where delta has
    bac_delta, spi_delta, cpi_delta keys.
    """
    a = _msp_evm_compute_metrics(file_path=file_path,
                                baseline_number=baseline_a)
    if a.get("status") != "ok":
        return {"status": "error",
                "error": f"baseline_a {baseline_a}: {a.get('error', 'load failed')}"}
    b = _msp_evm_compute_metrics(file_path=file_path,
                                baseline_number=baseline_b)
    if b.get("status") != "ok":
        return {"status": "error",
                "error": f"baseline_b {baseline_b}: {b.get('error', 'load failed')}"}
    delta = {
        "bac_delta": b["bac"] - a["bac"],
        "spi_delta": (b.get("spi") or 0) - (a.get("spi") or 0),
        "cpi_delta": (b.get("cpi") or 0) - (a.get("cpi") or 0),
    }
    return {"status": "ok",
            "baseline_a": a, "baseline_b": b, "delta": delta}


def _evm_snapshot_save(snapshot_path, snapshot):
    """Append snapshot dict to JSON file. Creates file with empty array if missing.

    Schema: {"snapshots": [<snapshot>, ...]} where snapshot has saved_at,
    metrics, forecast, earned_schedule, rag, tag, etc.
    """
    import json as _json
    if os.path.exists(snapshot_path):
        with open(snapshot_path, "r", encoding="utf-8") as f:
            data = _json.load(f)
    else:
        data = {"snapshots": []}
    data.setdefault("snapshots", []).append(snapshot)
    parent = os.path.dirname(snapshot_path) or "."
    if parent and parent != "." and not os.path.exists(parent):
        os.makedirs(parent, exist_ok=True)
    with open(snapshot_path, "w", encoding="utf-8") as f:
        _json.dump(data, f, indent=2, default=str)


def _evm_snapshot_load(snapshot_path, project_filter=None, baseline_filter=None):
    """Load snapshots from JSON file with optional filters.

    project_filter: substring match against project_name or project_file
    baseline_filter: exact baseline_number match (int)
    """
    import json as _json
    if not os.path.exists(snapshot_path):
        return []
    with open(snapshot_path, "r", encoding="utf-8") as f:
        data = _json.load(f)
    snaps = data.get("snapshots", []) or []
    if project_filter:
        snaps = [s for s in snaps
                 if project_filter in (s.get("project_name") or
                                       s.get("project_file") or "")]
    if baseline_filter is not None:
        snaps = [s for s in snaps if s.get("baseline_number") == baseline_filter]
    return snaps


def _msp_evm_save_period_snapshot(file_path=None, baseline_number=0,
                                  snapshot_path=None, tag=None):
    """Action 10: save_period_snapshot — append to JSON file.

    Bundles compute_metrics + forecast + earned_schedule + rag into
    a single snapshot entry. Default snapshot_path is
    ~/msproject_evm_snapshots.json.

    Phase 5a TAIL fix: optimized to share a single `compute_metrics`
    result across forecast + summary (was: 3 redundant COM iterations
    of ~200 tasks each). earned_schedule still requires its own load
    (PV curve build) — kept as separate call.
    """
    if not snapshot_path:
        snapshot_path = os.path.expanduser("~/msproject_evm_snapshots.json")
    cm = _msp_evm_compute_metrics(file_path=file_path,
                                  baseline_number=baseline_number)
    if cm.get("status") != "ok":
        return cm
    # Forecast = pure math from cm; no second COM iteration
    fc = _evm_forecast(bac=cm["bac"], ev=cm["ev"], ac=cm["ac"],
                       cpi=cm.get("cpi"), spi=cm.get("spi"))
    # Summary = RAG from cm; no second COM iteration
    completion_pct = (cm["ev"] / cm["bac"] * 100.0) if cm.get("bac", 0) > 0 else 0.0
    rag = _evm_rag(spi=cm.get("spi"), completion_pct=completion_pct)
    # ES still needs its own load + PV curve build — Phase 6 polish would
    # share via a unified `_evm_compute_full_snapshot` helper.
    es = _msp_evm_earned_schedule(file_path=file_path,
                                  baseline_number=baseline_number)
    snap = {
        "id": _dt5.datetime.now().strftime("%Y%m%d-%H%M%S"),
        "saved_at": _dt5.datetime.now().isoformat(),
        "project_file": file_path,
        "baseline_number": baseline_number,
        "metrics": {k: cm.get(k) for k in
                    ("bac", "pv", "ev", "ac", "spi", "cpi", "sv", "cv")},
        "forecast": {k: fc.get(k) for k in
                     ("eac_t1", "eac_t2", "eac_t3", "etc", "vac",
                      "tcpi_bac", "tcpi_eac")},
        "earned_schedule": {k: es.get(k) for k in
                            ("at", "es", "sv_t", "spi_t")}
                           if es.get("status") == "ok" else {},
        "rag": rag,
        "completion_pct": round(completion_pct, 2),
        "tag": tag,
    }
    try:
        _evm_snapshot_save(snapshot_path, snap)
        return {"status": "ok", "snapshot_path": snapshot_path,
                "snapshot_id": snap["id"]}
    except Exception as e:
        logger.exception(f"save_period_snapshot failed: {e}")
        return {"status": "error", "error": str(e)}


def _msp_evm_get_period_history(snapshot_path=None,
                                project_filter=None,
                                baseline_filter=None):
    """Action 11: get_period_history — list saved snapshots."""
    if not snapshot_path:
        snapshot_path = os.path.expanduser("~/msproject_evm_snapshots.json")
    snaps = _evm_snapshot_load(snapshot_path,
                              project_filter=project_filter,
                              baseline_filter=baseline_filter)
    return {"status": "ok", "count": len(snaps), "snapshots": snaps}


def _msp_evm_trend(snapshot_path=None, project_filter=None):
    """Action 12: trend — period-over-period series for SPI/CPI/EAC."""
    if not snapshot_path:
        snapshot_path = os.path.expanduser("~/msproject_evm_snapshots.json")
    snaps = _evm_snapshot_load(snapshot_path, project_filter=project_filter)
    snaps_sorted = sorted(snaps, key=lambda s: s.get("saved_at", ""))
    series = []
    for s in snaps_sorted:
        m = s.get("metrics", {}) or {}
        f = s.get("forecast", {}) or {}
        series.append({
            "saved_at": s.get("saved_at"),
            "tag": s.get("tag"),
            "spi": m.get("spi"),
            "cpi": m.get("cpi"),
            "eac_t3": f.get("eac_t3"),
            "rag": s.get("rag"),
        })
    return {"status": "ok", "count": len(series), "series": series}


@mcp.tool(
    name="msproject_evm",
    annotations={"title": "MS Project EVM Operations", "readOnlyHint": True},
)
async def msproject_evm(params: dict) -> str:
    """Earned Value Management — PMI PMBOK 8th § 7.4.2 + Lipke 2003 ES.

    Hybrid: file_path verilirse Phase 4 file path; yoksa Phase 1 COM.

    Actions:
    - compute_metrics: SPI/CPI/SV/CV (RULE 4)
    - forecast: EAC1/2/3 + ETC + VAC + TCPI(BAC/EAC) (RULE 9)
    - earned_schedule: AT, ES, SV(t), SPI(t) (RULE 8 Lipke)
    - summary: RAG + completion_pct + executive (RULE 12)
    - time_phased_evm: PV/EV/AC per period (bucket day/week/month, RULE 5)
    - period_delta: vs prev snapshot (RULE 6)
    - progress_data_quality: warnings (RULE 7)
    - variance_to_baseline: vs Baseline N (Phase 3a integration)
    - compare_baselines_evm: B_a vs B_b EVM delta
    - save_period_snapshot: append to JSON snapshot file
    - get_period_history: list saved snapshots (filter by project/baseline)
    - trend: SPI/CPI/EAC trajectory series
    - detect_currency_mode: hours vs cost (RULE 3)

    Phase 5a (30 Apr 2026). Tool count 8 -> 9.
    """
    import json
    action = params.get("action", "")
    p = {k: v for k, v in params.items() if k != "action"}
    try:
        if action == "compute_metrics":
            r = _msp_evm_compute_metrics(**p)
        elif action == "forecast":
            r = _msp_evm_forecast(**p)
        elif action == "earned_schedule":
            r = _msp_evm_earned_schedule(**p)
        elif action == "summary":
            r = _msp_evm_summary(**p)
        elif action == "time_phased_evm":
            r = _msp_evm_time_phased_evm(**p)
        elif action == "period_delta":
            r = _msp_evm_period_delta(**p)
        elif action == "progress_data_quality":
            r = _msp_evm_progress_data_quality(**p)
        elif action == "variance_to_baseline":
            r = _msp_evm_variance_to_baseline(**p)
        elif action == "compare_baselines_evm":
            r = _msp_evm_compare_baselines_evm(**p)
        elif action == "save_period_snapshot":
            r = _msp_evm_save_period_snapshot(**p)
        elif action == "get_period_history":
            r = _msp_evm_get_period_history(**p)
        elif action == "trend":
            r = _msp_evm_trend(**p)
        elif action == "detect_currency_mode":
            r = _msp_evm_detect_currency_mode(**p)
        elif action == "validate_currency_mode":
            r = _msp_evm_validate_currency_mode(**p)
        else:
            r = {"status": "error",
                 "error": (f"Unknown action '{action}'. Valid: "
                          "compute_metrics/forecast/earned_schedule/summary/"
                          "time_phased_evm/period_delta/progress_data_quality/"
                          "variance_to_baseline/compare_baselines_evm/"
                          "save_period_snapshot/get_period_history/trend/"
                          "detect_currency_mode/validate_currency_mode")}
    except TypeError as e:
        r = {"status": "error", "error": f"Invalid params for {action}: {e}"}
    except Exception as e:
        logger.exception(f"msproject_evm({action}) failed: {e}")
        r = {"status": "error", "error": str(e)}
    return json.dumps(r, default=str, ensure_ascii=False)


# ============================================================================
# PHASE 5B - DCMA TOOL
# ============================================================================
from dcma_checks import (
    DCMA_RULES,
    check_no_predecessor as _dcma_check_1,
    check_no_successor as _dcma_check_2,
    check_leads as _dcma_check_3,
    check_lags as _dcma_check_4,
    check_fs_link_pct as _dcma_check_5,
    check_hard_constraints as _dcma_check_6,
    check_high_float as _dcma_check_7,
    check_negative_float as _dcma_check_8,
    check_high_duration as _dcma_check_9,
    check_invalid_dates as _dcma_check_10,
    check_resources_missing as _dcma_check_11,
    check_missed_tasks as _dcma_check_12,
    check_critical_path as _dcma_check_13,
    check_bei as _dcma_check_14,
    assess_all as _dcma_assess_all,
    compute_overall_rag as _dcma_overall_rag,
)


def _dcma_load_links(file_path=None):
    """Hybrid: file_path -> Phase 4 _msp_file_read_links;
    None -> Phase 1 COM iter walking proj.Tasks predecessors.

    Returns list of {from_id, to_id, type, lag_days}.

    Phase 5e additive routing: file_path ending '.xer' delegates to
    XerFile.read_links (Phase 5d reader). Existing .xml/.mpp paths
    unchanged.
    """
    if file_path and isinstance(file_path, str) and file_path.lower().endswith(".xer"):
        from xer_parser import XerFile
        try:
            return XerFile(file_path).read_links()
        except Exception as e:
            logger.exception(f"_dcma_load_links XER routing failed: {e}")
            return []
    if file_path:
        r = _msp_file_read_links(file_path=file_path)
        if r.get("status") != "ok":
            return []
        return r.get("links", []) or []
    # COM path
    try:
        app = _validate_active_project()
        proj = app.ActiveProject
        out = []
        for i in range(1, proj.Tasks.Count + 1):
            try:
                t = proj.Tasks(i)
                if t is None:
                    continue
                # Use TaskDependencies for richer info (type + lag)
                try:
                    deps = t.TaskDependencies
                except Exception:
                    deps = None
                if deps:
                    for j in range(1, deps.Count + 1):
                        try:
                            d = deps(j)
                            if d is None:
                                continue
                            ft = d.From
                            tt = d.To
                            if ft is None or tt is None or tt.ID != t.ID:
                                continue
                            type_code = int(d.Type or 0)
                            type_str = ["FF", "FS", "SF", "SS"][type_code] if 0 <= type_code <= 3 else "FS"
                            lag_min = float(d.Lag or 0)
                            lag_days = lag_min / 480.0  # 8h/day default
                            out.append({
                                "from_id": ft.ID, "to_id": tt.ID,
                                "type": type_str, "lag_days": round(lag_days, 2),
                            })
                        except Exception:
                            continue
            except Exception:
                continue
        return out
    except Exception as e:
        logger.exception(f"_dcma_load_links COM path failed: {e}")
        return []


def _dcma_collect_full_data(file_path=None, baseline_number=0):
    """Aggregate Phase 5a data + Phase 5b extensions.

    Returns {status, tasks, links, assignments, resources, baseline,
             status_date}. tasks already include total_slack_days,
             critical, constraint_type fields when available (Phase 4
             file path). COM path adds these via task property reads.
    """
    if baseline_number not in BASELINE_NUMBERS:
        return {"status": "error",
                "error": f"baseline_number must be 0-10, got {baseline_number}"}
    base = _evm_load_task_data(file_path=file_path)
    if base.get("status") != "ok":
        return base
    bload = _evm_load_baseline_data(file_path=file_path,
                                    baseline_number=baseline_number)
    links = _dcma_load_links(file_path=file_path)
    tasks = base.get("tasks", []) or []
    # Phase 5a COM path returns assignments=[] for perf — re-collect here.
    assignments = base.get("assignments", []) or []
    # Enrich tasks with DCMA-specific fields when COM path
    if not file_path:
        try:
            app = _validate_active_project()
            proj = app.ActiveProject
            # Single proj.Tasks iteration: build O(1) map (avoid O(N^2) via
            # _find_task_by_id per task) AND collect assignments inline
            # (Phase 5a COM skips assignments for perf).
            tasks_by_com_id = {}
            com_assignments = []
            for i in range(1, proj.Tasks.Count + 1):
                try:
                    com_t = proj.Tasks(i)
                    if com_t is None:
                        continue
                    tid_int = int(com_t.ID)
                    tasks_by_com_id[tid_int] = com_t
                    # Walk task assignments
                    try:
                        a_coll = com_t.Assignments
                        if a_coll:
                            for ai in range(1, a_coll.Count + 1):
                                try:
                                    a = a_coll(ai)
                                    if a is None:
                                        continue
                                    com_assignments.append({
                                        "task_id": tid_int,
                                        "resource_id": int(a.ResourceID),
                                    })
                                except Exception:
                                    continue
                    except Exception:
                        pass
                except Exception:
                    continue
            assignments = com_assignments
            for t_dict in tasks:
                tid = int(t_dict["id"])
                com_t = tasks_by_com_id.get(tid)
                if com_t is None:
                    continue
                try:
                    slack_min = float(com_t.TotalSlack or 0)
                    t_dict["total_slack_days"] = round(slack_min / 480.0, 2)
                except Exception:
                    t_dict["total_slack_days"] = 0
                try:
                    t_dict["critical"] = bool(com_t.Critical)
                except Exception:
                    t_dict["critical"] = False
                try:
                    t_dict["constraint_type"] = int(com_t.ConstraintType or 0)
                except Exception:
                    t_dict["constraint_type"] = 0
                # Predecessors/successors as ID lists via TaskDependencies
                try:
                    deps = com_t.TaskDependencies
                    preds = []
                    succs = []
                    if deps:
                        for j in range(1, deps.Count + 1):
                            d = deps(j)
                            if d is None:
                                continue
                            if d.To and d.To.ID == tid and d.From:
                                preds.append(d.From.ID)
                            elif d.From and d.From.ID == tid and d.To:
                                succs.append(d.To.ID)
                    t_dict["predecessors"] = preds
                    t_dict["successors"] = succs
                except Exception:
                    t_dict.setdefault("predecessors", [])
                    t_dict.setdefault("successors", [])
        except Exception as e:
            logger.warning(f"_dcma_collect_full_data COM enrich failed: {e}")
    # File path: Phase 4 already provides total_float, critical, constraint_type
    else:
        for t_dict in tasks:
            t_dict.setdefault("total_slack_days", float(t_dict.get("total_float") or 0))
            t_dict.setdefault("critical", t_dict.get("critical", False))
            t_dict.setdefault("constraint_type", t_dict.get("constraint_type", 0))
    return {
        "status": "ok",
        "tasks": tasks,
        "links": links,
        "assignments": assignments,
        "resources": base.get("resources", []) or [],
        "baseline": bload if bload.get("status") == "ok" else None,
        "status_date": base.get("status_date"),
    }


def _msp_dcma_assess_all(file_path=None, baseline_number=0):
    """Action 1: assess_all - full DCMA 14-Point assessment."""
    data = _dcma_collect_full_data(file_path=file_path,
                                   baseline_number=baseline_number)
    if data.get("status") != "ok":
        return data
    result = _dcma_assess_all(
        tasks=data["tasks"],
        links=data["links"],
        assignments=data["assignments"],
        baseline=data.get("baseline"),
        status_date=data.get("status_date"),
    )
    return {"status": "ok", "baseline_number": baseline_number, **result}


def _msp_dcma_summary(file_path=None, baseline_number=0):
    """Action 2: summary - RAG + executive text only."""
    full = _msp_dcma_assess_all(file_path=file_path,
                                baseline_number=baseline_number)
    if full.get("status") != "ok":
        return full
    return {"status": "ok",
            "baseline_number": baseline_number,
            **full["summary"]}


def _msp_dcma_drill_down(file_path=None, rule_id=1, baseline_number=0):
    """Action 3: drill_down - per-rule failed task details.

    Single _dcma_collect_full_data call (was 2x — Phase 5b TAIL fix to
    avoid double COM enrichment on large projects).
    """
    if rule_id not in range(1, 15):
        return {"status": "error",
                "error": f"rule_id must be 1-14, got {rule_id}"}
    data = _dcma_collect_full_data(file_path=file_path,
                                   baseline_number=baseline_number)
    if data.get("status") != "ok":
        return data
    result = _dcma_assess_all(
        tasks=data["tasks"],
        links=data["links"],
        assignments=data["assignments"],
        baseline=data.get("baseline"),
        status_date=data.get("status_date"),
    )
    rule = next((r for r in result["rules"] if r["id"] == rule_id), None)
    if rule is None:
        return {"status": "error", "error": f"Rule {rule_id} not found"}
    tasks_by_id = {t["id"]: t for t in data.get("tasks", [])}
    failed_ids = rule.get("failed_task_ids", [])
    failed_tasks = []
    for tid in failed_ids:
        t = tasks_by_id.get(tid)
        if t:
            failed_tasks.append({"id": tid, "name": t.get("name", "")})
    return {
        "status": "ok",
        "baseline_number": baseline_number,
        "rule": {"id": rule["id"], "name": rule["name"],
                 "threshold": rule.get("threshold")},
        "actual": rule.get("actual"),
        "failed_count": rule.get("failed_count"),
        "total_count": rule.get("total_count"),
        "failed_tasks": failed_tasks,
    }


def _msp_dcma_compare(file_path=None, snapshot_path=None, baseline_number=0):
    """Action 4: compare current DCMA vs prev snapshot.

    Reuses Phase 5a _evm_snapshot_load to read prior DCMA dumps from
    the same JSON file (snapshots can include both EVM + DCMA data).
    """
    current = _msp_dcma_assess_all(file_path=file_path,
                                   baseline_number=baseline_number)
    if current.get("status") != "ok":
        return current
    if not snapshot_path:
        return {"status": "ok", "current": current["summary"], "prev": None,
                "delta": {"rules_improved": [], "rules_degraded": []}}
    snaps = _evm_snapshot_load(snapshot_path) if os.path.exists(snapshot_path) else []
    # Filter snaps that have DCMA data
    dcma_snaps = [s for s in snaps if s.get("dcma")]
    if not dcma_snaps:
        return {"status": "ok", "current": current["summary"], "prev": None,
                "delta": {"rules_improved": [], "rules_degraded": []}}
    dcma_snaps.sort(key=lambda s: s.get("saved_at", ""))
    prev = dcma_snaps[-1].get("dcma")
    # Compute delta
    improved = []
    degraded = []
    prev_rules = {r["id"]: r for r in (prev.get("rules") or [])}
    for cr in current["rules"]:
        pr = prev_rules.get(cr["id"])
        if pr is None:
            continue
        cur_actual = cr.get("actual", 0)
        prev_actual = pr.get("actual", 0)
        if pr.get("status") == "fail" and cr.get("status") == "pass":
            improved.append({"id": cr["id"], "name": cr["name"],
                             "from_actual": prev_actual, "to_actual": cur_actual})
        elif pr.get("status") == "pass" and cr.get("status") == "fail":
            degraded.append({"id": cr["id"], "name": cr["name"],
                             "from_actual": prev_actual, "to_actual": cur_actual})
    return {
        "status": "ok",
        "current": current["summary"],
        "prev": prev.get("summary") if isinstance(prev, dict) else None,
        "delta": {"rules_improved": improved, "rules_degraded": degraded},
    }


@mcp.tool(
    name="msproject_health",
    annotations={
        "title": "MS Project DCMA 14-Point Health Assessment",
        "readOnlyHint": True,
    },
)
async def msproject_health(params: dict) -> str:
    """DCMA 14-Point Schedule Health Assessment per CLAUDE.md RULE 10.

    Hybrid: file_path verilirse Phase 4 file path; yoksa Phase 1 COM.
    Read-only - no write actions.

    Actions:
    - assess_all: All 14 rules + summary + RAG
    - summary: Just RAG + executive text
    - drill_down: Per-rule failed task list (rule_id 1-14)
    - compare: Current DCMA vs prev snapshot (reuses Phase 5a snapshot file)

    Phase 5b (1 May 2026). Tool count 9 -> 10.
    """
    import json
    action = params.get("action", "")
    p = {k: v for k, v in params.items() if k != "action"}
    try:
        if action == "assess_all":
            r = _msp_dcma_assess_all(**p)
        elif action == "summary":
            r = _msp_dcma_summary(**p)
        elif action == "drill_down":
            r = _msp_dcma_drill_down(**p)
        elif action == "compare":
            r = _msp_dcma_compare(**p)
        else:
            r = {"status": "error",
                 "error": (f"Unknown action '{action}'. Valid: "
                           "assess_all/summary/drill_down/compare")}
    except TypeError as e:
        r = {"status": "error", "error": f"Invalid params for {action}: {e}"}
    except Exception as e:
        logger.exception(f"msproject_health({action}) failed: {e}")
        r = {"status": "error", "error": str(e)}
    return json.dumps(r, default=str, ensure_ascii=False)


# ============================================================================
# PHASE 5C - EXCEL TOOL
# ============================================================================
from excel_io import (
    build_tasks_sheet, build_evm_sheet, build_dcma_sheet,
    build_summary_sheet, build_hakedis_workbook,
    read_tasks_sheet, read_progress_sheet,
)
from openpyxl import Workbook as _XlWorkbook


def _excel_collect_full_data(file_path=None, baseline_number=0, bucket="week"):
    """Single-collect aggregator (Phase 5b TAIL lesson) - fetch tasks + EVM
    + DCMA once and translate Phase 5a/5b flat-lowercase shape to excel_io's
    nested-uppercase shape.

    Returns {status, tasks, evm: {metrics, forecast, earned_schedule, rag,
             time_phased}, dcma: {rules, summary, drilldowns}}.
    """
    if baseline_number not in BASELINE_NUMBERS:
        return {"status": "error",
                "error": f"baseline_number must be 0-10, got {baseline_number}"}
    base = _evm_load_task_data(file_path=file_path)
    if base.get("status") != "ok":
        return base
    tasks = base.get("tasks", []) or []

    # Phase 5a EVM compute (4 calls) - all return flat lowercase keys
    cm = _msp_evm_compute_metrics(file_path=file_path, baseline_number=baseline_number)
    fc = _msp_evm_forecast(file_path=file_path, baseline_number=baseline_number)
    es = _msp_evm_earned_schedule(file_path=file_path, baseline_number=baseline_number)
    sm = _msp_evm_summary(file_path=file_path, baseline_number=baseline_number)
    tp = _msp_evm_time_phased_evm(file_path=file_path,
                                  baseline_number=baseline_number, bucket=bucket)
    if tp.get("status") != "ok":
        return tp  # bubble up bucket validation errors

    # Translate flat lowercase -> nested UPPERCASE for excel_io
    metrics = {}
    if cm.get("status") == "ok":
        metrics = {"BAC": cm.get("bac"), "EV": cm.get("ev"), "AC": cm.get("ac"),
                   "PV": cm.get("pv"), "SV": cm.get("sv"), "CV": cm.get("cv"),
                   "SPI": cm.get("spi"), "CPI": cm.get("cpi")}
    forecast = {}
    if fc.get("status") == "ok":
        forecast = {"EAC1": fc.get("eac1"), "EAC2": fc.get("eac2"),
                    "EAC3": fc.get("eac3"), "ETC": fc.get("etc"),
                    "VAC": fc.get("vac"),
                    "TCPI_BAC": fc.get("tcpi_bac"), "TCPI_EAC": fc.get("tcpi_eac")}
    earned_schedule = {}
    if es.get("status") == "ok":
        earned_schedule = {"AT": es.get("at"), "ES": es.get("es"),
                           "SVt": es.get("sv_t"), "SPIt": es.get("spi_t")}

    # Build time_phased with cumulative columns
    time_phased = []
    cum_pv = cum_ev = cum_ac = 0.0
    for b in tp.get("buckets", []):
        pv = float(b.get("pv") or 0)
        ev = float(b.get("ev") or 0)
        ac = float(b.get("ac") or 0)
        cum_pv += pv
        cum_ev += ev
        cum_ac += ac
        time_phased.append({
            "period": str(b.get("period_start")) if b.get("period_start") else "",
            "PV": pv, "EV": ev, "AC": ac,
            "cum_PV": cum_pv, "cum_EV": cum_ev, "cum_AC": cum_ac,
        })

    # Phase 5b DCMA
    dcma_full = _msp_dcma_assess_all(file_path=file_path,
                                     baseline_number=baseline_number)
    drilldowns = {}
    dcma_rules = []
    dcma_summary = {}
    if dcma_full.get("status") == "ok":
        dcma_rules = dcma_full.get("rules", []) or []
        dcma_summary = dcma_full.get("summary", {}) or {}
        # Phase 5c TAIL fix: resolve failed_task_ids -> names from LOCAL
        # tasks list (already loaded above) instead of calling
        # _msp_dcma_drill_down per failed rule (each call = 1 collect = ~10s
        # for 200 tasks COM; 7 failed rules = 70s saved).
        tasks_by_id = {t["id"]: t for t in tasks}
        for rule in dcma_rules:
            if rule.get("status") == "fail":
                rid = rule["id"]
                failed_ids = rule.get("failed_task_ids") or []
                drilldowns[rid] = [
                    {"id": tid, "name": tasks_by_id[tid].get("name", "")}
                    for tid in failed_ids[:10]
                    if tid in tasks_by_id
                ]

    # RAG: prefer Phase 5b DCMA overall_rag (project health) over Phase 5a EVM
    rag = dcma_summary.get("overall_rag") or (sm.get("rag") if sm.get("status") == "ok" else None)

    return {
        "status": "ok",
        "tasks": tasks,
        "evm": {
            "metrics": metrics,
            "forecast": forecast,
            "earned_schedule": earned_schedule,
            "rag": rag,
            "time_phased": time_phased,
        },
        "dcma": {
            "rules": dcma_rules,
            "summary": dcma_summary,
            "drilldowns": drilldowns,
        },
    }


def _msp_excel_export_hakedis(file_path=None, xlsx_path=None, baseline_number=0):
    """Action 1 (HERO): export multi-sheet hakedis workbook."""
    if not xlsx_path:
        return {"status": "error", "error": "xlsx_path required"}
    data = _excel_collect_full_data(file_path=file_path,
                                    baseline_number=baseline_number)
    if data.get("status") != "ok":
        return data
    summary_for_sheet = {
        "BAC": data["evm"]["metrics"].get("BAC"),
        "EAC": data["evm"]["forecast"].get("EAC2"),
        "SPI": data["evm"]["metrics"].get("SPI"),
        "CPI": data["evm"]["metrics"].get("CPI"),
        "rag": data["dcma"]["summary"].get("overall_rag") or data["evm"].get("rag"),
        "executive_text": data["dcma"]["summary"].get("executive_text", ""),
    }
    try:
        build_hakedis_workbook(
            tasks=data["tasks"], evm=data["evm"], dcma=data["dcma"],
            summary=summary_for_sheet, xlsx_path=xlsx_path,
        )
    except Exception as e:
        logger.exception(f"export_hakedis failed: {e}")
        return {"status": "error", "error": str(e)}
    real_count = len([t for t in data["tasks"] if not t.get("summary")])
    return {
        "status": "ok",
        "xlsx_path": xlsx_path,
        "sheets_written": ["Summary", "Tasks", "EVM_Compute", "EVM_TimePhased",
                           "DCMA_Rules", "DCMA_Failed"],
        "rows_written": {
            "tasks": real_count,
            "evm_time_phased": len(data["evm"].get("time_phased", [])),
            "dcma_rules": len(data["dcma"].get("rules", [])),
        },
    }


def _msp_excel_export_tasks(file_path=None, xlsx_path=None):
    """Action 2: export tasks-only sheet."""
    if not xlsx_path:
        return {"status": "error", "error": "xlsx_path required"}
    base = _evm_load_task_data(file_path=file_path)
    if base.get("status") != "ok":
        return base
    try:
        wb = _XlWorkbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        build_tasks_sheet(wb, base.get("tasks", []) or [], sheet_name="Tasks")
        wb.save(xlsx_path)
    except Exception as e:
        logger.exception(f"export_tasks failed: {e}")
        return {"status": "error", "error": str(e)}
    real = len([t for t in base.get("tasks", []) if not t.get("summary")])
    return {"status": "ok", "xlsx_path": xlsx_path, "rows_written": real}


def _msp_excel_export_evm(file_path=None, xlsx_path=None, baseline_number=0,
                          bucket="week"):
    """Action 3: export EVM (Compute + TimePhased) sheets."""
    if not xlsx_path:
        return {"status": "error", "error": "xlsx_path required"}
    data = _excel_collect_full_data(file_path=file_path,
                                    baseline_number=baseline_number,
                                    bucket=bucket)
    if data.get("status") != "ok":
        return data
    try:
        wb = _XlWorkbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        build_evm_sheet(wb, data["evm"])
        wb.save(xlsx_path)
    except Exception as e:
        logger.exception(f"export_evm failed: {e}")
        return {"status": "error", "error": str(e)}
    return {"status": "ok", "xlsx_path": xlsx_path,
            "rows_written": {"compute": 19,
                             "time_phased": len(data["evm"].get("time_phased", []))}}


def _msp_excel_export_dcma(file_path=None, xlsx_path=None, baseline_number=0):
    """Action 4: export DCMA (Rules + Failed) sheets."""
    if not xlsx_path:
        return {"status": "error", "error": "xlsx_path required"}
    data = _excel_collect_full_data(file_path=file_path,
                                    baseline_number=baseline_number)
    if data.get("status") != "ok":
        return data
    try:
        wb = _XlWorkbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        build_dcma_sheet(wb, data["dcma"])
        wb.save(xlsx_path)
    except Exception as e:
        logger.exception(f"export_dcma failed: {e}")
        return {"status": "error", "error": str(e)}
    return {"status": "ok", "xlsx_path": xlsx_path,
            "rows_written": {"rules": len(data["dcma"].get("rules", [])),
                             "drilldowns": sum(len(v) for v in data["dcma"].get("drilldowns", {}).values())}}


def _msp_excel_import_tasks(xlsx_path=None, sheet_name="Tasks"):
    """Action 5: import tasks from xlsx via Phase 1 _msp_task_bulk_add.

    Reads xlsx Tasks sheet, converts each row to bulk_add items shape
    ({name, duration:'Nd'}), then delegates. Skips rows missing 'name'.
    """
    if not xlsx_path:
        return {"status": "error", "error": "xlsx_path required"}
    if not os.path.exists(xlsx_path):
        return {"status": "error", "error": f"File not found: {xlsx_path}"}
    try:
        rows = read_tasks_sheet(xlsx_path, sheet_name=sheet_name)
    except Exception as e:
        logger.exception(f"import_tasks read failed: {e}")
        return {"status": "error", "error": f"Read failed: {e}"}
    if not rows:
        return {"status": "ok", "rows_imported": 0, "task_ids": []}
    items = []
    for r in rows:
        if not r.get("name"):
            continue
        days = round(float(r.get("duration_h") or 0) / 8.0, 1)
        if days <= 0:
            days = 1.0
        items.append({"name": r["name"], "duration": f"{days}d"})
    if not items:
        return {"status": "ok", "rows_imported": 0, "task_ids": []}
    try:
        result = _msp_task_bulk_add(items=items)
    except Exception as e:
        logger.exception(f"_msp_task_bulk_add failed: {e}")
        return {"status": "error", "error": str(e)}
    if isinstance(result, dict) and result.get("status") == "error":
        return result
    return {
        "status": "ok",
        "rows_imported": len(items),
        "task_ids": result.get("task_ids", []) if isinstance(result, dict) else [],
    }


def _msp_excel_import_progress(xlsx_path=None, sheet_name="Progress"):
    """Action 6: import progress updates from xlsx via Phase 3b bulk_update."""
    if not xlsx_path:
        return {"status": "error", "error": "xlsx_path required"}
    if not os.path.exists(xlsx_path):
        return {"status": "error", "error": f"File not found: {xlsx_path}"}
    try:
        rows = read_progress_sheet(xlsx_path, sheet_name=sheet_name)
    except Exception as e:
        logger.exception(f"import_progress read failed: {e}")
        return {"status": "error", "error": f"Read failed: {e}"}
    if not rows:
        return {"status": "ok", "rows_imported": 0}
    items = [{"task_id": r["task_id"],
              "percent_complete": r.get("percent_complete", 0)}
             for r in rows if r.get("task_id") is not None]
    if not items:
        return {"status": "ok", "rows_imported": 0}
    try:
        result = _msp_progress_bulk_update(items=items)
    except Exception as e:
        logger.exception(f"_msp_progress_bulk_update failed: {e}")
        return {"status": "error", "error": str(e)}
    if isinstance(result, dict) and result.get("status") == "error":
        return result
    return {"status": "ok", "rows_imported": len(items)}


@mcp.tool(
    name="msproject_excel",
    annotations={
        "title": "MS Project Excel Hakedis Workbook + Bulk Import",
        "readOnlyHint": False,
    },
)
async def msproject_excel(params: dict) -> str:
    """Excel I/O for MSP - hakedis workbook export + bulk Excel->MSP import.

    Hybrid: file_path verilirse Phase 4 file path; yoksa Phase 1 COM.

    Actions:
    - export_hakedis: Multi-sheet workbook (Summary + Tasks + EVM + DCMA)
    - export_tasks: Tasks sheet only
    - export_evm: EVM_Compute + EVM_TimePhased
    - export_dcma: DCMA_Rules + DCMA_Failed
    - import_tasks: xlsx Tasks sheet -> _msp_task_bulk_add
    - import_progress: xlsx Progress sheet -> _msp_progress_bulk_update

    Phase 5c (1 May 2026). Tool count 10 -> 11.
    """
    import json
    action = params.get("action", "")
    p = {k: v for k, v in params.items() if k != "action"}
    try:
        if action == "export_hakedis":
            r = _msp_excel_export_hakedis(**p)
        elif action == "export_tasks":
            r = _msp_excel_export_tasks(**p)
        elif action == "export_evm":
            r = _msp_excel_export_evm(**p)
        elif action == "export_dcma":
            r = _msp_excel_export_dcma(**p)
        elif action == "import_tasks":
            r = _msp_excel_import_tasks(**p)
        elif action == "import_progress":
            r = _msp_excel_import_progress(**p)
        else:
            r = {"status": "error",
                 "error": (f"Unknown action '{action}'. Valid: "
                           "export_hakedis/export_tasks/export_evm/"
                           "export_dcma/import_tasks/import_progress")}
    except TypeError as e:
        r = {"status": "error", "error": f"Invalid params for {action}: {e}"}
    except Exception as e:
        logger.exception(f"msproject_excel({action}) failed: {e}")
        r = {"status": "error", "error": str(e)}
    return json.dumps(r, default=str, ensure_ascii=False)


# ============================================================================
# PHASE 5D - XER (PRIMAVERA P6) READER
# ============================================================================
from xer_parser import XerFile


def _xer_collect_full_data(file_path):
    """Single-collect aggregator (Phase 5b/5c TAIL lesson). Parses XER once,
    returns all 6 read shapes from a single XerFile instance.

    First calendar's day_hr_cnt drives total_float h->day conversion (CAU
    9.0 vs default 8.0). When no CALENDAR section, defaults to 8h/day.
    """
    if not file_path:
        return {"status": "error", "error": "file_path required"}
    try:
        xer = XerFile(file_path)
    except FileNotFoundError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        logger.exception(f"_xer_collect_full_data failed: {e}")
        return {"status": "error", "error": str(e)}
    cals = xer.read_calendars()
    day_hr_cnt = cals[0]["day_hr_cnt"] if cals else 8.0
    return {
        "status": "ok",
        "tasks": xer.read_tasks(day_hr_cnt=day_hr_cnt),
        "links": xer.read_links(),
        "resources": xer.read_resources(),
        "assignments": xer.read_assignments(),
        "calendars": cals,
        "progress": xer.read_progress(),
        "project": xer.read_project(),
    }


def _msp_xer_read_tasks(file_path=None, filters=None, limit=None):
    """Action 1: read tasks with optional filter dict + limit."""
    data = _xer_collect_full_data(file_path)
    if data.get("status") != "ok":
        return data
    tasks = data["tasks"]
    if filters:
        for k, v in filters.items():
            tasks = [t for t in tasks if t.get(k) == v]
    if limit:
        tasks = tasks[:int(limit)]
    return {"status": "ok", "count": len(tasks), "tasks": tasks}


def _msp_xer_read_links(file_path=None):
    """Action 2: read all links."""
    data = _xer_collect_full_data(file_path)
    if data.get("status") != "ok":
        return data
    return {"status": "ok", "count": len(data["links"]), "links": data["links"]}


def _msp_xer_read_resources(file_path=None):
    """Action 3: read all resources."""
    data = _xer_collect_full_data(file_path)
    if data.get("status") != "ok":
        return data
    return {"status": "ok", "count": len(data["resources"]),
            "resources": data["resources"]}


def _msp_xer_read_assignments(file_path=None):
    """Action 4: read all task-resource assignments."""
    data = _xer_collect_full_data(file_path)
    if data.get("status") != "ok":
        return data
    return {"status": "ok", "count": len(data["assignments"]),
            "assignments": data["assignments"]}


def _msp_xer_read_calendars(file_path=None):
    """Action 5: read all calendars."""
    data = _xer_collect_full_data(file_path)
    if data.get("status") != "ok":
        return data
    return {"status": "ok", "count": len(data["calendars"]),
            "calendars": data["calendars"]}


def _msp_xer_read_progress(file_path=None):
    """Action 6: read progress data (status_date + per-task progress)."""
    data = _xer_collect_full_data(file_path)
    if data.get("status") != "ok":
        return data
    return {"status": "ok", **data["progress"]}


@mcp.tool(
    name="msproject_xer",
    annotations={
        "title": "MS Project XER (Primavera P6) Reader",
        "readOnlyHint": True,
    },
)
async def msproject_xer(params: dict) -> str:
    """Pure-Python Primavera P6 XER file reader. Read-only.

    Bridges P6 XER projects (CAU baseline format) into Phase 5a EVM +
    Phase 5b DCMA + Phase 5c Excel pipelines. NO mpxj/Java dependency.

    Actions:
    - read_tasks(filters?, limit?): TASK section with optional dict filter
    - read_links: TASKPRED section
    - read_resources: RSRC section
    - read_assignments: TASKRSRC section
    - read_calendars: CALENDAR section (day_hr_cnt + week_hr_cnt)
    - read_progress: PROJECT.last_recalc_date + per-task progress

    Phase 5d (1 May 2026). Tool count 11 -> 12.
    """
    import json
    action = params.get("action", "")
    p = {k: v for k, v in params.items() if k != "action"}
    try:
        if action == "read_tasks":
            r = _msp_xer_read_tasks(**p)
        elif action == "read_links":
            r = _msp_xer_read_links(**p)
        elif action == "read_resources":
            r = _msp_xer_read_resources(**p)
        elif action == "read_assignments":
            r = _msp_xer_read_assignments(**p)
        elif action == "read_calendars":
            r = _msp_xer_read_calendars(**p)
        elif action == "read_progress":
            r = _msp_xer_read_progress(**p)
        else:
            r = {"status": "error",
                 "error": (f"Unknown action '{action}'. Valid: "
                           "read_tasks/read_links/read_resources/"
                           "read_assignments/read_calendars/read_progress")}
    except TypeError as e:
        r = {"status": "error", "error": f"Invalid params for {action}: {e}"}
    except Exception as e:
        logger.exception(f"msproject_xer({action}) failed: {e}")
        r = {"status": "error", "error": str(e)}
    return json.dumps(r, default=str, ensure_ascii=False)


# ============================================================================
# PHASE 5E - XER NATIVE INTEGRATION (Phase 5a loader extensions for .xer)
# ============================================================================
# Wires Phase 5d msproject_xer reader into Phase 5a EVM + Phase 5b DCMA +
# Phase 5c Excel pipelines via additive routing in _evm_load_task_data and
# _evm_load_baseline_data (single guard line each, NO behavior change for
# existing .xml/.mpp/COM paths). Adapters below translate XerFile output
# to Phase 5a expected shape.


def _xer_to_evm_task_shape(xer):
    """Translate XerFile output to Phase 5a _evm_load_task_data shape.

    CAU pattern (cost-loaded NO): baseline = target schedule. baseline_work
    = duration_h, baseline_start/finish = target dates.

    Derives: predecessors/successors lists from links, total_slack_days,
    critical (heuristic: total_slack_days <= 0 — XER lacks explicit
    critical flag).
    """
    cals = xer.read_calendars()
    day_hr_cnt = cals[0]["day_hr_cnt"] if cals else 8.0
    raw_tasks = xer.read_tasks(day_hr_cnt=day_hr_cnt)
    links = xer.read_links()
    progress = xer.read_progress()
    assignments = xer.read_assignments()

    # Pre-aggregate actual_work per task from XER assignments
    actual_by_task = {}
    for a in assignments:
        tid = a.get("task_id")
        if tid is not None:
            actual_by_task[tid] = actual_by_task.get(tid, 0.0) + float(
                a.get("actual_qty") or 0)

    # Pre-build predecessor/successor maps from links (single O(M) pass)
    preds_by_task = {}
    succs_by_task = {}
    for link in links:
        from_id = link.get("from_id")
        to_id = link.get("to_id")
        if to_id is not None and from_id is not None:
            preds_by_task.setdefault(to_id, []).append(from_id)
            succs_by_task.setdefault(from_id, []).append(to_id)

    # Build Phase 5a-shape task dicts (filter summaries)
    out_tasks = []
    for t in raw_tasks:
        if t.get("summary", False):
            continue
        tid = t["id"]
        slack = t.get("total_float", 0)
        out_tasks.append({
            **t,  # carry XER fields (id/name/code/duration_h/start/finish/...)
            "baseline_start": t.get("start"),
            "baseline_finish": t.get("finish"),
            "baseline_work": float(t.get("duration_h") or 0),
            "actual_work": actual_by_task.get(tid, 0.0),
            "total_slack_days": slack,
            "critical": float(slack) <= 0,
            "predecessors": preds_by_task.get(tid, []),
            "successors": succs_by_task.get(tid, []),
        })

    return {
        "status": "ok",
        "tasks": out_tasks,
        "resources": xer.read_resources(),
        "assignments": assignments,
        "status_date": progress.get("status_date"),
        "project_file": xer.file_path,
    }


def _xer_to_evm_baseline_shape(xer, baseline_number=0):
    """Translate XerFile to Phase 5a _evm_load_baseline_data shape.

    CAU pattern (cost-loaded NO): baseline = target schedule. Returns task
    baseline fields keyed by task_id. XER has only 1 implicit baseline per
    file (PROJECT.last_recalc_date snapshot); baseline_number ignored
    beyond passthrough.
    """
    cals = xer.read_calendars()
    day_hr_cnt = cals[0]["day_hr_cnt"] if cals else 8.0
    raw_tasks = xer.read_tasks(day_hr_cnt=day_hr_cnt)
    out = []
    for t in raw_tasks:
        if t.get("summary", False):
            continue
        out.append({
            "task_id": t["id"],
            "id": t["id"],
            "name": t.get("name", ""),
            "start": t.get("start"),
            "finish": t.get("finish"),
            "baseline_start": t.get("start"),
            "baseline_finish": t.get("finish"),
            "work_h": float(t.get("duration_h") or 0),
            "baseline_work": float(t.get("duration_h") or 0),
        })
    return {
        "status": "ok",
        "baseline_number": baseline_number,
        "tasks": out,
    }


# ============================================================================
# PHASE 7 - msproject_compare (XER vs XER / MSPDI vs MSPDI delta analysis)
# ============================================================================
# Bridges xer_compare pure module into a new tool. Reuses Phase 5a
# _evm_load_task_data + Phase 5b _dcma_load_links + _msp_evm_compute_metrics
# read-only — no DOKUNULMAZ contract changes. Tool count 12 -> 13.
from xer_compare import (
    diff_tasks as _xc_diff_tasks,
    diff_links as _xc_diff_links,
    diff_progress as _xc_diff_progress,
    diff_evm as _xc_diff_evm,
    summarize_compare as _xc_summarize,
)


def _msp_compare_tasks(file_path_a=None, file_path_b=None, fields=None):
    """Phase 7 Action 1: task_delta — added/removed/changed tasks."""
    a = _evm_load_task_data(file_path=file_path_a)
    if a.get("status") != "ok":
        return {"status": "error",
                "error": f"file_a load failed: {a.get('error')}"}
    b = _evm_load_task_data(file_path=file_path_b)
    if b.get("status") != "ok":
        return {"status": "error",
                "error": f"file_b load failed: {b.get('error')}"}
    diff = _xc_diff_tasks(a.get("tasks", []), b.get("tasks", []),
                          fields=fields)
    return {"status": "ok", **diff}


def _msp_compare_links(file_path_a=None, file_path_b=None):
    """Phase 7 Action 2: link_delta — added/removed/changed links."""
    links_a = _dcma_load_links(file_path=file_path_a)
    links_b = _dcma_load_links(file_path=file_path_b)
    if not isinstance(links_a, list):
        return {"status": "error",
                "error": f"file_a links unavailable: {links_a}"}
    if not isinstance(links_b, list):
        return {"status": "error",
                "error": f"file_b links unavailable: {links_b}"}
    diff = _xc_diff_links(links_a, links_b)
    return {"status": "ok", **diff}


def _msp_compare_progress(file_path_a=None, file_path_b=None):
    """Phase 7 Action 3: progress_delta — per-task pct + actual_work."""
    a = _evm_load_task_data(file_path=file_path_a)
    if a.get("status") != "ok":
        return {"status": "error",
                "error": f"file_a load failed: {a.get('error')}"}
    b = _evm_load_task_data(file_path=file_path_b)
    if b.get("status") != "ok":
        return {"status": "error",
                "error": f"file_b load failed: {b.get('error')}"}
    progress_a = {"status_date": a.get("status_date"),
                  "tasks": a.get("tasks", [])}
    progress_b = {"status_date": b.get("status_date"),
                  "tasks": b.get("tasks", [])}
    diff = _xc_diff_progress(progress_a, progress_b)
    return {"status": "ok", **diff}


def _msp_compare_evm(file_path_a=None, file_path_b=None, baseline_number=0):
    """Phase 7 Action 4: evm_delta — BAC/PV/EV/AC/SPI/CPI snapshot delta."""
    a = _msp_evm_compute_metrics(file_path=file_path_a,
                                 baseline_number=baseline_number)
    if a.get("status") != "ok":
        return {"status": "error",
                "error": f"file_a EVM failed: {a.get('error')}"}
    b = _msp_evm_compute_metrics(file_path=file_path_b,
                                 baseline_number=baseline_number)
    if b.get("status") != "ok":
        return {"status": "error",
                "error": f"file_b EVM failed: {b.get('error')}"}
    diff = _xc_diff_evm(a, b)
    return {"status": "ok", **diff}


def _msp_compare_summary(file_path_a=None, file_path_b=None,
                         baseline_number=0):
    """Phase 7 Action 5: summary — aggregate all delta types in one call."""
    task_d = _msp_compare_tasks(file_path_a, file_path_b)
    if task_d.get("status") != "ok":
        return task_d
    link_d = _msp_compare_links(file_path_a, file_path_b)
    if link_d.get("status") != "ok":
        return link_d
    progress_d = _msp_compare_progress(file_path_a, file_path_b)
    if progress_d.get("status") != "ok":
        return progress_d
    evm_d = _msp_compare_evm(file_path_a, file_path_b,
                             baseline_number=baseline_number)
    if evm_d.get("status") != "ok":
        return evm_d
    s = _xc_summarize(task_d, link_d, progress_d, evm_d)
    return {"status": "ok", **s}


def _msp_compare_monthly_report(file_path_a=None, file_path_b=None,
                                baseline_number=0, output_excel=None):
    """Phase 8.1 Action 6: monthly_report — bundles compare summary +
    EVM(both files) + optional hakediş Excel export.

    CAU monthly hakediş workflow: last_month.xer + this_month.xer →
    one call returns delta headline, both EVM RAG/SPI/CPI snapshots,
    and optionally writes a hakediş workbook for file_b (current).
    """
    summary = _msp_compare_summary(file_path_a, file_path_b,
                                   baseline_number=baseline_number)
    if summary.get("status") != "ok":
        return summary
    evm_a = _msp_evm_summary(file_path=file_path_a,
                             baseline_number=baseline_number)
    if evm_a.get("status") != "ok":
        return {"status": "error",
                "error": f"file_a EVM summary failed: {evm_a.get('error')}"}
    evm_b = _msp_evm_summary(file_path=file_path_b,
                             baseline_number=baseline_number)
    if evm_b.get("status") != "ok":
        return {"status": "error",
                "error": f"file_b EVM summary failed: {evm_b.get('error')}"}
    excel_result = None
    if output_excel:
        excel_result = _msp_excel_export_hakedis(
            file_path=file_path_b, xlsx_path=output_excel,
            baseline_number=baseline_number)
        if excel_result.get("status") != "ok":
            return {"status": "error",
                    "error": f"Excel export failed: {excel_result.get('error')}"}
    rag_a = evm_a.get("rag", "?")
    rag_b = evm_b.get("rag", "?")
    rag_segment = (f"RAG {rag_a}->{rag_b}" if rag_a != rag_b
                   else f"RAG {rag_a}")
    base_headline = summary.get("headline", "")
    headline = ((base_headline + ", " + rag_segment).strip(", ")
                if base_headline else rag_segment)
    return {
        "status": "ok",
        "headline": headline,
        "compare_summary": summary,
        "evm_a": {
            "rag": evm_a.get("rag"),
            "completion_pct": evm_a.get("completion_pct"),
            "spi": evm_a.get("spi"),
            "cpi": evm_a.get("cpi"),
        },
        "evm_b": {
            "rag": evm_b.get("rag"),
            "completion_pct": evm_b.get("completion_pct"),
            "spi": evm_b.get("spi"),
            "cpi": evm_b.get("cpi"),
        },
        "excel_path": output_excel if excel_result else None,
        "excel_export": excel_result,
    }


@mcp.tool(
    name="msproject_compare",
    annotations={
        "title": "MS Project Snapshot Comparison",
        "readOnlyHint": True,
    },
)
async def msproject_compare(params: dict) -> str:
    """Compare two project file snapshots (XER or MSPDI) for delta analysis.

    Use case: CAU monthly hakediş — last month's baseline XER vs this
    month's progress XER. Reports added/removed/changed tasks, link
    deltas, progress jumps, EVM (BAC/PV/EV/AC/SPI/CPI) deltas.

    Actions:
    - task_delta: added/removed/changed tasks (custom fields=[...] OK)
    - link_delta: added/removed/changed links (FS/FF/SS/SF + lag)
    - progress_delta: per-task percent_complete + actual_work delta
    - evm_delta: snapshot EVM metrics delta
    - summary: aggregate headline + counts + EVM deltas

    Phase 7 (2 May 2026). Tool count 12 -> 13.
    """
    import json
    action = params.get("action", "")
    p = {k: v for k, v in params.items() if k != "action"}
    # Friendly param aliases: file_a/file_b -> file_path_a/file_path_b
    if "file_a" in p:
        p["file_path_a"] = p.pop("file_a")
    if "file_b" in p:
        p["file_path_b"] = p.pop("file_b")
    try:
        if action == "task_delta":
            r = _msp_compare_tasks(**p)
        elif action == "link_delta":
            r = _msp_compare_links(**p)
        elif action == "progress_delta":
            r = _msp_compare_progress(**p)
        elif action == "evm_delta":
            r = _msp_compare_evm(**p)
        elif action == "summary":
            r = _msp_compare_summary(**p)
        elif action == "monthly_report":
            r = _msp_compare_monthly_report(**p)
        else:
            r = {"status": "error",
                 "error": (f"Unknown action '{action}'. Valid: "
                           "task_delta/link_delta/progress_delta/"
                           "evm_delta/summary/monthly_report")}
    except TypeError as e:
        r = {"status": "error", "error": f"Invalid params for {action}: {e}"}
    except Exception as e:
        logger.exception(f"msproject_compare({action}) failed: {e}")
        r = {"status": "error", "error": str(e)}
    return json.dumps(r, default=str, ensure_ascii=False)


def main():
    """Run MCP server (stdio)."""
    mcp.run()


if __name__ == "__main__":
    main()
