"""Phase 5c - Excel I/O for MS Project hakedis workflow.

Pure-Python openpyxl helpers. MSP/COM/file independent - takes plain
dicts, writes openpyxl Workbook objects. MCS brand styling per CLAUDE.md
RULE 14.
"""
from openpyxl.styles import Font, PatternFill, Alignment


# ---------- MCS brand constants (CLAUDE.md RULE 14) ----------

BRAND_LACIVERT = "FF0B1F4D"   # ARGB - main brand
BRAND_NAVY = "FF3D4663"
BRAND_TURQUOISE = "FF39B4CC"
BRAND_LABEL_GRAY = "FF6B7394"
ZEBRA = "FFF0F3F8"

# RAG status colors (industry standard, no competitor names per RULE 13)
RAG_GREEN = "FF8FBC8F"
RAG_AMBER = "FFFFCC66"
RAG_RED = "FFE57373"

WHITE = "FFFFFFFF"


def apply_header_style(ws, row=1):
    """Lacivert background + white bold Calibri 11pt header (CLAUDE.md RULE 14)."""
    fill = PatternFill(start_color=BRAND_LACIVERT, end_color=BRAND_LACIVERT,
                       fill_type="solid")
    font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def apply_rag_fill(cell, status):
    """Apply RAG color fill based on status string (pass/fail/amber/green/red)."""
    s = (status or "").lower()
    if s in ("pass", "green", "ok"):
        color = RAG_GREEN
    elif s in ("fail", "red"):
        color = RAG_RED
    elif s == "amber":
        color = RAG_AMBER
    else:
        return
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def apply_zebra_fill(cell, row_idx):
    """Light gray zebra stripe on even rows (1-indexed). Odd rows untouched."""
    if row_idx % 2 == 0:
        cell.fill = PatternFill(start_color=ZEBRA, end_color=ZEBRA, fill_type="solid")


# ---------- T95: Tasks + Summary sheet builders + readers ----------

from openpyxl import Workbook, load_workbook


TASKS_HEADERS = ["ID", "Name", "Duration (d)", "Start", "Finish",
                 "%Complete", "Critical", "Summary"]


def _real_only(tasks):
    """Filter out summary tasks (real work rows only)."""
    return [t for t in tasks if not t.get("summary", False)]


def _hours_to_days(h):
    """Convert duration in hours to working days (8h/day)."""
    try:
        return round(float(h or 0) / 8.0, 2)
    except (ValueError, TypeError):
        return 0


def _ensure_sheet(wb, sheet_name):
    """Get or create a worksheet by name. Always create_sheet for new names.

    Caller (composer) is responsible for removing the default 'Sheet' if present.
    Don't probe wb.active.cell(...) — that instantiates the cell and breaks
    subsequent append(), causing headers to land on row 2 instead of row 1.
    """
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name)


def build_tasks_sheet(wb, tasks, sheet_name="Tasks"):
    """Build Tasks sheet from list of task dicts. Summary tasks excluded."""
    ws = _ensure_sheet(wb, sheet_name)
    ws.append(TASKS_HEADERS)
    apply_header_style(ws, row=1)
    real = _real_only(tasks)
    for idx, t in enumerate(real, start=2):
        ws.cell(row=idx, column=1, value=t.get("id"))
        ws.cell(row=idx, column=2, value=t.get("name", ""))
        ws.cell(row=idx, column=3, value=_hours_to_days(t.get("duration_h")))
        ws.cell(row=idx, column=4, value=t.get("start"))
        ws.cell(row=idx, column=5, value=t.get("finish"))
        pc = float(t.get("percent_complete") or 0)
        ws.cell(row=idx, column=6, value=pc / 100.0 if pc > 1 else pc)
        ws.cell(row=idx, column=6).number_format = "0%"
        ws.cell(row=idx, column=7, value=bool(t.get("critical", False)))
        ws.cell(row=idx, column=8, value=bool(t.get("summary", False)))
        for col in range(1, len(TASKS_HEADERS) + 1):
            apply_zebra_fill(ws.cell(row=idx, column=col), row_idx=idx)
    ws.freeze_panes = "A2"
    return ws


def read_tasks_sheet(xlsx_path, sheet_name="Tasks"):
    """Read Tasks sheet back into list of dicts (for import workflow)."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    out = []
    for row in rows[1:]:
        d = {}
        for i, val in enumerate(row):
            if i >= len(headers):
                continue
            h = headers[i]
            if h == "ID":
                d["id"] = int(val) if val is not None else None
            elif h == "Name":
                d["name"] = str(val) if val is not None else ""
            elif h == "Duration (d)":
                d["duration_h"] = float(val or 0) * 8.0
            elif h == "Start":
                d["start"] = str(val) if val else None
            elif h == "Finish":
                d["finish"] = str(val) if val else None
            elif h == "%Complete":
                pc = val or 0
                try:
                    pc = float(pc)
                except (ValueError, TypeError):
                    pc = 0
                d["percent_complete"] = pc * 100.0 if 0 < pc <= 1 else pc
            elif h == "Critical":
                d["critical"] = bool(val)
            elif h == "Summary":
                d["summary"] = bool(val)
        if d.get("name"):
            out.append(d)
    return out


def build_summary_sheet(wb, summary):
    """Hakedis executive summary sheet (BAC/EAC/SPI/CPI/RAG + exec text)."""
    ws = _ensure_sheet(wb, "Summary")
    ws.cell(row=1, column=1, value="Project Health Summary")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=16,
                                         color=BRAND_LACIVERT)
    metrics = [
        ("BAC (Budget at Completion)", summary.get("BAC")),
        ("EAC (Estimate at Completion)", summary.get("EAC")),
        ("SPI (Schedule Performance Index)", summary.get("SPI")),
        ("CPI (Cost Performance Index)", summary.get("CPI")),
        ("Overall RAG", str(summary.get("rag", "")).upper()),
    ]
    for i, (label, val) in enumerate(metrics, start=3):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=1).font = Font(name="Calibri", bold=True,
                                             color=BRAND_NAVY)
        ws.cell(row=i, column=2, value=val if val is not None else "N/A")
        if label == "Overall RAG":
            apply_rag_fill(ws.cell(row=i, column=2), str(val or "").lower())
    exec_text = summary.get("executive_text", "")
    if exec_text:
        ws.cell(row=10, column=1, value="Executive Summary")
        ws.cell(row=10, column=1).font = Font(name="Calibri", bold=True,
                                              color=BRAND_LACIVERT)
        ws.cell(row=11, column=1, value=exec_text)
        ws.merge_cells(start_row=11, start_column=1, end_row=11, end_column=4)
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 20
    return ws


# ---------- T96: EVM sheet builder (Compute + TimePhased) ----------

EVM_COMPUTE_ROWS = [
    ("BAC", "metrics", "BAC", "currency"),
    ("EV (Earned Value)", "metrics", "EV", "currency"),
    ("AC (Actual Cost)", "metrics", "AC", "currency"),
    ("PV (Planned Value)", "metrics", "PV", "currency"),
    ("SV (Schedule Variance)", "metrics", "SV", "currency"),
    ("CV (Cost Variance)", "metrics", "CV", "currency"),
    ("SPI", "metrics", "SPI", "ratio"),
    ("CPI", "metrics", "CPI", "ratio"),
    ("EAC1", "forecast", "EAC1", "currency"),
    ("EAC2", "forecast", "EAC2", "currency"),
    ("EAC3", "forecast", "EAC3", "currency"),
    ("ETC", "forecast", "ETC", "currency"),
    ("VAC", "forecast", "VAC", "currency"),
    ("TCPI(BAC)", "forecast", "TCPI_BAC", "ratio"),
    ("TCPI(EAC)", "forecast", "TCPI_EAC", "ratio"),
    ("AT (Actual Time, weeks)", "earned_schedule", "AT", "weeks"),
    ("ES (Earned Schedule, weeks)", "earned_schedule", "ES", "weeks"),
    ("SV(t) (weeks)", "earned_schedule", "SVt", "weeks"),
    ("SPI(t)", "earned_schedule", "SPIt", "ratio"),
]


def build_evm_sheet(wb, evm):
    """Build EVM_Compute + EVM_TimePhased sheets from EVM data dict.

    evm = {metrics, forecast?, earned_schedule?, rag, time_phased[]}
    """
    # Compute sheet
    ws = _ensure_sheet(wb, "EVM_Compute")
    ws.append(["Metric", "Value", "Unit"])
    apply_header_style(ws, row=1)
    row = 2
    for label, src, key, unit in EVM_COMPUTE_ROWS:
        section = evm.get(src) if isinstance(evm.get(src), dict) else {}
        val = section.get(key)
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val if val is not None else "N/A")
        if unit == "currency" and isinstance(val, (int, float)):
            ws.cell(row=row, column=2).number_format = "#,##0.00"
        elif unit == "ratio" and isinstance(val, (int, float)):
            ws.cell(row=row, column=2).number_format = "0.000"
        ws.cell(row=row, column=3, value=unit)
        for col in range(1, 4):
            apply_zebra_fill(ws.cell(row=row, column=col), row_idx=row)
        row += 1
    # RAG row
    ws.cell(row=row, column=1, value="Overall RAG")
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True,
                                           color=BRAND_NAVY)
    rag = str(evm.get("rag", "")).lower()
    ws.cell(row=row, column=2, value=rag.upper() if rag else "N/A")
    apply_rag_fill(ws.cell(row=row, column=2), rag)
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.freeze_panes = "A2"

    # Time-phased sheet
    ws_tp = _ensure_sheet(wb, "EVM_TimePhased")
    ws_tp.append(["Period", "PV", "EV", "AC", "Cum PV", "Cum EV", "Cum AC"])
    apply_header_style(ws_tp, row=1)
    for idx, p in enumerate(evm.get("time_phased") or [], start=2):
        ws_tp.cell(row=idx, column=1, value=p.get("period"))
        for col, key in enumerate(["PV", "EV", "AC", "cum_PV", "cum_EV", "cum_AC"],
                                  start=2):
            ws_tp.cell(row=idx, column=col, value=p.get(key, 0))
            ws_tp.cell(row=idx, column=col).number_format = "#,##0.00"
        for col in range(1, 8):
            apply_zebra_fill(ws_tp.cell(row=idx, column=col), row_idx=idx)
    ws_tp.freeze_panes = "B2"
    return ws, ws_tp


# ---------- T97: DCMA sheet builder (Rules + Failed) ----------

def build_dcma_sheet(wb, dcma):
    """Build DCMA_Rules + DCMA_Failed sheets.

    dcma = {rules: [...], summary: {...}, drilldowns: {rule_id: [{id, name}]}}
    """
    # Rules sheet
    ws = _ensure_sheet(wb, "DCMA_Rules")
    ws.append(["Rule #", "Name", "Threshold", "Actual", "Status",
               "Failed Count", "Total"])
    apply_header_style(ws, row=1)
    for idx, rule in enumerate(dcma.get("rules") or [], start=2):
        ws.cell(row=idx, column=1, value=rule.get("id"))
        ws.cell(row=idx, column=2, value=rule.get("name"))
        ws.cell(row=idx, column=3, value=rule.get("threshold"))
        actual = rule.get("actual")
        unit = rule.get("actual_unit", "")
        ws.cell(row=idx, column=4,
                value=f"{actual}{unit}" if actual is not None else "N/A")
        status = rule.get("status", "")
        ws.cell(row=idx, column=5, value=status.upper())
        ws.cell(row=idx, column=6, value=rule.get("failed_count", 0))
        ws.cell(row=idx, column=7, value=rule.get("total_count", 0))
        # Zebra all cells first; then RAG overrides status column color
        for col in range(1, 8):
            if col == 5:
                continue  # status col gets RAG fill, not zebra
            apply_zebra_fill(ws.cell(row=idx, column=col), row_idx=idx)
        apply_rag_fill(ws.cell(row=idx, column=5), status)
    for col, w in zip("ABCDEFG", [8, 28, 12, 14, 10, 14, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # Failed drilldown sheet
    ws_f = _ensure_sheet(wb, "DCMA_Failed")
    ws_f.append(["Rule #", "Rule Name", "Task ID", "Task Name"])
    apply_header_style(ws_f, row=1)
    drills = dcma.get("drilldowns") or {}
    rules_by_id = {r["id"]: r for r in (dcma.get("rules") or [])}
    row = 2
    for rid, tasks in drills.items():
        rule_name = (rules_by_id.get(rid) or {}).get("name", "")
        for t in (tasks or [])[:10]:
            ws_f.cell(row=row, column=1, value=rid)
            ws_f.cell(row=row, column=2, value=rule_name)
            ws_f.cell(row=row, column=3, value=t.get("id"))
            ws_f.cell(row=row, column=4, value=t.get("name", ""))
            for col in range(1, 5):
                apply_zebra_fill(ws_f.cell(row=row, column=col), row_idx=row)
            row += 1
    for col, w in zip("ABCD", [8, 28, 10, 28]):
        ws_f.column_dimensions[col].width = w
    ws_f.freeze_panes = "A2"
    return ws, ws_f


# ---------- T98: Hakedis workbook composer + progress reader ----------

def build_hakedis_workbook(tasks, evm, dcma, summary, xlsx_path):
    """Compose multi-sheet hakedis workbook and save to xlsx_path.

    Sheets in display order: Summary, Tasks, EVM_Compute, EVM_TimePhased,
    DCMA_Rules, DCMA_Failed. Default empty 'Sheet' removed.
    """
    wb = Workbook()
    # Remove default empty sheet (composer is responsible per _ensure_sheet
    # convention).
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    build_summary_sheet(wb, summary or {})
    build_tasks_sheet(wb, tasks or [], sheet_name="Tasks")
    build_evm_sheet(wb, evm or {})
    build_dcma_sheet(wb, dcma or {})
    wb.save(xlsx_path)
    return wb


def read_progress_sheet(xlsx_path, sheet_name="Progress"):
    """Read Progress sheet -> [{task_id, percent_complete, actual_work_h?}].

    %Complete cells: accepts 50, 50.0, or 0.5 (treats values 0<v<=1 as fraction).
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    out = []
    for row in rows[1:]:
        d = {}
        for i, val in enumerate(row):
            if i >= len(headers):
                continue
            h = headers[i]
            if h == "Task ID":
                d["task_id"] = int(val) if val is not None else None
            elif h == "%Complete":
                try:
                    pc = float(val or 0)
                except (ValueError, TypeError):
                    pc = 0
                d["percent_complete"] = pc * 100.0 if 0 < pc <= 1 else pc
            elif h == "Actual Work (h)":
                try:
                    d["actual_work_h"] = float(val or 0)
                except (ValueError, TypeError):
                    d["actual_work_h"] = 0
        if d.get("task_id") is not None:
            out.append(d)
    return out
